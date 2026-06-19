import pandas as pd

from phenex.reporting.reporter import Reporter
from phenex.util import create_logger
from ibis import _

logger = create_logger(__name__)


class _ComponentPhenotypeView:
    """
    Wraps a phenotype to override its ``display_name`` for indented display inside
    Table1 when ``include_component_phenotypes_level`` is set.  All other
    attributes are delegated to the underlying phenotype unchanged.
    """

    def __init__(self, phenotype, display_name: str, level: int = 0):
        # Store directly in __dict__ to avoid triggering __getattr__
        object.__setattr__(self, "_phenotype", phenotype)
        object.__setattr__(self, "_display_name", display_name)
        object.__setattr__(self, "_component_level", level)

    @property
    def display_name(self) -> str:
        return object.__getattribute__(self, "_display_name")

    @property
    def _level(self) -> int:
        return object.__getattribute__(self, "_component_level")

    def __getattr__(self, name: str):
        return getattr(object.__getattribute__(self, "_phenotype"), name)


class Table1(Reporter):
    """
    Table1 is a common term used in epidemiology to describe a table that shows an overview of the baseline characteristics of a cohort. It contains the counts and percentages of the cohort that have each characteristic, for both boolean and value characteristics. In addition, summary statistics are provided for value characteristics (mean, std, median, min, max).

    Table1 by default reports on all phenotypes in the cohort's characteristics, but a custom list of phenotypes can be provided to the execute() method. When using the default cohort.characteristics, the section structure defined on the cohort is preserved in the Table1 output for better organization and display.

    Parameters:
        decimal_places: Number of decimal places to round to. Default: 1
        include_component_phenotypes_level: When set to an integer, component
            (child) phenotypes are expanded inline beneath each parent phenotype,
            indented according to their nesting depth.  ``None`` (default) disables
            expansion.  Set to a large number (e.g. 100) to include all levels.
    """

    def __init__(self, include_component_phenotypes_level=None, **kwargs):
        super().__init__(**kwargs)
        self.include_component_phenotypes_level = include_component_phenotypes_level
        self.characteristic_sections = None

    # ------------------------------------------------------------------
    # Component-phenotype expansion helpers
    # ------------------------------------------------------------------

    def _expand_with_components(self, phenotypes):
        """Return a flat, ordered list that interleaves component phenotypes.

        Each top-level phenotype is followed (depth-first) by its children up
        to *include_component_phenotypes_level* levels deep.  Component phenotypes
        are wrapped in a :class:`_ComponentPhenotypeView` whose ``display_name``
        is prefixed with two spaces per nesting level so the hierarchy is visible
        in the final table.
        """
        result = []
        for p in phenotypes:
            result.append(p)
            self._collect_components(p, result, level=1)
        return result

    def _collect_components(self, phenotype, result, level):
        if level > self.include_component_phenotypes_level:
            return
        children = getattr(phenotype, "children", None) or []
        for child in children:
            indent = "\u00a0\u00a0" * level  # non-breaking spaces for visual indent
            view = _ComponentPhenotypeView(
                child, f"{indent}{child.display_name}", level=level
            )
            result.append(view)
            self._collect_components(child, result, level + 1)

    def execute(
        self, cohort: "Cohort", phenotypes: "Optional[Union[List, Dict]]" = None
    ) -> pd.DataFrame:
        self.cohort = cohort

        if phenotypes is None:
            self._phenotypes = cohort.characteristics
            self.characteristic_sections = getattr(
                cohort, "characteristic_sections", None
            )
        elif isinstance(phenotypes, dict):
            self.characteristic_sections = {
                section: [p.display_name for p in phenos]
                for section, phenos in phenotypes.items()
            }
            self._phenotypes = [p for phenos in phenotypes.values() for p in phenos]
        else:
            self._phenotypes = phenotypes
            self.characteristic_sections = None

        if len(self._phenotypes) == 0:
            logger.info("No phenotypes. table1 is empty")
            return pd.DataFrame()

        # Optionally expand each phenotype with its component children
        if self.include_component_phenotypes_level is not None:
            self._phenotypes = self._expand_with_components(self._phenotypes)

        self.cohort_names_in_order = [x.name for x in self._phenotypes]
        self.N = (
            cohort.index_table.filter(cohort.index_table.BOOLEAN == True)
            .select("PERSON_ID")
            .distinct()
            .count()
            .execute()
        )
        logger.debug("Starting with categorical columns for table1")
        self.df_categoricals = self._report_categorical_columns()
        logger.debug("Starting with boolean columns for table1")
        self.df_booleans = self._report_boolean_columns()
        logger.debug("Starting with value columns for table1")
        self.df_values = self._report_value_columns()
        logger.debug("Collecting value distributions for histogram visualization")
        self._value_distributions = self._collect_value_distributions()

        # add the full cohort size as the first row
        df_n = pd.DataFrame(
            {"N": [self.N], "inex_order": [-1], "_level": [0]}, index=["Cohort"]
        )
        # add percentage column
        dfs = [
            df
            for df in [df_n, self.df_booleans, self.df_values, self.df_categoricals]
            if df is not None
        ]
        if len(dfs) > 1:
            self.df = pd.concat(dfs)
        elif len(dfs) == 1:
            self.df = dfs[0]
        else:
            self.df = None
        if self.df is not None:
            self.df["Pct"] = 100 * self.df["N"] / self.N
            # reorder columns so N and Pct are first
            first_cols = ["N", "Pct"]
            column_order = first_cols + [
                x for x in self.df.columns if x not in first_cols
            ]
            self.df = self.df[column_order]
        logger.debug("Finished creating table1")

        self.df = self.df.reset_index()
        self.df.columns = ["Name"] + list(self.df.columns[1:])

        self.df = self.df.sort_values(by=["inex_order", "Name"])
        self.df = self.df.reset_index()[
            [x for x in self.df.columns if x not in ["index", "inex_order"]]
        ]
        # Strip the "NNNN_" sort-order prefix that BinPhenotype embeds in bin
        # labels (e.g. "Age group=0003_[30-40)" → "Age group=[30-40)").
        self.df["Name"] = self.df["Name"].str.replace(r"=\d{4}_", "=", regex=True)
        return self.df

    def _get_boolean_characteristics(self):
        return [x for x in self._phenotypes if x.output_display_type == "boolean"]

    def _get_value_characteristics(self):
        return [x for x in self._phenotypes if x.output_display_type == "value"]

    def _get_categorical_characteristics(self):
        return [x for x in self._phenotypes if x.output_display_type == "categorical"]

    def _get_boolean_count_for_phenotype(self, phenotype):
        result = (
            phenotype.table.select(["PERSON_ID", "BOOLEAN"])
            .distinct()["BOOLEAN"]
            .sum()
            .execute()
        )
        # Return 0 if result is None or NaN (no rows with BOOLEAN=True)
        return (
            0
            if result is None or (isinstance(result, float) and pd.isna(result))
            else int(result)
        )

    def _report_boolean_columns(self):
        # get list of all boolean columns
        boolean_phenotypes = self._get_boolean_characteristics()
        logger.debug(
            f"Found {len(boolean_phenotypes)} : {[x.name for x in boolean_phenotypes]}"
        )
        if len(boolean_phenotypes) == 0:
            return None
        # get count of 'Trues' in the boolean columns i.e. the phenotype counts
        df_t1 = pd.DataFrame()
        df_t1["N"] = [
            self._get_boolean_count_for_phenotype(phenotype)
            for phenotype in boolean_phenotypes
        ]
        df_t1.index = [x.display_name for x in boolean_phenotypes]
        df_t1["inex_order"] = [
            self.cohort_names_in_order.index(x.name) for x in boolean_phenotypes
        ]
        df_t1["_level"] = [getattr(x, "_level", 0) for x in boolean_phenotypes]
        return df_t1

    def _report_value_columns(self):
        value_phenotypes = self._get_value_characteristics()
        logger.debug(
            f"Found {len(value_phenotypes)} : {[x.name for x in value_phenotypes]}"
        )

        if len(value_phenotypes) == 0:
            return None

        names = []
        dfs = []
        for phenotype in value_phenotypes:
            _table = phenotype.table.select(["PERSON_ID", "VALUE"]).distinct()
            # Cast VALUE to float to avoid integer-overflow in variance/std
            # computations on fixed-precision backends (e.g. Snowflake computes
            # SUM(VALUE^2), which overflows NUMBER(38,0) for large values).
            _value = _table["VALUE"].cast("float64")
            d = {
                "N": self._get_boolean_count_for_phenotype(phenotype),
                "Mean": _value.mean().execute(),
                "STD": _value.std().execute(),
                "Min": _value.min().execute(),
                "P10": _value.quantile(0.10).execute(),
                "P25": _value.quantile(0.25).execute(),
                "Median": _value.median().execute(),
                "P75": _value.quantile(0.75).execute(),
                "P90": _value.quantile(0.90).execute(),
                "Max": _value.max().execute(),
                "inex_order": self.cohort_names_in_order.index(phenotype.name),
                "_level": getattr(phenotype, "_level", 0),
            }
            dfs.append(pd.DataFrame.from_dict([d]))
            names.append(phenotype.display_name)
        if len(dfs) == 1:
            df = dfs[0]
        else:
            df = pd.concat(dfs)
        df.index = names
        return df

    def _report_categorical_columns(self):
        categorical_phenotypes = self._get_categorical_characteristics()
        logger.debug(
            f"Found {len(categorical_phenotypes)} : {[x.name for x in categorical_phenotypes]}"
        )
        if len(categorical_phenotypes) == 0:
            return None
        dfs = []
        names = []
        for phenotype in categorical_phenotypes:
            name = phenotype.display_name
            _table = phenotype.table.select(["PERSON_ID", "VALUE"])
            # Get counts for each category.
            # Keep the raw VALUE (which may carry a "NNNN_" sort prefix from
            # BinPhenotype) in the index so that the final sort_values("Name")
            # in execute() orders bins correctly. The prefix is stripped there.
            cat_counts = (
                _table.distinct().group_by("VALUE").aggregate(N=_.count()).execute()
            )
            cat_counts.index = [
                f"{name}={v if v is not None else 'None'}" for v in cat_counts["VALUE"]
            ]
            _df = pd.DataFrame(cat_counts["N"])
            _df["inex_order"] = self.cohort_names_in_order.index(phenotype.name)
            _df["_level"] = getattr(phenotype, "_level", 0)
            dfs.append(_df)
            names.extend(cat_counts.index)
        if len(dfs) == 1:
            df = dfs[0]
        else:
            df = pd.concat(dfs)
        df.index = names
        return df

    def _collect_value_distributions(self):
        """Compute KDE curves for numeric phenotypes.

        Stores ``{"x": [...], "y": [...]}`` per phenotype where *y* is
        normalised so the peak equals 100.  This is far more compact than
        raw patient-level values and avoids binning decisions at display time.
        """
        import numpy as np
        from scipy.stats import gaussian_kde

        N_POINTS = 200
        PADDING = 0.10  # 10% range padding on each side

        value_phenotypes = self._get_value_characteristics()
        distributions = {}
        for phenotype in value_phenotypes:
            try:
                values = np.array(
                    phenotype.table.select(["PERSON_ID", "VALUE"])
                    .distinct()["VALUE"]
                    .execute()
                    .dropna()
                    .tolist(),
                    dtype=float,
                )
                if len(values) < 2:
                    continue
                # For integer-valued data, widen the bandwidth to avoid
                # spiky peaks at each integer and produce smooth plateaus.
                is_integer = np.allclose(values, np.round(values))
                bw = 1.5 if is_integer else None
                kde = gaussian_kde(values, bw_method=bw)
                lo, hi = float(values.min()), float(values.max())
                pad = (hi - lo) * PADDING if hi > lo else 1.0
                x = np.linspace(lo - pad, hi + pad, N_POINTS)
                y = kde(x)
                y = y / y.max() * 100  # normalise peak to 100
                distributions[phenotype.display_name] = {
                    "x": np.round(x, 4).tolist(),
                    "y": np.round(y, 2).tolist(),
                }
            except Exception:
                pass
        return distributions

    def get_pretty_display(self) -> pd.DataFrame:
        """
        Return a formatted version of the Table1 results for display.

        Formats numeric columns and converts counts to strings to avoid NaN display.

        Returns:
            pd.DataFrame: Formatted copy of the results
        """
        # Create a copy to avoid modifying the original
        pretty_df = self.df.copy()

        # Drop the internal _level column from display output
        pretty_df = pretty_df.drop(columns=["_level"], errors="ignore")

        # cast counts to integer and to str, so that we can display without 'NaNs'
        pretty_df["N"] = pretty_df["N"].astype("Int64").astype(str)

        pretty_df = pretty_df.round(self.decimal_places)

        to_prettify = [
            "Pct",
            "Mean",
            "STD",
            "Min",
            "P10",
            "P25",
            "Median",
            "P75",
            "P90",
            "Max",
        ]
        for column in to_prettify:
            if column in pretty_df.columns:
                pretty_df[column] = pretty_df[column].astype(str)

        pretty_df = pretty_df.replace("<NA>", "").replace("nan", "")

        return pretty_df

    def to_json(self, filename: str) -> str:
        """Export Table1 to JSON, including section metadata when available."""
        import json
        from pathlib import Path

        if not hasattr(self, "df"):
            raise AttributeError("Call execute() first before calling to_json().")

        filepath = Path(filename)
        if filepath.suffix != ".json":
            filepath = filepath.with_suffix(".json")
        filepath.parent.mkdir(parents=True, exist_ok=True)

        payload = {
            "reporter_type": self.__class__.__name__,
            "rows": self.df.to_dict(orient="records"),
        }
        if self.characteristic_sections:
            payload["sections"] = self.characteristic_sections

        if hasattr(self, "_value_distributions") and self._value_distributions:
            payload["kdes"] = self._value_distributions

        with filepath.open("w") as f:
            json.dump(payload, f, indent=2, default=str)

        if hasattr(self, "_value_distributions") and self._value_distributions:
            dist_path = filepath.with_stem(f"{filepath.stem}_value_distributions")
            with dist_path.open("w") as f:
                json.dump(self._value_distributions, f, indent=2, default=str)

        return str(filepath.absolute())

    def to_excel(self, filename: str) -> str:
        """Export Table1 to Excel, applying progressive gray fills for component rows."""
        import openpyxl
        from openpyxl.styles import PatternFill
        from pathlib import Path

        if not hasattr(self, "df"):
            raise AttributeError("Call execute() first before calling to_excel().")

        filepath = Path(filename)
        if filepath.suffix != ".xlsx":
            filepath = filepath.with_suffix(".xlsx")
        filepath.parent.mkdir(parents=True, exist_ok=True)

        # Write pretty display (strips _level)
        pretty_df = self.get_pretty_display()
        pretty_df.to_excel(filepath, index=False)

        # Apply gray fills based on _level if present
        if "_level" in self.df.columns:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            # header row is row 1; data starts at row 2
            for row_idx, level in enumerate(
                self.df["_level"].fillna(0).astype(int).values, start=2
            ):
                hex_color = self._level_to_gray_hex(level)
                if hex_color:
                    fill = PatternFill(
                        start_color=hex_color, end_color=hex_color, fill_type="solid"
                    )
                    for cell in ws[row_idx]:
                        cell.fill = fill
            wb.save(filepath)

        return str(filepath.absolute())

    @staticmethod
    def _level_to_gray_hex(level: int) -> str:
        """Return a 6-char hex fill color for a component nesting level (empty = no fill)."""
        if level <= 0:
            return ""
        value = max(235 - 20 * (level - 1), 100)
        return f"{value:02X}{value:02X}{value:02X}"
