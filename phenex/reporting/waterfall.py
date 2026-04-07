import pandas as pd
import numpy as np

from .reporter import Reporter
from phenex.util import create_logger

logger = create_logger(__name__)


class Waterfall(Reporter):
    """
    A waterfall diagram, also known as an attrition table, shows how inclusion/exclusion criteria contribute to a final population size. Each inclusion/exclusion criteria is a row in the table, and the number of patients remaining after applying that criteria are shown on that row.

    | Column name | Description |
    | --- | --- |
    | Type | The type of the phenotype, either entry, inclusion or exclusion |
    | Name | The name of entry, inclusion or exclusion criteria |
    | N | The absolute number of patients that fulfill that phenotype. For the entry criterium this is the absolute number in the dataset. For inclusion/exclusion criteria this is the number of patients that fulfill the entry criterium AND the phenotype and that row. |
    | Remaining | The number of patients remaining in the cohort after sequentially applying the inclusion/exclusion criteria in the order that they are listed in this table. |
    | % | The percentage of patients who fulfill the entry criterion who are remaining in the cohort after application of the phenotype on that row |
    | Delta | The change in number of patients that occurs by applying the phenotype on that row. |

    """

    def __init__(
        self,
        decimal_places: int = 1,
        include_component_phenotypes_level=None,
    ):
        super().__init__(decimal_places=decimal_places)
        self.include_component_phenotypes_level = include_component_phenotypes_level

    def execute(self, cohort: "Cohort") -> pd.DataFrame:
        self.cohort = cohort
        logger.debug(f"Beginning execution of waterfall. Calculating N patents")
        N = (
            cohort.index_table.filter(cohort.index_table.BOOLEAN == True)
            .select("PERSON_ID")
            .distinct()
            .count()
            .execute()
        )
        logger.debug(f"Cohort has {N} patients")
        # create info dictionaries for each phenotype containing counts
        self.ds = []
        table = cohort.entry_criterion.table
        N_entry = table.select("PERSON_ID").distinct().count().execute()
        index = 1
        self.ds.append(
            {
                "Type": "entry",
                "Level": 0,
                "Index": str(index),
                "Name": (cohort.entry_criterion.display_name),
                "N": N_entry,
                "Remaining": N_entry,
            }
        )

        if self.include_component_phenotypes_level is not None:
            self._append_components_recursively(
                cohort.entry_criterion, table, parent_index=str(index)
            )

        for inclusion in cohort.inclusions:
            index += 1
            table = self.append_phenotype_to_waterfall(
                table, inclusion, "inclusion", level=0, index=index
            )
            if self.include_component_phenotypes_level is not None:
                self._append_components_recursively(
                    inclusion, table, parent_index=str(index)
                )

        for exclusion in cohort.exclusions:
            index += 1
            table = self.append_phenotype_to_waterfall(
                table, exclusion, "exclusion", level=0, index=index
            )
            if self.include_component_phenotypes_level is not None:
                self._append_components_recursively(
                    exclusion, table, parent_index=str(index)
                )

        # Calculate deltas before adding first/last rows
        self.ds = self.append_delta(self.ds)
        # Entry row delta: remaining after entry vs N persons in database
        self.ds[0]["Delta"] = N_entry - cohort.n_persons_in_source_database

        # create dataframe with phenotype counts (without first/last rows)
        self.df = pd.DataFrame(self.ds)

        # calculate percentage of entry criterion
        self.df["Pct_Remaining"] = self.df["Remaining"] / N_entry * 100
        self.df["Pct_N"] = self.df["N"] / N_entry * 100

        # Calculate Pct Source Database column before rounding
        # Entry row gets a percentage, middle rows get NaN, last row will be added after concat
        entry_pct = N_entry / cohort.n_persons_in_source_database * 100

        # Round numeric (float) columns; preserve integer columns like Level
        float_cols = self.df.select_dtypes(include="float").columns
        self.df[float_cols] = self.df[float_cols].round(self.decimal_places)

        # first row data
        first_row_data = {
            "Type": "info",
            "Name": "N persons in database",
            "N": cohort.n_persons_in_source_database,
            "Remaining": cohort.n_persons_in_source_database,
            "Level": 0,
            "Index": "",
        }

        # last rows via concatenation (they won't have percentages calculated)
        last_row_data = {
            "Type": "info",
            "Name": "Final Cohort Size",
            "Remaining": N,
            "Pct_Remaining": round(100 * N / N_entry, self.decimal_places),
            "Level": 0,
            "Index": "",
        }

        # Concatenate: first row + main dataframe + last row
        first_row_df = pd.DataFrame([first_row_data])
        last_row_df = pd.DataFrame([last_row_data])
        self.df = pd.concat([first_row_df, self.df, last_row_df], ignore_index=True)

        entry_pct = round(
            N_entry / cohort.n_persons_in_source_database * 100, self.decimal_places
        )
        final_pct = round(
            N / cohort.n_persons_in_source_database * 100, self.decimal_places
        )

        self.df["Pct_Source_Database"] = (
            [np.nan, entry_pct] + [np.nan] * (self.df.shape[0] - 3) + [final_pct]
        )

        # Do final column selection (never include _color in raw df)
        columns_to_select = [
            "Type",
            "Index",
            "Name",
            "Level",
            "N",
            "Pct_N",
            "Remaining",
            "Pct_Remaining",
            "Delta",
            "Pct_Source_Database",
        ]

        self.df = self.df[columns_to_select]

        return self.df

    def to_json(self, filename: str) -> str:
        """Serialise waterfall results to JSON.

        Compared to the base implementation:
        - Drops the internal ``Level`` column (used only for visual indentation)
        - Stores N / Remaining / Delta as integers (not floats)
        - Serialises NaN cells as JSON ``null``
        """
        import json
        import math
        from pathlib import Path

        df = self.df.drop(columns=["Level"], errors="ignore").copy()

        COUNT_COLS = {"N", "Remaining", "Delta"}

        records = []
        for row in df.to_dict(orient="records"):
            clean = {}
            for k, v in row.items():
                if isinstance(v, float) and math.isnan(v):
                    clean[k] = None
                elif k in COUNT_COLS and isinstance(v, float):
                    clean[k] = int(v)
                else:
                    clean[k] = v
            records.append(clean)

        filepath = Path(filename)
        if filepath.suffix != ".json":
            filepath = filepath.with_suffix(".json")
        filepath.parent.mkdir(parents=True, exist_ok=True)

        payload = {
            "reporter_type": self.__class__.__name__,
            "rows": records,
        }

        with filepath.open("w") as f:
            json.dump(payload, f, indent=2, default=str)

        return str(filepath.absolute())

    def _append_components_recursively(
        self, current_phenotype, table, level=1, parent_index=""
    ):
        if level <= self.include_component_phenotypes_level:
            for i, child in enumerate(current_phenotype.children):
                current_index = f"{parent_index}.{i+1}"
                current_name = child.display_name
                self.append_phenotype_to_waterfall(
                    table,
                    child,
                    "component",
                    full_name=current_name,
                    index=current_index,
                    level=level,
                )
                self._append_components_recursively(
                    child, table, level + 1, parent_index=current_index
                )

    def append_phenotype_to_waterfall(
        self, table, phenotype, type, level, index=None, full_name=None
    ):
        if type == "inclusion":
            table = table.inner_join(
                phenotype.table, table["PERSON_ID"] == phenotype.table["PERSON_ID"]
            )
        elif type == "exclusion":
            table = table.filter(~table["PERSON_ID"].isin(phenotype.table["PERSON_ID"]))
        elif type == "component":
            table = table
        else:
            raise ValueError("type must be either inclusion or exclusion")
        logger.debug(f"Starting {type} criteria {phenotype.name}")

        if full_name is None:
            full_name = phenotype.display_name

        self.ds.append(
            {
                "Type": type,
                "Name": full_name,
                "Level": level,
                "Index": index if index is not None else str(level),
                "N": phenotype.table.select("PERSON_ID").distinct().count().execute(),
                "Remaining": (
                    table.select("PERSON_ID").distinct().count().execute()
                    if type != "component"
                    else np.nan
                ),
            }
        )
        logger.debug(
            f"Finished {type} criteria {phenotype.name}: N = {self.ds[-1]['N']} waterfall = {self.ds[-1]['Remaining']}"
        )
        return table.select("PERSON_ID")

    def get_pretty_display(self) -> pd.DataFrame:
        """
        Return a formatted version of the waterfall results for display.

        Formatting includes:
        - Formatting numeric columns as strings with thousand separators
        - Replacing NAs with empty strings
        - Creating sparse type column

        Returns:
            pd.DataFrame: Formatted copy of the results without color styling
        """
        return self.get_pretty_display(color=False)

    def get_pretty_display(self, color: bool = True) -> pd.DataFrame:
        """
        Return a formatted version of the waterfall results for display.

        Formatting includes:
        - Formatting numeric columns as strings with thousand separators
        - Replacing NAs with empty strings
        - Creating sparse type column
        - Optionally adding row colors based on type and level

        Args:
            color: If True, add color styling to rows. Default: True

        Returns:
            pd.DataFrame: Formatted copy of the results
        """
        # Create a copy to avoid modifying the original
        pretty_df = self.df.copy()

        # Temporarily swap self.df so helper methods work
        original_df = self.df
        self.df = pretty_df

        try:
            # Add colors before any transformations (if requested)
            if color:
                self._add_row_colors()

            # Drop Level column now that coloring is done
            self.df = self.df.drop(columns=["Level"], errors="ignore")

            # Format numeric columns as strings
            self._format_numeric_columns()

            # Replace NAs and None values with empty strings
            self.df = self.df.replace("<NA>", "")

            # Create sparse type column (show type only once per section)
            self._create_sparse_type_column()

            result = self.df
        finally:
            # Restore original df
            self.df = original_df

        return result

    def _add_row_colors(self):
        """Add HSL colors to each row based on type and level"""
        color_map = self._get_color_map()

        self.df["_color"] = None
        last_parent_color = None

        for idx, row in self.df.iterrows():
            row_type = str(row["Type"]) if row["Type"] != "" else "component"
            # Derive level from the Index column (always present): the number of
            # dots encodes nesting depth, e.g. "2" → 0, "2.1" → 1, "2.1.3" → 2.
            # This avoids relying on the Level column which is not saved to Excel.
            index_val = (
                str(row.get("Index", "")) if pd.notna(row.get("Index", None)) else ""
            )
            level = index_val.count(".")

            # Determine base color and apply lightness for component depth
            if row_type == "component":
                # Inherit parent color and lighten based on nesting level
                base_color = last_parent_color
                adjusted_color = self._adjust_brightness(base_color, level)
            else:
                # Non-component rows: use base color (level is always 0 here)
                base_color = color_map.get(row_type, (0, 0, 100))
                last_parent_color = base_color
                adjusted_color = base_color

            self.df.at[idx, "_color"] = self._hsl_to_string(adjusted_color)

    def _format_numeric_columns(self):
        """Convert numeric columns to formatted strings with thousand separators"""
        # Format integer columns with commas
        self.df["N"] = self.df["N"].apply(
            lambda x: f"{int(x):,}" if pd.notna(x) else ""
        )
        self.df["Delta"] = self.df["Delta"].apply(
            lambda x: f"{int(x):,}" if pd.notna(x) else ""
        )
        self.df["Remaining"] = self.df["Remaining"].apply(
            lambda x: f"{int(x):,}" if pd.notna(x) else ""
        )

        # Format percentage columns without commas (they won't need them)
        self.df["Pct_Remaining"] = (
            self.df["Pct_Remaining"].astype("Float64").astype(str)
        )
        self.df["Pct_N"] = self.df["Pct_N"].astype("Float64").astype(str)
        self.df["Pct_Source_Database"] = (
            self.df["Pct_Source_Database"].astype("Float64").astype(str)
        )

    def _apply_styling(self):
        """Apply background colors to dataframe rows"""

        def apply_row_color(row):
            color = row.get("_color")
            if color and pd.notna(color):
                return [f"background-color: {color}" for _ in row]
            return ["" for _ in row]

        # Create styled dataframe (keeping _color column for now)
        styled_df = self.df.style.apply(apply_row_color, axis=1)

        # Hide the _color column in display
        styled_df = styled_df.hide(subset=["_color"], axis="columns")

        return styled_df

    def to_excel(self, filename: str, sheet_name: str = "Waterfall") -> str:
        """
        Export waterfall report to Excel with color styling.
        All cells are formatted as text to prevent Excel auto-formatting.

        Args:
            filename: Path to the output file (relative or absolute, with or without .xlsx extension)
            sheet_name: Name of the Excel sheet (default: 'Waterfall')

        Returns:
            str: Full path to the created file
        """
        from pathlib import Path

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        # Convert to Path object and ensure .xlsx extension
        filename = Path(filename)
        if not filename.suffix == ".xlsx":
            filename = filename.with_suffix(".xlsx")

        # Create parent directories if needed
        filename.parent.mkdir(parents=True, exist_ok=True)

        # Get formatted dataframe with colors
        colored_df = self.get_pretty_display(color=True)

        # Extract colors if they exist
        if "_color" in colored_df.columns:
            export_df = colored_df.drop(columns=["_color"])
            colors = colored_df["_color"].tolist()
        else:
            export_df = colored_df
            colors = [None] * len(colored_df)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers
        headers = list(export_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Style header
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Write data rows and apply styling
        for row_idx, (df_idx, row_data) in enumerate(export_df.iterrows(), start=2):
            color = colors[df_idx] if df_idx < len(colors) else None

            # Get fill pattern for this row
            fill = None
            if color and pd.notna(color):
                hex_color = self._hsl_to_hex(color)
                if hex_color:
                    fill = PatternFill(
                        start_color=hex_color, end_color=hex_color, fill_type="solid"
                    )

            # Write each cell as text
            for col_idx, value in enumerate(row_data, start=1):
                # Convert value to string and write
                cell_value = str(value) if pd.notna(value) else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # Force cell to be text format (prevents Excel auto-formatting)
                cell.number_format = "@"  # '@' is the Excel format code for text
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Apply background color
                if fill:
                    cell.fill = fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(str(filename))
        logger.info(f"Waterfall report exported to {filename}")
        return str(filename)

    def _hsl_to_hex(self, hsl_string):
        """Convert HSL color string to hex for Excel"""
        import re

        # Parse HSL string like 'hsl(284, 16%, 24%)'
        match = re.match(r"hsl\((\d+),\s*(\d+)%,\s*(\d+)%\)", hsl_string)
        if not match:
            return None

        h, s, l = int(match.group(1)), int(match.group(2)), int(match.group(3))

        # Convert HSL to RGB
        h = h / 360.0
        s = s / 100.0
        l = l / 100.0

        def hue_to_rgb(p, q, t):
            if t < 0:
                t += 1
            if t > 1:
                t -= 1
            if t < 1 / 6:
                return p + (q - p) * 6 * t
            if t < 1 / 2:
                return q
            if t < 2 / 3:
                return p + (q - p) * (2 / 3 - t) * 6
            return p

        if s == 0:
            r = g = b = l
        else:
            q = l * (1 + s) if l < 0.5 else l + s - l * s
            p = 2 * l - q
            r = hue_to_rgb(p, q, h + 1 / 3)
            g = hue_to_rgb(p, q, h)
            b = hue_to_rgb(p, q, h - 1 / 3)

        # Convert to hex
        r_hex = format(int(r * 255), "02x")
        g_hex = format(int(g * 255), "02x")
        b_hex = format(int(b * 255), "02x")

        return f"{r_hex}{g_hex}{b_hex}".upper()

    def append_delta(self, ds):
        ds[0]["Delta"] = np.nan
        previous_remaining = ds[0]["Remaining"]
        for i in range(1, len(ds)):
            d_current = ds[i]
            if pd.isna(d_current["Remaining"]):
                d_current["Delta"] = np.nan
                continue
            d_current["Delta"] = d_current["Remaining"] - previous_remaining
            previous_remaining = d_current["Remaining"]
        return ds

    def _get_color_map(self):
        """Return HSL color definitions for each row type"""
        return {
            "entry": (208, 67, 75),  # Blue
            "inclusion": (88, 51, 66),  # Green
            "exclusion": (347, 62, 77),  # Rasperry
            "component": None,  # Inherits from parent
            "final_cohort": (0, 0, 100),  # Light gray
        }

    def _get_effective_type(self, row):
        """Get the effective type of a row (component if Type is empty)"""
        return row["Type"] if row["Type"] != "" else "component"

    def _adjust_brightness(self, hsl_tuple, level):
        """Increase lightness based on component nesting level"""
        if hsl_tuple is None:
            return None
        h, s, l = hsl_tuple
        brightness_increase = min(level * 10, 30)  # +10% per level, max +30%
        adjusted_l = min(l + brightness_increase, 95)  # Cap at 95%
        return (h, s, adjusted_l)

    def _hsl_to_string(self, hsl_tuple):
        """Convert HSL tuple to CSS color string"""
        if hsl_tuple is None:
            return None
        h, s, l = hsl_tuple
        return f"hsl({int(round(h))}, {int(round(s))}%, {int(round(l))}%)"

    def _create_sparse_type_column(self):
        """Show type label only once per section (not repeated on each row)"""
        previous_type = None
        sparse_types = []
        for _type in self.df["Type"].values:
            if _type != previous_type and _type != "component":
                sparse_types.append(_type)
                previous_type = _type
            else:
                sparse_types.append("")
        self.df["Type"] = sparse_types
