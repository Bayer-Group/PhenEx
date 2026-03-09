"""
Tests for Study execution and OutputConcatenator output.

Verifies that Study writes the correct files to the correct locations and
that OutputConcatenator produces a correctly structured Excel file.
"""

import json
import unittest
from pathlib import Path

import openpyxl

from phenex import Database
from phenex.core.study import Study
from phenex.test.cohort.test_cohort_various_phenotypes_as_inex import (
    CohortWithContinuousCoverageAndExclusionTestGenerator,
    CohortWithContinuousCoverageTestGenerator,
    CohortWithLogicPhenotypeAsInclusionTestGenerator,
)
from phenex.test.cohort.test_mappings import TestDomains


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_cohort(gen_cls, name):
    gen = gen_cls()
    gen.define_mapped_tables()
    cohort = gen.define_cohort()
    cohort.name = name
    cohort.database = Database(connector=gen.con, mapper=TestDomains)
    return cohort


def _latest_exec_dir(study_dir: Path) -> Path:
    dirs = sorted(
        [d for d in study_dir.iterdir() if d.is_dir() and d.name.startswith("D20")],
        key=lambda d: d.stat().st_mtime,
    )
    return dirs[-1]


# ---------------------------------------------------------------------------
# Study output location and structure
# ---------------------------------------------------------------------------


class TestStudyOutput(unittest.TestCase):
    """Verify files written to the correct locations with correct names."""

    COHORT_NAMES = ("CohortWithExclusion", "CohortWithoutExclusion", "CohortWithComponents")
    STUDY_NAME = "output_test"

    @classmethod
    def setUpClass(cls):
        artifacts = Path(__file__).parent / "artifacts"
        artifacts.mkdir(parents=True, exist_ok=True)

        cohorts = [
            _make_cohort(
                CohortWithContinuousCoverageAndExclusionTestGenerator,
                "CohortWithExclusion",
            ),
            _make_cohort(
                CohortWithContinuousCoverageTestGenerator,
                "CohortWithoutExclusion",
            ),
            _make_cohort(
                CohortWithLogicPhenotypeAsInclusionTestGenerator,
                "CohortWithComponents",
            ),
        ]
        study = Study(name=cls.STUDY_NAME, path=str(artifacts), cohorts=cohorts)
        study.execute(overwrite=True)

        cls.exec_dir = _latest_exec_dir(artifacts / cls.STUDY_NAME)

        xlsx_files = list(cls.exec_dir.glob("results_*.xlsx"))
        cls.results_file = xlsx_files[0] if xlsx_files else None
        cls.wb = openpyxl.load_workbook(cls.results_file) if cls.results_file else None

    @classmethod
    def tearDownClass(cls):
        if cls.wb:
            cls.wb.close()

    # --- results file ---

    def test_results_file_exists(self):
        self.assertIsNotNone(self.results_file, "No results_*.xlsx found in exec dir")

    def test_results_file_name(self):
        """Filename must be results_{study_name}_{exec_dirname}.xlsx."""
        expected = f"results_{self.STUDY_NAME}_{self.exec_dir.name}.xlsx"
        self.assertEqual(self.results_file.name, expected)

    # --- sheet structure ---

    def test_sheet_names_and_order(self):
        self.assertEqual(
            self.wb.sheetnames, ["Waterfall", "WaterfallDetailed", "Table1"]
        )

    def test_each_sheet_has_both_cohort_headers(self):
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            row1 = {
                sheet.cell(row=1, column=c).value
                for c in range(1, sheet.max_column + 1)
            }
            for name in self.COHORT_NAMES:
                self.assertIn(
                    name, row1, f"Sheet '{sheet_name}': missing header '{name}'"
                )

    # --- per-cohort JSON files ---

    def test_cohort_directories_created(self):
        dirs = {d.name for d in self.exec_dir.iterdir() if d.is_dir()}
        for name in self.COHORT_NAMES:
            self.assertIn(name, dirs)

    def test_json_reports_written_per_cohort(self):
        for name in self.COHORT_NAMES:
            for fname in ("table1.json", "waterfall.json", "waterfall_detailed.json"):
                self.assertTrue(
                    (self.exec_dir / name / fname).exists(),
                    f"{name}/{fname} missing",
                )

    def test_frozen_cohort_json_written(self):
        for name in self.COHORT_NAMES:
            frozen = self.exec_dir / name / f"frozen_{name}.json"
            self.assertTrue(frozen.exists(), f"Frozen cohort JSON missing: {frozen}")

    # --- info file ---

    def test_info_file_contains_version_info(self):
        info = (self.exec_dir / "info.txt").read_text()
        self.assertIn("Python Version", info)
        self.assertIn("PhenEx Version", info)

    # --- WaterfallDetailed component phenotype rows ---

    def _waterfall_detailed_values_lower(self) -> set:
        sheet = self.wb["WaterfallDetailed"]
        return {
            str(cell).lower()
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        }

    def test_waterfall_detailed_contains_logic_phenotype(self):
        vals = self._waterfall_detailed_values_lower()
        self.assertTrue(
            any("combined cc and prior drug" in v for v in vals),
            "WaterfallDetailed missing LogicPhenotype inclusion name",
        )

    def test_waterfall_detailed_contains_component_phenotypes(self):
        vals = self._waterfall_detailed_values_lower()
        self.assertTrue(
            any("continuous coverage" in v for v in vals),
            "WaterfallDetailed missing component 'continuous_coverage'",
        )
        self.assertTrue(
            any("prior drug d1" in v for v in vals),
            "WaterfallDetailed missing component 'prior_drug_d1'",
        )


# ---------------------------------------------------------------------------
# Waterfall JSON content
# ---------------------------------------------------------------------------


class TestWaterfallJsonContent(unittest.TestCase):
    """Verify correctness of the serialised waterfall JSON."""

    COHORT_NAMES = ("CohortWithExclusion", "CohortWithoutExclusion")
    STUDY_NAME = "waterfall_json_test"

    @classmethod
    def setUpClass(cls):
        artifacts = Path(__file__).parent / "artifacts"
        artifacts.mkdir(parents=True, exist_ok=True)

        cohorts = [
            _make_cohort(
                CohortWithContinuousCoverageAndExclusionTestGenerator,
                "CohortWithExclusion",
            ),
            _make_cohort(
                CohortWithContinuousCoverageTestGenerator,
                "CohortWithoutExclusion",
            ),
        ]
        study = Study(name=cls.STUDY_NAME, path=str(artifacts), cohorts=cohorts)
        study.execute(overwrite=True)

        cls.exec_dir = _latest_exec_dir(artifacts / cls.STUDY_NAME)

    def _load_waterfall(self, cohort_name: str) -> list:
        path = self.exec_dir / cohort_name / "waterfall.json"
        return json.loads(path.read_text())["rows"]

    def test_count_columns_are_integers(self):
        """N, Remaining and Delta must serialise as JSON integers (no .0)."""
        for name in self.COHORT_NAMES:
            for row in self._load_waterfall(name):
                for col in ("N", "Remaining", "Delta"):
                    val = row.get(col)
                    if val is not None:
                        self.assertIsInstance(
                            val, int,
                            f"{name} waterfall '{col}' = {val!r} should be int",
                        )

    def test_level_column_absent(self):
        """Level is internal and must not appear in JSON output."""
        for name in self.COHORT_NAMES:
            for row in self._load_waterfall(name):
                self.assertNotIn(
                    "Level", row, f"{name}: 'Level' must not be serialised"
                )


# ---------------------------------------------------------------------------
# WaterfallDetailed component phenotypes
# ---------------------------------------------------------------------------


class TestWaterfallDetailedComponents(unittest.TestCase):
    """Verify WaterfallDetailed contains rows for component phenotypes."""

    @classmethod
    def setUpClass(cls):
        artifacts = Path(__file__).parent / "artifacts"
        artifacts.mkdir(parents=True, exist_ok=True)

        cohort = _make_cohort(
            CohortWithLogicPhenotypeAsInclusionTestGenerator,
            "CohortWithComponents",
        )
        study = Study(
            name="components_test", path=str(artifacts), cohorts=[cohort]
        )
        study.execute(overwrite=True)

        cls.exec_dir = _latest_exec_dir(artifacts / "components_test")
        xlsx_files = list(cls.exec_dir.glob("results_*.xlsx"))
        cls.wb = openpyxl.load_workbook(xlsx_files[0]) if xlsx_files else None

    @classmethod
    def tearDownClass(cls):
        if cls.wb:
            cls.wb.close()

    def _all_cell_values_lower(self, sheet_name: str) -> set:
        sheet = self.wb[sheet_name]
        return {
            str(cell).lower()
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        }

    def test_waterfall_detailed_sheet_present(self):
        self.assertIsNotNone(self.wb)
        self.assertIn("WaterfallDetailed", self.wb.sheetnames)

    def test_logic_phenotype_name_present(self):
        vals = self._all_cell_values_lower("WaterfallDetailed")
        self.assertTrue(
            any("combined cc and prior drug" in v for v in vals),
            "WaterfallDetailed missing LogicPhenotype name",
        )

    def test_component_phenotype_names_present(self):
        vals = self._all_cell_values_lower("WaterfallDetailed")
        self.assertTrue(
            any("continuous coverage" in v for v in vals),
            "WaterfallDetailed missing component 'continuous_coverage'",
        )
        self.assertTrue(
            any("prior drug d1" in v for v in vals),
            "WaterfallDetailed missing component 'prior_drug_d1'",
        )


if __name__ == "__main__":
    unittest.main()


