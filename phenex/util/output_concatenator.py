"""
Excel Output Concatenator for Study Reports

Reads per-cohort JSON reporter outputs and assembles them into a single
multi-sheet Excel file for cross-cohort comparison.
"""

import colorsys
import json
import math
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from phenex.util import create_logger

logger = create_logger(__name__)


# ---------------------------------------------------------------------------
# Cohort grouping
# ---------------------------------------------------------------------------


@dataclass
class CohortGroup:
    """A main cohort and its associated subcohorts."""

    name: str
    main_dir: Path
    subcohort_dirs: List[Path] = field(default_factory=list)

    @property
    def all_dirs(self) -> List[Path]:
        return [self.main_dir] + self.subcohort_dirs

    def display_name(self, cohort_dir: Path) -> str:
        if cohort_dir == self.main_dir:
            return self.name
        return cohort_dir.name.replace(f"{self.name}__", "")


# ---------------------------------------------------------------------------
# Shared sheet-writing utilities
# ---------------------------------------------------------------------------


class _BaseSheetWriter:
    """Low-level cell and sheet helpers shared by all writers."""

    # Row layout constants (1-indexed Excel rows)
    _SPACING_ROW = 1
    _TITLE_ROW = 2
    _SUBTITLE_ROW = 3
    _HEADER_ROW = 4
    _DATA_START_ROW = 5

    _SPACING_SIZE = 3
    _ROW_BACKGROUND_1 = "ededed"
    _SECTION_HEADER_COLOR = "FFFFFF"
    _GRAY_TEXT = "808080"
    _FONT = "Arial"
    _BOOLEAN_COLUMNS = {"N", "Pct"}
    _COLUMN_DISPLAY_NAMES = {"Pct": "%"}

    _COHORT_COLORS: List[str] = [
        "D6E4F0",  # light blue
        "D9EAD3",  # light green
        "FDE5CC",  # light orange
        "F4CCCC",  # light red
        "D9D2E9",  # light purple
        "D0E0E3",  # light teal
        "FFF2CC",  # light yellow
        "EAD1DC",  # light pink
        "CFE2F3",  # light sky
        "E6D8C3",  # light tan
    ]

    # Row type -> background fill colour (ARGB hex, no leading #)
    _WATERFALL_COLORS: Dict[str, str] = {
        "info": "DCDCDC",
        "entry": "9DC3E6",
        "inclusion": "C5E0B4",
        "exclusion": "F4B8A0",
        "component": "E2EFD9",
    }

    def _write_cell(
        self,
        sheet,
        row: int,
        col: int,
        value,
        bold: bool = False,
        italic: bool = False,
        size: int = 14,
        horizontal: str = "left",
        vertical: str = "center",
        indent: int = 0,
        fill_color: Optional[str] = None,
        number_format: Optional[str] = None,
        wrap_text: bool = False,
        font_color: Optional[str] = None,
    ):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = Font(
            name=self._FONT, bold=bold, italic=italic, size=size, color=font_color
        )
        cell.alignment = Alignment(
            horizontal=horizontal, vertical="center", indent=indent, wrap_text=wrap_text
        )
        if fill_color:
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
        if number_format:
            cell.number_format = number_format
        return cell

    def _add_group_title(
        self, sheet, title: str, start_col: int, end_col: int, color: str = "D3D3D3"
    ):
        """Write main cohort name in the title row, merged across group columns."""
        cell = self._write_cell(
            sheet,
            self._TITLE_ROW,
            start_col,
            title,
            italic=True,
            size=18,
            horizontal="left",
            indent=2,
            fill_color=color,
        )
        cell.border = Border()
        if end_col > start_col:
            sheet.merge_cells(
                start_row=self._TITLE_ROW,
                start_column=start_col,
                end_row=self._TITLE_ROW,
                end_column=end_col,
            )

    def _add_subcohort_header(
        self,
        sheet,
        name: str,
        start_col: int,
        num_cols: int,
        color: str = "E8E8E8",
        font_color: Optional[str] = None,
    ):
        """Write individual cohort/subcohort name in the subtitle row."""
        self._write_cell(
            sheet,
            self._SUBTITLE_ROW,
            start_col,
            name,
            italic=True,
            size=14,
            horizontal="left",
            indent=2,
            fill_color=color,
            font_color=font_color,
        )
        if num_cols > 1:
            sheet.merge_cells(
                start_row=self._SUBTITLE_ROW,
                start_column=start_col,
                end_row=self._SUBTITLE_ROW,
                end_column=start_col + num_cols - 1,
            )

    def _apply_group_border(
        self, sheet, start_col: int, end_col: int, color: str = "000000"
    ):
        """Draw a thin border around the cohort group from title row to last data row."""
        max_row = sheet.max_row
        top_row = self._TITLE_ROW
        side = Side(style="thin", color=color)
        for row in range(top_row, max_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                existing = cell.border
                cell.border = Border(
                    left=side if col == start_col else existing.left,
                    right=side if col == end_col else existing.right,
                    top=side if row == top_row else existing.top,
                    bottom=side if row == max_row else existing.bottom,
                )

    def _apply_spacing(self, sheet, spacing_col: int):
        """Set spacing column width and spacing row height."""
        sheet.column_dimensions[get_column_letter(spacing_col)].width = (
            self._SPACING_SIZE
        )
        sheet.row_dimensions[self._SPACING_ROW].height = self._SPACING_SIZE * 5

    def _apply_left_border_to_column(self, sheet, col: int, color: str = "000000"):
        """Draw a thin left border on every cell in a column."""
        side = Side(style="thin", color=color)
        for row in range(self._TITLE_ROW, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            existing = cell.border
            cell.border = Border(
                left=side,
                right=existing.right,
                top=existing.top,
                bottom=existing.bottom,
            )

    @staticmethod
    def _number_format_for_value(value) -> Optional[str]:
        """Return an Excel number format code suitable for value."""
        if value is None:
            return None
        if isinstance(value, int):
            return "#,##0"
        if isinstance(value, float):
            return "0.0"
        return None

    @staticmethod
    def _clean_numeric(value):
        """Convert whole-number floats to int (e.g. 98.0 -> 98)."""
        if isinstance(value, float) and not math.isnan(value) and value == int(value):
            return int(value)
        return value

    @staticmethod
    def _level_to_gray_hex(level) -> Optional[str]:
        """Return a 6-char ARGB hex fill colour for a component nesting level, or None."""
        try:
            lvl = int(level)
        except (TypeError, ValueError):
            return None
        if lvl <= 0:
            return None
        value = max(235 - 20 * (lvl - 1), 100)
        return f"{value:02X}{value:02X}{value:02X}"

    @staticmethod
    def _cohort_text_colors(hex_color: str) -> Tuple[str, str]:
        """Return (text_dark, text_light) derived from a cohort background hex colour.

        Both share the same hue as *hex_color*.  ``text_dark`` is very dark
        (for prominent text like %), ``text_light`` is medium-dark (for
        secondary text like N).
        """
        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        # text_dark – deep, saturated
        rd, gd, bd = colorsys.hls_to_rgb(h, 0.25, min(s * 1.5, 1.0))
        dark = f"{int(rd * 255):02X}{int(gd * 255):02X}{int(bd * 255):02X}"
        # text_light – medium
        rl, gl, bl = colorsys.hls_to_rgb(h, 0.45, min(s * 1.3, 1.0))
        light = f"{int(rl * 255):02X}{int(gl * 255):02X}{int(bl * 255):02X}"
        return dark, light

    def _set_default_row_height(self, sheet, height: float = 24.0) -> None:
        sheet.sheet_format.defaultRowHeight = height
        sheet.sheet_format.customHeight = True
        sheet.row_dimensions[self._TITLE_ROW].height = 34
        sheet.row_dimensions[self._SUBTITLE_ROW].height = 28

    @staticmethod
    def _load_json(path: Path) -> Dict:
        with path.open() as f:
            return json.load(f)


# ---------------------------------------------------------------------------
# Generic side-by-side writer  (Waterfall and custom reporters)
# ---------------------------------------------------------------------------


class GenericSheetWriter(_BaseSheetWriter):
    """Writes multiple JSON report files side-by-side into a single worksheet.

    For Waterfall reports the ``Type`` column drives row background colours.
    Columns are reordered and renamed according to ``_COLUMN_ORDER`` and
    ``_COLUMN_HEADERS``.  Columns not listed in the order list are appended
    at the end in their original order.
    """

    # Desired column order and display headers for Waterfall-type reports.
    _COLUMN_ORDER = [
        "Type",
        "Index",
        "Name",
        "Pct_Source_Database",
        "Pct_Remaining",
        "Remaining",
        "Delta",
        "Pct_N",
        "N",
    ]
    _COLUMN_HEADERS: Dict[str, str] = {
        "Type": "",
        "Index": "",
        "Name": "",
        "Pct_Source_Database": "% source",
        "Pct_Remaining": "%",
        "Remaining": "remaining",
        "Delta": "\u0394",
        "Pct_N": "%",
        "N": "N",
    }

    # Columns that should be sized for large numbers (2-digit millions).
    _WIDE_COLUMNS = {"N", "Remaining", "Delta"}

    def write(
        self,
        sheet,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List["CohortGroup"],
    ):
        self._set_default_row_height(sheet)
        raw_columns = self._get_raw_column_order(report_files)
        if not raw_columns:
            return

        columns = self._reorder_columns(raw_columns)
        dir_to_report = dict(zip(cohort_dirs, report_files))
        num_cols = len(columns)
        current_col = 1
        sheet.freeze_panes = sheet.cell(row=self._DATA_START_ROW, column=1)

        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]

            for cohort_dir in group.all_dirs:
                report_file = dir_to_report.get(cohort_dir)
                if report_file is None:
                    continue
                is_subcohort = cohort_dir != group.main_dir

                # Spacer before every cohort card
                self._apply_spacing(sheet, spacing_col=current_col)
                current_col += 1

                try:
                    data = self._load_json(report_file)
                    rows = data.get("rows", [])
                    display_name = group.display_name(cohort_dir)
                    title = (
                        f"{group.name} \u00b7 {display_name}"
                        if is_subcohort
                        else display_name
                    )

                    card_start = current_col
                    card_end = card_start + num_cols - 1

                    # Title row (row 2) – cohort color, merged
                    self._write_cell(
                        sheet,
                        self._TITLE_ROW,
                        card_start,
                        title,
                        bold=True,
                        italic=True,
                        size=16,
                        horizontal="left",
                        indent=1,
                        fill_color=group_color,
                    )
                    for c in range(card_start + 1, card_end + 1):
                        self._write_cell(
                            sheet, self._TITLE_ROW, c, None, fill_color=group_color
                        )
                    if card_end > card_start:
                        sheet.merge_cells(
                            start_row=self._TITLE_ROW,
                            start_column=card_start,
                            end_row=self._TITLE_ROW,
                            end_column=card_end,
                        )

                    self._write_column_headers(sheet, columns, card_start)
                    self._write_data_rows(sheet, rows, columns, card_start)
                    self._set_column_widths(sheet, rows, columns, card_start)
                    self._apply_group_border(
                        sheet, card_start, card_end, color=group_color
                    )

                    current_col = card_end + 1
                except Exception as e:
                    logger.warning(f"Failed to process {report_file}: {e}")

    # ------------------------------------------------------------------

    def _reorder_columns(self, raw_columns: List[str]) -> List[str]:
        """Return columns in the preferred order, appending any extras."""
        ordered = [c for c in self._COLUMN_ORDER if c in raw_columns]
        extras = [c for c in raw_columns if c not in self._COLUMN_ORDER]
        return ordered + extras

    @staticmethod
    def _get_raw_column_order(report_files: List[Optional[Path]]) -> List[str]:
        for f in report_files:
            if f is None:
                continue
            try:
                with f.open() as fh:
                    data = json.load(fh)
                rows = data.get("rows", [])
                if rows:
                    return list(rows[0].keys())
            except Exception:
                pass
        return []

    def _write_column_headers(self, sheet, columns: List[str], start_col: int):
        for offset, col_name in enumerate(columns):
            display = self._COLUMN_HEADERS.get(col_name, col_name)
            col = start_col + offset
            self._write_cell(
                sheet,
                self._HEADER_ROW,
                col,
                display,
                bold=True,
                size=14,
                horizontal="center",
            )

    def _write_data_rows(self, sheet, rows: list, columns: List[str], start_col: int):
        display_rows = self._sparsify_type(rows) if "Type" in columns else rows
        for row_idx, (orig_row, disp_row) in enumerate(
            zip(rows, display_rows),
            start=self._DATA_START_ROW,
        ):
            row_type = str(orig_row.get("Type", "")).lower()
            fill_color = (
                None if row_type == "info" else self._WATERFALL_COLORS.get(row_type)
            )
            for offset, col_name in enumerate(columns):
                value = self._clean_numeric(disp_row.get(col_name))
                fmt = self._number_format_for_value(value)
                self._write_cell(
                    sheet,
                    row_idx,
                    start_col + offset,
                    value,
                    size=14,
                    horizontal="center",
                    fill_color=fill_color,
                    number_format=fmt,
                )

    @staticmethod
    def _sparsify_type(rows: list) -> list:
        """Return a copy of rows with Type showing only on the first row of each group."""
        previous_type = None
        result = []
        for row in rows:
            row_copy = dict(row)
            type_val = str(row_copy.get("Type", ""))
            if (
                type_val
                and type_val != previous_type
                and type_val.lower() not in ("component", "info")
            ):
                previous_type = type_val
            else:
                row_copy["Type"] = ""
            result.append(row_copy)
        return result

    def _set_column_widths(self, sheet, rows: list, columns: List[str], start_col: int):
        font_scale = 14 / 11
        for offset, col_name in enumerate(columns):
            if col_name in self._WIDE_COLUMNS:
                width = 14  # fits "10,000,000"
            else:
                display = self._COLUMN_HEADERS.get(col_name, col_name)
                values = [
                    str(r.get(col_name, ""))
                    for r in rows
                    if r.get(col_name) is not None
                ]
                max_val_len = max((len(v) for v in values), default=0)
                content_len = max(len(display), max_val_len)
                width = min(max(content_len * font_scale + 2, 6), 40)
            sheet.column_dimensions[get_column_letter(start_col + offset)].width = width


# ---------------------------------------------------------------------------
# Table1 aligned writer with section headers
# ---------------------------------------------------------------------------


class Table1SheetWriter(_BaseSheetWriter):
    """Writes Table1 JSON files into a single aligned worksheet.

    Layout::

        Row 1  : spacing row
        Row 2  : main cohort name banner per group (merged)
        Row 3  : individual cohort/subcohort name per block
        Row 4  : column labels per cohort block
        Row 5+ : characteristic rows; section-header rows interleaved

    Column A is a narrow spacer that mirrors row colours.
    Column B holds the characteristic names.

    Section headers appear as full-width blue-grey rows when the first cohort
    JSON contains a ``sections`` dict mapping section names to characteristic
    display names.
    """

    def write(
        self,
        sheet,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List["CohortGroup"],
        boolean_only: bool = False,
    ):
        self._set_default_row_height(sheet)
        master_names, sections, name_to_level = self._build_master_names_and_sections(
            report_files
        )
        if not master_names:
            return

        expanded = self._build_expanded_rows(master_names, sections, name_to_level)
        alt_mask = self._compute_alternate_fills(expanded, name_to_level)
        self._write_name_column(sheet, expanded, name_to_level, alt_mask)

        dir_to_report = dict(zip(cohort_dirs, report_files))
        current_col = 3
        spacing_cols: List[int] = []

        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            text_dark, text_light = self._cohort_text_colors(group_color)
            spacing_cols.append(current_col)
            self._apply_spacing(sheet, spacing_col=current_col)
            current_col += 1
            group_start_col = current_col

            for cohort_dir in group.all_dirs:
                report_file = dir_to_report.get(cohort_dir)
                if report_file is None:
                    continue
                is_subcohort = cohort_dir != group.main_dir
                try:
                    data = self._load_json(report_file)
                    rows = data.get("rows", [])
                    columns = [
                        c
                        for c in (rows[0].keys() if rows else [])
                        if c not in ("Name", "_level")
                    ]
                    if boolean_only:
                        columns = [c for c in columns if c in self._BOOLEAN_COLUMNS]
                    if not columns:
                        continue
                    num_value_cols = len(columns)

                    display_name = group.display_name(cohort_dir)
                    self._add_subcohort_header(
                        sheet,
                        display_name,
                        current_col,
                        num_value_cols,
                        color=group_color,
                        font_color=text_dark,
                    )
                    pct_offset = self._write_column_labels(
                        sheet,
                        columns,
                        current_col,
                        text_dark,
                        text_light,
                    )
                    self._write_value_rows(
                        sheet,
                        rows,
                        columns,
                        expanded,
                        current_col,
                        pct_offset,
                        name_to_level,
                        alt_mask,
                        group_color,
                        text_dark,
                        text_light,
                    )
                    self._set_value_column_widths(sheet, rows, columns, current_col)

                    if is_subcohort:
                        self._apply_left_border_to_column(
                            sheet, current_col, color=group_color
                        )

                    current_col += num_value_cols

                except Exception as e:
                    logger.warning(f"Failed to process {report_file}: {e}")

            group_end_col = current_col - 1
            if group_end_col >= group_start_col:
                self._add_group_title(
                    sheet, group.name, group_start_col, group_end_col, color=group_color
                )
                self._apply_group_border(
                    sheet, group_start_col, group_end_col, color=group_color
                )

        # Apply section-header spans, skipping spacer columns
        self._apply_section_header_spans(
            sheet, expanded, sheet.max_column, spacing_cols
        )

    # ------------------------------------------------------------------

    def _build_master_names_and_sections(
        self, report_files: List[Optional[Path]]
    ) -> Tuple[List[str], Optional[Dict], Dict[str, int]]:
        """Return (master_names, sections, name_to_level) aggregated across all JSON files."""
        master_names: List[str] = []
        existing_names: set = set()
        first_file_done = False
        sections = None
        name_to_level: Dict[str, int] = {}
        for report_file in report_files:
            if report_file is None:
                continue
            try:
                data = self._load_json(report_file)
                if sections is None:
                    sections = data.get("sections")
                for row in data.get("rows", []):
                    name = row.get("Name")
                    if name is not None:
                        name_str = str(name)
                        if not first_file_done:
                            # First file: preserve all names including duplicates
                            master_names.append(name_str)
                            existing_names.add(name_str)
                        elif name_str not in existing_names:
                            # Subsequent files: only add genuinely new names
                            master_names.append(name_str)
                            existing_names.add(name_str)
                        level = row.get("_level", 0)
                        if name_str not in name_to_level and level:
                            name_to_level[name_str] = int(level)
                first_file_done = True
            except Exception as e:
                logger.warning(f"Could not read {report_file}: {e}")
        return master_names, sections, name_to_level

    def _build_expanded_rows(
        self,
        master_names: List[str],
        sections: Optional[Dict],
        name_to_level: Dict[str, int] = None,
    ) -> List[Tuple[str, str]]:
        """Return list of (row_type, value) with section headers interleaved.

        ``row_type`` is ``"section"`` for section-header rows and ``"row"``
        for data rows.
        """
        if not sections:
            return [("row", n) for n in master_names]

        insert_at: Dict[int, str] = {}
        search_start = 0
        for section_name, char_names in sections.items():
            for idx in range(search_start, len(master_names)):
                if self._name_belongs_to_chars(master_names[idx], char_names):
                    insert_at[idx] = section_name
                    # Advance past all rows for this section by counting
                    # base char_names matches (skips over binned variants)
                    count = 0
                    end = idx
                    while end < len(master_names) and count < len(char_names):
                        if master_names[end] in char_names:
                            count += 1
                        end += 1
                    search_start = end
                    break

        result: List[Tuple[str, str]] = []
        section_start_indices = sorted(insert_at.keys())

        # Build a mapping: master_names index -> which section it belongs to
        current_section_idx = -1
        owner: List[int] = []
        for idx in range(len(master_names)):
            if idx in insert_at:
                current_section_idx = idx
            owner.append(current_section_idx)

        for idx, name in enumerate(master_names):
            if idx in insert_at:
                result.append(("section", insert_at[idx]))
            result.append(("row", name))
            # Add spacer after the last row of this section
            is_last = idx == len(master_names) - 1
            next_in_different_section = not is_last and owner[idx + 1] != owner[idx]
            if owner[idx] >= 0 and (is_last or next_in_different_section):
                result.append(("spacer", ""))
        return self._insert_binned_spacers(result, name_to_level or {})

    @staticmethod
    def _is_binned(name: str, name_to_level: Dict[str, int]) -> bool:
        """A row is binned if it contains '=' and is not a component (level 0)."""
        return "=" in name and name_to_level.get(name, 0) == 0

    @staticmethod
    def _insert_binned_spacers(
        expanded: List[Tuple[str, str]],
        name_to_level: Dict[str, int],
    ) -> List[Tuple[str, str]]:
        """Insert a spacer row before each group of binned rows.

        A spacer is not inserted if the binned group is the first data row
        or immediately follows a section header.
        """
        result: List[Tuple[str, str]] = []
        for i, (row_type, value) in enumerate(expanded):
            if row_type == "row" and "=" in value and name_to_level.get(value, 0) == 0:
                prev_type = result[-1][0] if result else None
                prev_is_binned = (
                    (
                        prev_type == "row"
                        and "=" in result[-1][1]
                        and name_to_level.get(result[-1][1], 0) == 0
                    )
                    if result
                    else False
                )
                if not prev_is_binned and prev_type not in (None, "section"):
                    result.append(("spacer", ""))
            result.append((row_type, value))
        return result

    def _compute_alternate_fills(
        self,
        expanded: List[Tuple[str, str]],
        name_to_level: Dict[str, int],
    ) -> List[bool]:
        """Return per-row boolean mask: True where an alternating fill applies.

        Resets at each section.  The first eligible row after a section
        header gets no fill; subsequent eligible rows alternate.
        Component rows (level > 0) and spacer rows are always False.
        """
        mask: List[bool] = []
        counter = 0
        for row_type, value in expanded:
            if row_type == "section":
                counter = 0
                mask.append(False)
            elif row_type == "spacer":
                mask.append(False)
            else:
                if name_to_level.get(value, 0) > 0:
                    mask.append(False)
                else:
                    mask.append(counter % 2 == 1)
                    counter += 1
        return mask

    @staticmethod
    def _name_belongs_to_chars(name: str, char_names: List[str]) -> bool:
        for char in char_names:
            if name == char or name.startswith(char + "="):
                return True
        return False

    _NAME_SPACER_COL = 1
    _NAME_COL = 2

    def _write_name_column(
        self,
        sheet,
        expanded: List[Tuple[str, str]],
        name_to_level: Dict[str, int] = None,
        alt_mask: List[bool] = None,
    ):
        """Populate col B with characteristic names; col A is a narrow colour spacer."""
        name_to_level = name_to_level or {}
        sheet.column_dimensions[get_column_letter(self._NAME_SPACER_COL)].width = (
            self._SPACING_SIZE
        )

        for i, (row_type, value) in enumerate(expanded):
            out_row = self._DATA_START_ROW + i
            if row_type == "section":
                self._write_cell(
                    sheet,
                    out_row,
                    self._NAME_COL,
                    value,
                    bold=True,
                    size=14,
                    horizontal="left",
                    indent=2,
                    fill_color=self._SECTION_HEADER_COLOR,
                )
                self._write_cell(
                    sheet,
                    out_row,
                    self._NAME_SPACER_COL,
                    None,
                    fill_color=self._SECTION_HEADER_COLOR,
                )
                sheet.row_dimensions[out_row].height = 36
            elif row_type == "spacer":
                sheet.cell(row=out_row, column=self._NAME_COL, value=None)
                sheet.cell(row=out_row, column=self._NAME_SPACER_COL, value=None)
            else:
                is_cohort = value == "Cohort"
                level = name_to_level.get(value, 0)
                gray = self._level_to_gray_hex(level)
                alt = self._ROW_BACKGROUND_1 if (alt_mask and alt_mask[i]) else None
                fill = gray if gray else alt
                display = None if is_cohort else value
                if display and "=" in display and name_to_level.get(value, 0) == 0:
                    self._write_binned_name_cell(
                        sheet,
                        out_row,
                        self._NAME_COL,
                        display,
                        fill_color=fill,
                    )
                else:
                    self._write_cell(
                        sheet,
                        out_row,
                        self._NAME_COL,
                        display,
                        bold=is_cohort,
                        size=14,
                        horizontal="right",
                        indent=1,
                        fill_color=fill,
                    )
                self._write_cell(
                    sheet,
                    out_row,
                    self._NAME_SPACER_COL,
                    None,
                    fill_color=fill,
                )

        max_name_len = max((len(t[1]) for t in expanded if t[0] == "row"), default=10)
        sheet.column_dimensions[get_column_letter(self._NAME_COL)].width = min(
            max(max_name_len * 2.4, 28), 60
        )
        sheet.freeze_panes = sheet.cell(
            row=self._DATA_START_ROW + 1, column=self._NAME_COL + 1
        )

    def _write_binned_name_cell(
        self,
        sheet,
        row: int,
        col: int,
        value: str,
        fill_color: Optional[str] = None,
    ):
        """Write a binned phenotype name with the part after '=' in bold."""
        prefix, suffix = value.split("=", 1)
        rich = CellRichText(
            TextBlock(InlineFont(rFont=self._FONT, sz=14), prefix + "="),
            TextBlock(InlineFont(rFont=self._FONT, sz=14, b=True), suffix),
        )
        cell = sheet.cell(row=row, column=col, value=rich)
        cell.alignment = Alignment(horizontal="right", indent=1)
        if fill_color:
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    def _write_column_labels(
        self,
        sheet,
        columns: List[str],
        start_col: int,
        text_dark: Optional[str] = None,
        text_light: Optional[str] = None,
    ) -> Optional[int]:
        """Write column labels in the header row; return 0-based offset of the Pct column."""
        pct_offset = None
        for offset, col_name in enumerate(columns):
            is_pct = col_name.strip().lower() == "pct"
            is_n = col_name.strip().lower() == "n"
            if is_pct:
                pct_offset = offset
            display_name = self._COLUMN_DISPLAY_NAMES.get(col_name, col_name)
            horizontal = "right" if is_n else ("left" if is_pct else "center")
            fc = (
                text_light
                if is_n and text_light
                else (text_dark if is_pct and text_dark else None)
            )
            cell = self._write_cell(
                sheet,
                self._HEADER_ROW,
                start_col + offset,
                display_name,
                bold=is_pct,
                size=14,
                horizontal=horizontal,
                font_color=fc,
            )
            cell.border = Border()
        return pct_offset

    def _write_value_rows(
        self,
        sheet,
        rows,
        columns: List[str],
        expanded: List[Tuple[str, str]],
        start_col: int,
        pct_offset: Optional[int],
        name_to_level: Dict[str, int] = None,
        alt_mask: List[bool] = None,
        group_color: Optional[str] = None,
        text_dark: Optional[str] = None,
        text_light: Optional[str] = None,
    ):
        name_to_level = name_to_level or {}
        # Build per-name list of rows to handle duplicate names (e.g. TPA periods)
        _name_row_lists: Dict[str, List] = defaultdict(list)
        for r in rows:
            if "Name" in r:
                _name_row_lists[str(r["Name"])].append(r)
        _name_cursors: Dict[str, int] = defaultdict(int)
        for i, (row_type, value) in enumerate(expanded):
            out_row = self._DATA_START_ROW + i
            if row_type == "section":
                for offset in range(len(columns)):
                    self._write_cell(
                        sheet,
                        out_row,
                        start_col + offset,
                        None,
                        fill_color=self._SECTION_HEADER_COLOR,
                    )
                continue
            if row_type == "spacer":
                continue

            cursor = _name_cursors[value]
            entries = _name_row_lists.get(value, [])
            row_data = entries[cursor] if cursor < len(entries) else None
            _name_cursors[value] = cursor + 1
            is_cohort = value == "Cohort"
            level = name_to_level.get(value, 0)
            gray = self._level_to_gray_hex(level)
            should_alt = alt_mask[i] if alt_mask else False
            row_fill = gray if gray else (group_color if should_alt else None)
            for offset, col_name in enumerate(columns):
                raw_value = (
                    self._clean_numeric(row_data.get(col_name)) if row_data else None
                )
                is_pct_col = offset == pct_offset
                is_n_col = col_name.strip().lower() == "n"
                horizontal = (
                    "right" if is_n_col else ("left" if is_pct_col else "center")
                )
                fmt = self._number_format_for_value(raw_value)
                fc = text_light if is_n_col else (text_dark if is_pct_col else None)
                self._write_cell(
                    sheet,
                    out_row,
                    start_col + offset,
                    raw_value,
                    bold=is_pct_col,
                    size=14,
                    horizontal=horizontal,
                    fill_color=row_fill,
                    number_format=fmt,
                    font_color=fc,
                )

    def _apply_section_header_spans(
        self,
        sheet,
        expanded: List[Tuple[str, str]],
        max_col: int,
        spacing_cols: List[int] = None,
    ):
        """Apply fill colour across the full row width for section headers."""
        if max_col < 2:
            return
        skip = set(spacing_cols or [])
        fill = PatternFill(
            start_color=self._SECTION_HEADER_COLOR,
            end_color=self._SECTION_HEADER_COLOR,
            fill_type="solid",
        )
        for i, (row_type, _) in enumerate(expanded):
            if row_type == "section":
                out_row = self._DATA_START_ROW + i
                for col in range(1, max_col + 1):
                    if col in skip:
                        continue
                    cell = sheet.cell(row=out_row, column=col)
                    if (
                        not cell.fill
                        or cell.fill.fgColor is None
                        or cell.fill.fgColor.rgb == "00000000"
                    ):
                        cell.fill = fill

    def _set_value_column_widths(self, sheet, rows, columns: List[str], start_col: int):
        font_scale = 14 / 11
        for offset, col_name in enumerate(columns):
            if col_name.strip().lower() == "n":
                width = 14  # fits "10,000,000"
            else:
                display_name = self._COLUMN_DISPLAY_NAMES.get(col_name, col_name)
                values = [
                    str(r.get(col_name, ""))
                    for r in rows
                    if r.get(col_name) is not None
                ]
                max_val_len = max((len(v) for v in values), default=0)
                content_len = max(len(display_name), max_val_len)
                width = min(max(content_len * font_scale + 1, 6), 18)
            sheet.column_dimensions[get_column_letter(start_col + offset)].width = width


# ---------------------------------------------------------------------------
# Table1 numeric + categorical horizontal writer
# ---------------------------------------------------------------------------


class Table1NumericSheetWriter(_BaseSheetWriter):
    """Writes numeric and categorical characteristics in horizontally-stacked blocks.

    Layout::

        Col A  : narrow spacer (no fill)
        Col B  : cohort display names (cohort colour, right-justified, frozen)
        Col C+ : phenotype blocks separated by spacer columns

    Header rows (all frozen)::

        Row 1  : spacing
        Row 2  : phenotype names (large text, merged across block columns)
        Row 3  : stat column labels (numeric) / category names (categorical)
        Row 4  : blank (numeric) / N % sub-headers (categorical)
        Row 5+ : one row per cohort/subcohort
    """

    _NUMERIC_COL_ORDER = [
        "Mean",
        "STD",
        "Min",
        "P10",
        "P25",
        "Median",
        "P75",
        "P90",
        "Max",
        "N",
        "Pct",
    ]
    _GRAYED_STATS = {"STD", "N", "Pct"}

    _SPACER_COL = 1
    _NAME_COL = 2
    _DATA_START_COL = 3

    def write(
        self,
        sheet,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List["CohortGroup"],
    ):
        self._set_default_row_height(sheet)
        dir_to_report = dict(zip(cohort_dirs, report_files))

        blocks = self._collect_blocks(report_files)
        if not blocks:
            return

        # Build ordered cohort entries with pre-loaded data
        cohort_entries: List[Dict] = []
        for gi, group in enumerate(cohort_groups):
            gc = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            td, tl = self._cohort_text_colors(gc)
            for cd in group.all_dirs:
                rf = dir_to_report.get(cd)
                if rf is None:
                    continue
                try:
                    data = self._load_json(rf)
                except Exception:
                    continue
                cohort_entries.append(
                    {
                        "display_name": group.display_name(cd),
                        "group_color": gc,
                        "text_dark": td,
                        "text_light": tl,
                        "is_subcohort": cd != group.main_dir,
                        "data": data,
                    }
                )

        if not cohort_entries:
            return

        # Fixed columns
        sheet.column_dimensions[get_column_letter(self._SPACER_COL)].width = (
            self._SPACING_SIZE
        )
        sheet.column_dimensions[get_column_letter(self._NAME_COL)].width = 24
        sheet.row_dimensions[self._SPACING_ROW].height = self._SPACING_SIZE * 5
        sheet.row_dimensions[self._TITLE_ROW].height = 32

        # Write cohort name column
        for ri, entry in enumerate(cohort_entries):
            row = self._DATA_START_ROW + ri
            self._write_cell(
                sheet,
                row,
                self._NAME_COL,
                entry["display_name"],
                size=14,
                horizontal="right",
                indent=6 if entry["is_subcohort"] else 4,
                fill_color=entry["group_color"],
            )

        # Write phenotype blocks
        current_col = self._DATA_START_COL
        last_data_row = self._DATA_START_ROW + len(cohort_entries) - 1
        for block in blocks:
            self._apply_spacing(sheet, spacing_col=current_col)
            current_col += 1
            block_start = current_col
            if block["kind"] == "numeric":
                current_col = self._write_numeric_block(
                    sheet,
                    current_col,
                    block,
                    cohort_entries,
                )
            else:
                current_col = self._write_categorical_block(
                    sheet,
                    current_col,
                    block,
                    cohort_entries,
                )
            self._apply_block_border(
                sheet,
                self._TITLE_ROW,
                last_data_row,
                block_start,
                current_col - 1,
            )

        # Freeze header rows + cohort name columns
        sheet.freeze_panes = sheet.cell(
            row=self._DATA_START_ROW,
            column=self._NAME_COL + 1,
        )

    # ------------------------------------------------------------------

    def _collect_blocks(self, report_files: List[Optional[Path]]) -> List[Dict]:
        """Return phenotype block descriptors (numeric and categorical) in data order."""
        order: List[Tuple[str, str]] = []
        seen: set = set()
        stat_cols_available: set = set()
        categorical_phenos: Dict[str, Dict[str, None]] = {}

        for f in report_files:
            if f is None:
                continue
            try:
                data = self._load_json(f)
                for row in data.get("rows", []):
                    name = row.get("Name", "")
                    level = row.get("_level", 0)
                    if "=" in name and int(level) == 0:
                        pheno, cat = name.split("=", 1)
                        if pheno not in categorical_phenos:
                            categorical_phenos[pheno] = {}
                        if pheno not in seen:
                            order.append(("categorical", pheno))
                            seen.add(pheno)
                        if cat not in categorical_phenos[pheno]:
                            categorical_phenos[pheno][cat] = None
                    elif name and "=" not in name and self._has_numeric_mean(row):
                        if name not in seen:
                            order.append(("numeric", name))
                            seen.add(name)
                        for col in self._NUMERIC_COL_ORDER:
                            if row.get(col) is not None:
                                stat_cols_available.add(col)
            except Exception:
                pass

        available = [c for c in self._NUMERIC_COL_ORDER if c in stat_cols_available]

        blocks: List[Dict] = []
        for kind, name in order:
            if kind == "numeric":
                blocks.append({"kind": "numeric", "name": name, "stat_cols": available})
            else:
                blocks.append(
                    {
                        "kind": "categorical",
                        "name": name,
                        "categories": list(categorical_phenos[name].keys()),
                    }
                )
        return blocks

    def _write_numeric_block(
        self,
        sheet,
        start_col: int,
        block: Dict,
        cohort_entries: List[Dict],
    ) -> int:
        """Write headers and data for one numeric phenotype. Returns next free column."""
        name = block["name"]
        stat_cols = block["stat_cols"]
        num_cols = len(stat_cols)
        end_col = start_col + num_cols - 1

        # Row 2: phenotype name (merged)
        self._write_cell(
            sheet,
            self._TITLE_ROW,
            start_col,
            name,
            bold=True,
            size=16,
            horizontal="center",
        )
        if num_cols > 1:
            for c in range(start_col + 1, end_col + 1):
                sheet.cell(row=self._TITLE_ROW, column=c)
            sheet.merge_cells(
                start_row=self._TITLE_ROW,
                start_column=start_col,
                end_row=self._TITLE_ROW,
                end_column=end_col,
            )

        # Row 3: stat column labels
        for i, col_name in enumerate(stat_cols):
            display = self._COLUMN_DISPLAY_NAMES.get(col_name, col_name)
            grayed = col_name in self._GRAYED_STATS
            self._write_cell(
                sheet,
                self._SUBTITLE_ROW,
                start_col + i,
                display,
                bold=not grayed,
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT if grayed else None,
            )

        # Data rows + track max content length for widths
        font_scale = 14 / 11
        col_max_lens = {c: len(self._COLUMN_DISPLAY_NAMES.get(c, c)) for c in stat_cols}

        for ri, entry in enumerate(cohort_entries):
            row = self._DATA_START_ROW + ri
            row_data = self._find_char_row(entry["data"], name)
            alt_fill = entry["group_color"] if ri % 2 == 1 else None
            td, tl = entry["text_dark"], entry["text_light"]
            for i, col_name in enumerate(stat_cols):
                value = (
                    self._clean_numeric(row_data.get(col_name)) if row_data else None
                )
                fmt = self._number_format_for_value(value)
                if col_name == "STD":
                    fc = self._GRAY_TEXT
                elif col_name == "N":
                    fc = tl
                elif col_name == "Pct":
                    fc = td
                else:
                    fc = td
                self._write_cell(
                    sheet,
                    row,
                    start_col + i,
                    value,
                    size=14,
                    horizontal="center",
                    fill_color=alt_fill,
                    number_format=fmt,
                    font_color=fc,
                )
                if value is not None:
                    col_max_lens[col_name] = max(
                        col_max_lens[col_name],
                        len(str(value)),
                    )

        # Column widths
        for i, col_name in enumerate(stat_cols):
            if col_name == "N":
                w = 14
            else:
                content_len = col_max_lens.get(col_name, 4)
                w = min(max(content_len * font_scale + 1, 6), 18)
            sheet.column_dimensions[get_column_letter(start_col + i)].width = w

        return end_col + 1

    def _write_categorical_block(
        self,
        sheet,
        start_col: int,
        block: Dict,
        cohort_entries: List[Dict],
    ) -> int:
        """Write headers and data for one categorical phenotype. Returns next free column."""
        name = block["name"]
        categories = block["categories"]
        num_cols = len(categories) * 2
        end_col = start_col + num_cols - 1

        # Row 2: phenotype name (merged)
        self._write_cell(
            sheet,
            self._TITLE_ROW,
            start_col,
            name,
            bold=True,
            size=16,
            horizontal="center",
        )
        if num_cols > 1:
            for c in range(start_col + 1, end_col + 1):
                sheet.cell(row=self._TITLE_ROW, column=c)
            sheet.merge_cells(
                start_row=self._TITLE_ROW,
                start_column=start_col,
                end_row=self._TITLE_ROW,
                end_column=end_col,
            )

        # Row 3: category names (merged over N/% pairs)
        for ci, cat_name in enumerate(categories):
            n_col = start_col + ci * 2
            pct_col = n_col + 1
            cell = self._write_cell(
                sheet,
                self._SUBTITLE_ROW,
                n_col,
                cat_name,
                bold=True,
                size=14,
                horizontal="center",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="bottom",
                wrap_text=True,
            )
            sheet.merge_cells(
                start_row=self._SUBTITLE_ROW,
                start_column=n_col,
                end_row=self._SUBTITLE_ROW,
                end_column=pct_col,
            )

        # Adjust row 3 height for wrapped category names
        font_scale = 14 / 11
        chars_per_col = int(14 / font_scale)
        max_lines = 1
        for cat in categories:
            lines = max(1, math.ceil(len(cat) / max(chars_per_col, 1)))
            max_lines = max(max_lines, lines)
        needed = max(24, max_lines * 20)
        current = sheet.row_dimensions[self._SUBTITLE_ROW].height or 24
        sheet.row_dimensions[self._SUBTITLE_ROW].height = max(current, needed)

        # Row 4: N/% sub-headers
        for ci in range(len(categories)):
            n_col = start_col + ci * 2
            pct_col = n_col + 1
            self._write_cell(
                sheet,
                self._HEADER_ROW,
                n_col,
                "N",
                size=14,
                horizontal="right",
                font_color=self._GRAY_TEXT,
            )
            self._write_cell(
                sheet,
                self._HEADER_ROW,
                pct_col,
                "%",
                bold=True,
                size=14,
                horizontal="left",
            )

        # Data rows
        for ri, entry in enumerate(cohort_entries):
            row = self._DATA_START_ROW + ri
            cat_data = self._get_category_data(entry["data"], name, categories)
            alt_fill = entry["group_color"] if ri % 2 == 1 else None
            td, tl = entry["text_dark"], entry["text_light"]
            for ci, cat in enumerate(categories):
                n_col = start_col + ci * 2
                pct_col = n_col + 1
                vals = cat_data.get(cat, {})
                n_val = self._clean_numeric(vals.get("N"))
                pct_val = self._clean_numeric(vals.get("Pct"))
                self._write_cell(
                    sheet,
                    row,
                    n_col,
                    n_val,
                    size=14,
                    horizontal="right",
                    fill_color=alt_fill,
                    number_format=self._number_format_for_value(n_val),
                    font_color=tl,
                )
                self._write_cell(
                    sheet,
                    row,
                    pct_col,
                    pct_val,
                    bold=True,
                    size=14,
                    horizontal="left",
                    fill_color=alt_fill,
                    number_format=self._number_format_for_value(pct_val),
                    font_color=td,
                )

        # Column widths
        for ci in range(len(categories)):
            n_col = start_col + ci * 2
            pct_col = n_col + 1
            sheet.column_dimensions[get_column_letter(n_col)].width = 14
            sheet.column_dimensions[get_column_letter(pct_col)].width = 6

        return end_col + 1

    # ------------------------------------------------------------------

    def _apply_block_border(
        self,
        sheet,
        top_row: int,
        bottom_row: int,
        start_col: int,
        end_col: int,
    ):
        """Draw a thin black border around a phenotype block."""
        side = Side(style="thin", color="000000")
        for row in range(top_row, bottom_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                existing = cell.border
                cell.border = Border(
                    left=side if col == start_col else existing.left,
                    right=side if col == end_col else existing.right,
                    top=side if row == top_row else existing.top,
                    bottom=side if row == bottom_row else existing.bottom,
                )

    @staticmethod
    def _has_numeric_mean(row: Dict) -> bool:
        """Return True if the row has a real (non-NaN) Mean value."""
        val = row.get("Mean")
        if val is None:
            return False
        if isinstance(val, float) and math.isnan(val):
            return False
        if isinstance(val, str) and val.lower() in ("nan", ""):
            return False
        return True

    @staticmethod
    def _find_char_row(data: Dict, char_name: str) -> Optional[Dict]:
        for row in data.get("rows", []):
            if row.get("Name") == char_name:
                return row
        return None

    @staticmethod
    def _get_category_data(
        data: Dict,
        phenotype_name: str,
        categories: List[str],
    ) -> Dict[str, Dict[str, object]]:
        result: Dict[str, Dict] = {}
        for row in data.get("rows", []):
            name = row.get("Name", "")
            if "=" in name:
                pheno, cat = name.split("=", 1)
                if pheno == phenotype_name and cat in categories:
                    result[cat] = {"N": row.get("N"), "Pct": row.get("Pct")}
        return result


# ---------------------------------------------------------------------------
# Simplified attrition card writer
# ---------------------------------------------------------------------------


class SimplifiedAttritionTable(_BaseSheetWriter):
    """Writes attrition cards side-by-side, one per cohort/subcohort.

    Each card is a compact table with downward arrows between steps.
    Component rows are filtered out.  A coloured border surrounds each
    card and horizontal dividers separate entry/inclusion/exclusion
    sections.

    Columns per card: Type | Index | Name | % entry | N | Δ | % source
    """

    _ARROW = "\u2193"  # ↓
    _NUM_CARD_COLS = 7

    # Column offsets within a card (0-based from card_start)
    _OFF_TYPE = 0
    _OFF_INDEX = 1
    _OFF_NAME = 2
    _OFF_PCT = 3  # Pct_Remaining
    _OFF_N = 4  # Remaining
    _OFF_DELTA = 5
    _OFF_SRC = 6  # Pct_Source_Database

    _CARD_WIDTHS = [14, 4, 28, 10, 14, 14, 12]

    def write(self, sheet, report_files, cohort_dirs, cohort_groups):
        sheet.sheet_format.defaultRowHeight = 28.0
        sheet.sheet_format.customHeight = True
        dir_to_report = dict(zip(cohort_dirs, report_files))

        expanded = self._build_template(report_files)
        if not expanded:
            return

        title_row = 2
        header_row = 3
        data_start = 4
        sheet.row_dimensions[1].height = self._SPACING_SIZE * 5

        # Row heights
        for i, exp in enumerate(expanded):
            row = data_start + i
            kind = exp["kind"]
            if kind == "spacer":
                sheet.row_dimensions[row].height = 8
            elif kind == "arrow":
                sheet.row_dimensions[row].height = 22
            else:
                sheet.row_dimensions[row].height = 28

        last_data_row = data_start + len(expanded) - 1
        sheet.freeze_panes = sheet.cell(row=data_start, column=1)

        current_col = 1
        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            for cohort_dir in group.all_dirs:
                report_file = dir_to_report.get(cohort_dir)
                if report_file is None:
                    continue

                self._apply_spacing(sheet, spacing_col=current_col)
                current_col += 1

                try:
                    data = self._load_json(report_file)
                    rows = self._filter_components(data.get("rows", []))
                    display = group.display_name(cohort_dir)
                    is_sub = cohort_dir != group.main_dir
                    title = f"{group.name} \u00b7 {display}" if is_sub else display

                    cs = current_col  # card start
                    ce = cs + self._NUM_CARD_COLS - 1

                    self._write_title(sheet, title_row, cs, ce, title, group_color)
                    self._write_headers(sheet, header_row, cs)
                    self._write_data(sheet, data_start, cs, rows, expanded, group_color)
                    self._set_widths(sheet, cs)
                    self._apply_card_border(
                        sheet, title_row, last_data_row, cs, ce, group_color
                    )
                    self._apply_section_dividers(
                        sheet, data_start, cs, ce, expanded, group_color
                    )

                    current_col = ce + 1
                except Exception as e:
                    logger.warning(f"Failed to process {report_file}: {e}")

    # ------------------------------------------------------------------

    @staticmethod
    def _filter_components(rows: list) -> list:
        return [r for r in rows if str(r.get("Type", "")).lower() != "component"]

    def _build_template(self, report_files):
        """Build expanded row template from first available waterfall file."""
        for f in report_files:
            if f is None:
                continue
            try:
                data = self._load_json(f)
                rows = self._filter_components(data.get("rows", []))
                if rows:
                    return self._expand_rows(rows)
            except Exception:
                pass
        return []

    def _expand_rows(self, rows):
        """Build display row template with arrows and spacers interleaved."""
        expanded = []
        prev_type = None

        for i, r in enumerate(rows):
            rtype = str(r.get("Type", "")).lower()
            is_first = i == 0
            is_last = i == len(rows) - 1

            # Database info row (first row)
            if is_first and rtype == "info":
                expanded.append({"kind": "db", "idx": i})
                expanded.append({"kind": "spacer"})
                prev_type = rtype
                continue

            # Spacer before a new type group
            if prev_type and prev_type != rtype and prev_type != "info":
                expanded.append({"kind": "spacer"})

            # Final cohort row – no arrow before it
            if is_last and rtype == "info":
                expanded.append({"kind": "final", "idx": i})
                prev_type = rtype
                continue

            # Arrow row
            show_type = rtype != prev_type
            expanded.append(
                {
                    "kind": "arrow",
                    "type_label": rtype if show_type else None,
                    "idx": i,
                }
            )

            # Data row
            expanded.append({"kind": "data", "idx": i})
            prev_type = rtype

        return expanded

    # ------------------------------------------------------------------

    def _write_title(self, sheet, row, cs, ce, title, color):
        self._write_cell(
            sheet,
            row,
            cs,
            title,
            bold=True,
            italic=True,
            size=16,
            horizontal="left",
            indent=1,
            fill_color=color,
        )
        for c in range(cs + 1, ce + 1):
            self._write_cell(sheet, row, c, None, fill_color=color)
        sheet.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
        sheet.row_dimensions[row].height = 32

    def _write_headers(self, sheet, row, cs):
        for offset, hdr, gray in (
            (self._OFF_PCT, "% entry", False),
            (self._OFF_N, "N", False),
            (self._OFF_DELTA, "\u0394", True),
            (self._OFF_SRC, "% source", True),
        ):
            self._write_cell(
                sheet,
                row,
                cs + offset,
                hdr,
                bold=True,
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT if gray else None,
            )

    def _write_data(self, sheet, start_row, cs, rows, expanded, group_color):
        prev_rem = {}
        prev_val = None
        for i, r in enumerate(rows):
            prev_rem[i] = prev_val
            rtype = str(r.get("Type", "")).lower()
            if i == 0 and rtype == "info":
                prev_val = r.get("N")
            else:
                rem = r.get("Remaining")
                if rem is not None:
                    prev_val = rem

        for ei, exp in enumerate(expanded):
            row = start_row + ei
            kind = exp["kind"]

            if kind == "spacer":
                continue
            if kind == "db":
                self._write_db_row(sheet, row, cs, rows[exp["idx"]])
            elif kind == "arrow":
                self._write_arrow_row(sheet, row, cs, rows, exp, prev_rem)
            elif kind == "final":
                self._write_criterion_row(
                    sheet, row, cs, rows[exp["idx"]], True, group_color
                )
            else:
                self._write_criterion_row(sheet, row, cs, rows[exp["idx"]], False, None)

    def _write_db_row(self, sheet, row, cs, dr):
        db_n = dr.get("N")
        if db_n is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_N,
                self._clean_numeric(db_n),
                size=14,
                horizontal="center",
                number_format=self._number_format_for_value(db_n),
            )
        self._write_cell(
            sheet,
            row,
            cs + self._OFF_SRC,
            100,
            size=14,
            horizontal="center",
            font_color=self._GRAY_TEXT,
        )

    def _write_arrow_row(self, sheet, row, cs, rows, exp, prev_rem):
        tl = exp.get("type_label")
        if tl:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_TYPE,
                tl,
                italic=True,
                size=14,
                horizontal="right",
            )
        self._write_cell(
            sheet,
            row,
            cs + self._OFF_N,
            self._ARROW,
            size=14,
            horizontal="center",
        )
        di = exp["idx"]
        rem = rows[di].get("Remaining")
        prev = prev_rem.get(di)
        if rem is not None and prev is not None:
            delta = rem - prev
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_DELTA,
                self._clean_numeric(delta),
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT,
                number_format=self._number_format_for_value(delta),
            )

    def _write_criterion_row(self, sheet, row, cs, dr, is_final, fill_color):
        if not is_final:
            idx_val = dr.get("Index")
            if idx_val:
                self._write_cell(
                    sheet,
                    row,
                    cs + self._OFF_INDEX,
                    idx_val,
                    size=14,
                    horizontal="center",
                    fill_color=fill_color,
                )
        self._write_cell(
            sheet,
            row,
            cs + self._OFF_NAME,
            dr.get("Name"),
            bold=is_final,
            size=14,
            horizontal="left",
            fill_color=fill_color,
        )
        pct_rem = dr.get("Pct_Remaining")
        if pct_rem is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_PCT,
                self._clean_numeric(pct_rem),
                bold=is_final,
                size=14,
                horizontal="center",
                number_format=self._number_format_for_value(pct_rem),
                fill_color=fill_color,
            )
        remaining = dr.get("Remaining")
        if remaining is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_N,
                self._clean_numeric(remaining),
                bold=is_final,
                size=14,
                horizontal="center",
                number_format=self._number_format_for_value(remaining),
                fill_color=fill_color,
            )
        pct_src = dr.get("Pct_Source_Database")
        if pct_src is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_SRC,
                self._clean_numeric(pct_src),
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT,
                number_format=self._number_format_for_value(pct_src),
                fill_color=fill_color,
            )
        # Fill remaining empty cells in final row
        if is_final and fill_color:
            for off in range(self._NUM_CARD_COLS):
                cell = sheet.cell(row=row, column=cs + off)
                if cell.value is None:
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid",
                    )

    def _set_widths(self, sheet, cs):
        for i, w in enumerate(self._CARD_WIDTHS):
            sheet.column_dimensions[get_column_letter(cs + i)].width = w

    def _apply_card_border(self, sheet, top_row, bottom_row, cs, ce, color):
        """Draw a thin border around the entire card."""
        side = Side(style="thin", color=color)
        for r in range(top_row, bottom_row + 1):
            for c in range(cs, ce + 1):
                cell = sheet.cell(row=r, column=c)
                existing = cell.border
                cell.border = Border(
                    left=side if c == cs else existing.left,
                    right=side if c == ce else existing.right,
                    top=side if r == top_row else existing.top,
                    bottom=side if r == bottom_row else existing.bottom,
                )

    def _apply_section_dividers(self, sheet, data_start, cs, ce, expanded, color):
        """Draw a horizontal top border on arrow rows that introduce a new type group."""
        side = Side(style="thin", color=color)
        for ei, exp in enumerate(expanded):
            if exp["kind"] == "arrow" and exp.get("type_label"):
                row = data_start + ei
                for c in range(cs, ce + 1):
                    cell = sheet.cell(row=row, column=c)
                    existing = cell.border
                    cell.border = Border(
                        left=existing.left,
                        right=existing.right,
                        top=side,
                        bottom=existing.bottom,
                    )


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------


class InfoSheetWriter(_BaseSheetWriter):
    """Writes the Info sheet: PhenEx version, cohort sizes, and study description.

    Layout::

        Row 1  : "Executed with PhenEx v<version>"
        Row 2  : blank
        Row 3  : column headers  Cohort | Final N
        Row 4+ : one row per cohort
        blank row
        description lines starting in col B, one row per line;
        Markdown # / ## / ### headers are rendered with larger font.
    """

    _HEADER_SIZES: Dict[int, int] = {1: 20, 2: 16, 3: 14}

    _SHEET_INFO_TABLE = [
        ("Sheet", "Description"),
        (
            "INCLUSION EXCLUSION",
            "Shows how the study entry criterion, inclusion and exclusion criteria result in the final cohort sizes.",
        ),
        ("CHARACTERISTICS", "Characterizes populations at study entry date."),
        ("OUTCOMES", "Characterizes populations after study entry date."),
        (
            "… all",
            "Displays all boolean, numeric and categorical study elements within a single table.",
        ),
        (
            "… boolean",
            "Displays only boolean study elements — the number of patients that have or do not have a given element. "
            "For numeric values, this identifies missingness.",
        ),
        (
            "… numeric",
            "Displays numeric and categorical study elements horizontally — summary statistics for numeric values, N and % for each category.",
        ),
        (
            "… (detailed)",
            "Displays subcomponents of study elements (e.g. if Diabetes is composed of Type 1 and Type 2, counts for "
            "each component are shown). Identical to the non-detailed version when no subcomponents are present.",
        ),
    ]

    def write(
        self,
        sheet,
        cohort_dirs: List[Path],
        study_path: Path,
        description: Optional[str],
        cohort_groups: Optional[List["CohortGroup"]] = None,
    ) -> None:
        self._set_default_row_height(sheet)
        sheet.column_dimensions["A"].width = self._SPACING_SIZE
        sheet.column_dimensions["B"].width = 34
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 80

        current_row = 1

        # Description at the top
        if description:
            for line in description.splitlines():
                text, font_size, bold = self._parse_markdown_line(line)
                if text is not None:
                    self._write_cell(
                        sheet, current_row, 2, text, bold=bold, size=font_size
                    )
                current_row += 1
            current_row += 1  # blank row after description

        # Cohort table header
        self._write_cell(sheet, current_row, 2, "Cohort", bold=True, size=14)
        self._write_cell(
            sheet, current_row, 3, "Final N", bold=True, size=14, horizontal="right"
        )
        current_row += 1

        # Build lookup from dir -> (display_name, color, is_sub)
        dir_to_display: Dict[Path, tuple] = {}
        if cohort_groups:
            for gi, group in enumerate(cohort_groups):
                color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
                for cohort_dir in group.all_dirs:
                    display = group.display_name(cohort_dir)
                    is_sub = cohort_dir != group.main_dir
                    dir_to_display[cohort_dir] = (display, color, is_sub)

        for cohort_dir in cohort_dirs:
            final_n = self._read_final_n(cohort_dir)
            display, color, is_sub = dir_to_display.get(
                cohort_dir, (cohort_dir.name, None, False)
            )
            # Spacer col: no fill
            self._write_cell(
                sheet,
                current_row,
                2,
                display,
                size=14,
                indent=4 if is_sub else 2,
                fill_color=color,
            )
            self._write_cell(
                sheet,
                current_row,
                3,
                final_n,
                size=14,
                horizontal="right",
                fill_color=color,
                number_format="#,##0" if isinstance(final_n, int) else None,
            )
            current_row += 1

        current_row += 1  # blank row

        # Info section
        self._write_cell(
            sheet,
            current_row,
            2,
            "The following sheets allow comparison of all cohorts, subcohorts and stratifications within this study. "
            "Cohorts are arranged side by side for easy comparison.",
            size=14,
        )
        sheet.row_dimensions[current_row].height = 36
        current_row += 2  # blank row

        # Sheet description table — header row
        sheet_col, desc_col = 2, 4
        thin = Side(style="thin")
        header_bottom = Border(bottom=thin)
        row_border = Border(bottom=Side(style="hair"))

        self._write_cell(
            sheet, current_row, sheet_col, "Sheet name", bold=True, size=14
        )
        self._write_cell(
            sheet, current_row, desc_col, "Description", bold=True, size=14
        )
        for col in (sheet_col, sheet_col + 1, desc_col):
            sheet.cell(row=current_row, column=col).border = header_bottom
        current_row += 1

        # col D width in chars ≈ 80 / 1.27 (ratio of 14pt to default 11pt)
        chars_per_line = 63
        line_height_pt = 20  # 14pt font line height

        for sheet_name, sheet_desc in self._SHEET_INFO_TABLE[1:]:
            name_cell = self._write_cell(
                sheet,
                current_row,
                sheet_col,
                sheet_name,
                size=14,
                bold=(not sheet_name.startswith("…")),
            )
            desc_cell = self._write_cell(
                sheet, current_row, desc_col, sheet_desc, size=14
            )
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            name_cell.alignment = Alignment(vertical="top")
            for col in (sheet_col, sheet_col + 1, desc_col):
                sheet.cell(row=current_row, column=col).border = row_border
            lines = max(1, math.ceil(len(sheet_desc) / chars_per_line))
            sheet.row_dimensions[current_row].height = max(24, lines * line_height_pt)
            current_row += 1

        current_row += 1  # blank row

        # PhenEx version at the bottom
        phenex_version = self._read_phenex_version(study_path)
        self._write_cell(
            sheet,
            current_row,
            2,
            f"Executed with PhenEx v{phenex_version}",
            bold=False,
            size=11,
            font_color=self._GRAY_TEXT,
        )

    # ------------------------------------------------------------------

    def _read_phenex_version(self, study_path: Path) -> str:
        info_file = study_path / "info.txt"
        if info_file.exists():
            for line in info_file.read_text().splitlines():
                if line.startswith("PhenEx Version:"):
                    return line.split(":", 1)[1].strip()
        return "unknown"

    def _read_final_n(self, cohort_dir: Path):
        """Return the final cohort size from Waterfall.json ('Remaining' of last row)."""
        json_file = cohort_dir / "Waterfall.json"
        if json_file.exists():
            try:
                data = self._load_json(json_file)
                rows = data.get("rows", [])
                if rows:
                    return rows[-1].get("Remaining", "-")
            except Exception:
                pass
        return "-"

    def _parse_markdown_line(self, line: str):
        """Return (text, font_size, bold) for a single markdown line.

        Returns (None, ...) for blank lines to signal an empty row.
        """
        # ATX headers: up to three levels
        for level in (3, 2, 1):
            prefix = "#" * level + " "
            if line.startswith(prefix):
                return line[len(prefix) :], self._HEADER_SIZES.get(level, 11), True

        # Unordered list
        if line.startswith("- ") or line.startswith("* "):
            return "\u2022 " + line[2:], 14, False

        # Ordered list
        m = re.match(r"^\d+\.\s+(.*)", line)
        if m:
            return line, 14, False

        # Blank line
        if not line.strip():
            return None, 14, False

        # Plain text
        return line, 14, False


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------


class OutputConcatenator:
    """Concatenates per-cohort JSON reports into a single multi-sheet study file.

    Reads ``<cohort_dir>/<reporter_type>.json`` files produced by
    ``Cohort.write_reports_to_json()`` and assembles them into
    ``study_results.xlsx``.
    """

    _SHEET_ORDER_PREFIX = [
        "Info",
        "Waterfall",
        "WaterfallDetailed",
        "Table1Boolean",
        "Table1Numeric",
        "Table1",
        "Table1OutcomesNumeric",
        "Table1Outcomes",
        "Table1OutcomesAll",
        "Table1BooleanDetailed",
        "Table1NumericDetailed",
        "Table1Detailed",
        "Table1OutcomesDetailed",
        "Table1OutcomesAllDetailed",
    ]
    _CANONICAL_NAMES = {
        "waterfall": "Waterfall",
        "waterfall_detailed": "WaterfallDetailed",
        "table1": "Table1",
        "table1_detailed": "Table1Detailed",
        "table1_outcomes": "Table1Outcomes",
        "table1_outcomes_detailed": "Table1OutcomesDetailed",
    }
    _SHEET_DISPLAY_NAMES = {
        "Info": "OVERVIEW",
        "Waterfall": "INCLUSION EXCLUSION",
        "WaterfallDetailed": "INCLUSION EXCLUSION (detailed)",
        "Table1": "CHARACTERISTICS all",
        "Table1Detailed": "CHARACTERISTICS all (detailed)",
        "Table1Boolean": "CHARACTERISTICS boolean",
        "Table1BooleanDetailed": "CHARACTERISTICS bool (detailed)",
        "Table1Numeric": "CHARACTERISTICS numeric",
        "Table1NumericDetailed": "CHARACTERISTICS num (detailed)",
        "Table1OutcomesAll": "OUTCOMES all",
        "Table1OutcomesAllDetailed": "OUTCOMES all (detailed)",
        "Table1Outcomes": "OUTCOMES boolean",
        "Table1OutcomesDetailed": "OUTCOMES boolean (detailed)",
        "Table1OutcomesNumeric": "OUTCOMES numeric",
    }

    def __init__(
        self,
        study_execution_path: str,
        study_name: str = "study",
        cohort_names: Optional[List[str]] = None,
        description: Optional[str] = None,
    ) -> None:
        self.study_path = Path(study_execution_path)
        self.cohort_names = cohort_names
        self.description = description
        timestamp = self.study_path.name
        self.output_file = self.study_path / f"results_{study_name}_{timestamp}.xlsx"
        self._info_writer = InfoSheetWriter()
        self._generic_writer = GenericSheetWriter()
        self._table1_writer = Table1SheetWriter()
        self._numeric_writer = Table1NumericSheetWriter()
        self._attrition_writer = GenericSheetWriter()

    # ------------------------------------------------------------------

    def concatenate_all_reports(self) -> None:
        cohort_dirs = self._get_cohort_directories()
        if not cohort_dirs:
            logger.warning(f"No cohort directories found in {self.study_path}")
            return

        reports_by_type = self._group_reports_by_type(cohort_dirs)
        if not reports_by_type:
            logger.warning("No report files found in cohort directories")
            return

        # Add boolean views of Table1 data (same source files, N+Pct only)
        _BOOLEAN_SOURCES = {
            "Table1Boolean": "Table1",
            "Table1BooleanDetailed": "Table1Detailed",
        }
        for bool_type, source_type in _BOOLEAN_SOURCES.items():
            if source_type in reports_by_type:
                reports_by_type[bool_type] = reports_by_type[source_type]

        # Add all-columns outcome views (same source files as boolean outcomes)
        _ALL_SOURCES = {
            "Table1OutcomesAll": "Table1Outcomes",
            "Table1OutcomesAllDetailed": "Table1OutcomesDetailed",
        }
        for all_type, source_type in _ALL_SOURCES.items():
            if source_type in reports_by_type:
                reports_by_type[all_type] = reports_by_type[source_type]

        # Add numeric views of Table1 data (same source files, numeric rows only)
        _NUMERIC_SOURCES = {
            "Table1Numeric": "Table1",
            "Table1NumericDetailed": "Table1Detailed",
            "Table1OutcomesNumeric": "Table1Outcomes",
        }
        for num_type, source_type in _NUMERIC_SOURCES.items():
            if source_type in reports_by_type:
                reports_by_type[num_type] = reports_by_type[source_type]

        cohort_groups = self._group_cohorts(cohort_dirs)

        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)

        # Info sheet is always first
        info_sheet = output_wb.create_sheet(
            title=self._SHEET_DISPLAY_NAMES.get("Info", "Info")
        )
        info_sheet.sheet_view.showGridLines = False
        self._info_writer.write(
            info_sheet, cohort_dirs, self.study_path, self.description, cohort_groups
        )

        for report_type in self._sheet_order(reports_by_type):
            if report_type in self._SANKEY_TYPES:
                continue  # handled separately as HTML
            display_name = self._SHEET_DISPLAY_NAMES.get(report_type, report_type)
            display_name = display_name[:31]  # Excel sheet name limit
            logger.info(f"Concatenating {report_type} reports...")
            sheet = output_wb.create_sheet(title=display_name)
            sheet.sheet_view.showGridLines = False
            self._write_sheet(
                sheet,
                report_type,
                reports_by_type[report_type],
                cohort_dirs,
                cohort_groups,
            )

        output_wb.save(self.output_file)
        self._suppress_number_as_text_warnings(self.output_file)
        logger.info(f"Successfully created: {self.output_file}")

        for report_type in self._SANKEY_TYPES:
            if report_type in reports_by_type:
                self._generate_sankey_html(
                    report_type, reports_by_type[report_type], cohort_dirs
                )

    # ------------------------------------------------------------------

    def _sheet_order(self, reports_by_type: Dict[str, List[Path]]) -> List[str]:
        prefix = [n for n in self._SHEET_ORDER_PREFIX if n in reports_by_type]
        rest = sorted(
            k
            for k in reports_by_type
            if k not in self._SHEET_ORDER_PREFIX and k not in self._SANKEY_TYPES
        )
        return prefix + rest

    def _generate_sankey_html(
        self,
        report_type: str,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
    ) -> None:
        """Generate a combined sankey HTML for all cohorts that have data."""
        from phenex.reporting.treatment_pattern_analysis_sankey import (
            _build_sankey_html,
        )

        all_entries = []
        for cohort_dir, json_file in zip(cohort_dirs, report_files):
            if json_file is None:
                continue
            try:
                with json_file.open() as f:
                    data = json.load(f)
                for entry in data.get("sankey_data", []):
                    labeled = dict(entry)
                    labeled["tpa_name"] = (
                        f"{cohort_dir.name} — {entry.get('tpa_name', '')}"
                    )
                    all_entries.append(labeled)
            except Exception as e:
                logger.warning(f"Could not read sankey data from {json_file}: {e}")

        if not all_entries:
            logger.warning(
                f"No sankey data found for {report_type}; skipping HTML generation."
            )
            return

        html_path = self.output_file.with_name(
            self.output_file.stem + f"_{report_type}.html"
        )
        html_path.write_text(_build_sankey_html(all_entries), encoding="utf-8")
        logger.info(f"Generated sankey HTML: {html_path}")

    _TABLE1_TYPES = {
        "Table1",
        "Table1Detailed",
        "Table1OutcomesAll",
        "Table1OutcomesAllDetailed",
        "Table1Outcomes",
        "Table1OutcomesDetailed",
        "Table1Boolean",
        "Table1BooleanDetailed",
    }
    _TABLE1_BOOLEAN_TYPES = {
        "Table1Boolean",
        "Table1BooleanDetailed",
        "Table1Outcomes",
        "Table1OutcomesDetailed",
    }
    _TABLE1_NUMERIC_TYPES = {
        "Table1Numeric",
        "Table1NumericDetailed",
        "Table1OutcomesNumeric",
    }
    _SANKEY_TYPES = {
        "TreatmentPatternSankey",
    }

    def _write_sheet(
        self,
        sheet,
        report_type: str,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List[CohortGroup],
    ) -> None:
        if report_type == "Waterfall":
            self._attrition_writer.write(
                sheet, report_files, cohort_dirs, cohort_groups
            )
        elif report_type in self._TABLE1_NUMERIC_TYPES:
            self._numeric_writer.write(sheet, report_files, cohort_dirs, cohort_groups)
        elif report_type in self._TABLE1_TYPES:
            boolean_only = report_type in self._TABLE1_BOOLEAN_TYPES
            self._table1_writer.write(
                sheet,
                report_files,
                cohort_dirs,
                cohort_groups,
                boolean_only=boolean_only,
            )
        else:
            self._generic_writer.write(sheet, report_files, cohort_dirs, cohort_groups)

    def _get_cohort_directories(self) -> List[Path]:
        dirs = {
            d.name: d
            for d in self.study_path.iterdir()
            if d.is_dir() and not d.name.startswith(".")
        }
        if self.cohort_names:
            return [dirs[name] for name in self.cohort_names if name in dirs]
        return sorted(dirs.values(), key=lambda x: x.name)

    def _group_cohorts(self, cohort_dirs: List[Path]) -> List[CohortGroup]:
        """Group cohort directories into main cohorts with their subcohorts.

        Main cohorts are those whose name does not start with
        ``<other_name>__``.  Subcohorts are matched to their parent by
        the ``parent__child`` naming convention.
        """
        names = [d.name for d in cohort_dirs]
        dir_by_name = {d.name: d for d in cohort_dirs}

        main_names = [
            name
            for name in names
            if not any(
                name.startswith(other + "__") for other in names if other != name
            )
        ]

        groups: List[CohortGroup] = []
        for main_name in main_names:
            subcohort_dirs = [
                dir_by_name[n] for n in names if n.startswith(main_name + "__")
            ]
            groups.append(
                CohortGroup(
                    name=main_name,
                    main_dir=dir_by_name[main_name],
                    subcohort_dirs=subcohort_dirs,
                )
            )
        return groups

    def _group_reports_by_type(
        self, cohort_dirs: List[Path]
    ) -> Dict[str, List[Optional[Path]]]:
        """Group report files by type, aligned to cohort_dirs.

        Returns a dict mapping report type name to a list of Optional[Path] with
        one entry per cohort dir (None when that cohort has no file of that type).
        """
        # {report_type: {cohort_dir: path}}
        type_to_cohort_path: Dict[str, Dict[Path, Path]] = {}
        for cohort_dir in cohort_dirs:
            for json_file in sorted(cohort_dir.glob("*.json")):
                if json_file.name.startswith("frozen_"):
                    continue
                report_type = self._CANONICAL_NAMES.get(
                    json_file.stem.lower(), json_file.stem
                )
                type_to_cohort_path.setdefault(report_type, {})[cohort_dir] = json_file
        # Build aligned lists; None where a cohort has no file for a given type
        return {
            report_type: [cohort_paths.get(d) for d in cohort_dirs]
            for report_type, cohort_paths in type_to_cohort_path.items()
        }

    def _suppress_number_as_text_warnings(self, xlsx_path: Path) -> None:
        """Inject <ignoredErrors> into each sheet XML to suppress green triangles."""
        tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
        try:
            with zipfile.ZipFile(xlsx_path, "r") as zin:
                with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename.startswith(
                            "xl/worksheets/sheet"
                        ) and item.filename.endswith(".xml"):
                            data = self._inject_ignored_errors(data)
                        zout.writestr(item, data)
            tmp_path.replace(xlsx_path)
        except Exception as e:
            logger.warning(f"Could not suppress number-as-text warnings: {e}")
            if tmp_path.exists():
                tmp_path.unlink()

    @staticmethod
    def _inject_ignored_errors(xml_bytes: bytes) -> bytes:
        text = xml_bytes.decode("utf-8")
        if "<ignoredErrors>" in text:
            return xml_bytes
        m = re.search(r'<dimension ref="([^"]+)"', text)
        ref = m.group(1) if m else "A1:XFD1048576"
        snippet = (
            f"<ignoredErrors>"
            f'<ignoredError sqref="{ref}" numberStoredAsText="1"/>'
            f"</ignoredErrors>"
        )
        return text.replace("</worksheet>", snippet + "</worksheet>").encode("utf-8")
