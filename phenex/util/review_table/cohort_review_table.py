"""Excel review table for a single cohort.

Produces one worksheet that mirrors, in spreadsheet form, exactly what the app's
CohortCardViewer shows: the entry criterion, inclusion, exclusion, baseline and
outcome phenotypes, each on its own row, with component phenotypes on the rows
below their parent. The index / type / name columns are frozen (pinned), and
each parameter cell is either filled (parameter not applicable to the phenotype
class), empty (applicable but unset), or a short text description of the value.
"""

import copy
from collections import defaultdict
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import colors, formatters

# Sections in display order, and their title-row labels.
SECTION_ORDER = ["entry", "inclusion", "exclusion", "baseline", "outcome"]
SECTION_TITLES = {
    "entry": "Entry",
    "inclusion": "Inclusion",
    "exclusion": "Exclusion",
    "baseline": "Baseline",
    "outcome": "Outcomes",
}

# Parameter -> phenotype classes it applies to. A parameter cell is "not
# applicable" (and thus filled) when the phenotype's class is not in the list.
# Copied from app/ui/src/assets/phenotype_applicable_parameters.ts.
APPLICABLE_CLASSES: Dict[str, List[str]] = {
    "return_date": [
        "CategoricalPhenotype", "CodelistPhenotype", "EventCountPhenotype",
        "MeasurementChangePhenotype", "MeasurementPhenotype", "ScorePhenotype",
        "ArithmeticPhenotype", "LogicPhenotype",
    ],
    "domain": [
        "AgePhenotype", "CategoricalPhenotype", "CodelistPhenotype",
        "DeathPhenotype", "MeasurementPhenotype", "TimeRangePhenotype",
    ],
    "relative_time_range": [
        "CategoricalPhenotype", "CodelistPhenotype", "DeathPhenotype",
        "EventCountPhenotype", "MeasurementPhenotype", "TimeRangePhenotype",
    ],
    "value_filter": ["AgePhenotype", "EventCountPhenotype", "MeasurementPhenotype"],
    "categorical_filter": [
        "CategoricalPhenotype", "CodelistPhenotype", "MeasurementPhenotype",
    ],
    "date_range": [
        "CategoricalPhenotype", "CodelistPhenotype", "MeasurementPhenotype",
    ],
    "expression": ["ScorePhenotype", "ArithmeticPhenotype", "LogicPhenotype"],
    "codelist": ["CodelistPhenotype", "MeasurementPhenotype"],
    "value_aggregation": ["MeasurementPhenotype"],
}


@dataclass(frozen=True)
class _Column:
    field: str
    header: str
    width: int
    formatter: Callable[[Any], str]
    pinned: bool = False


# Column layout, matching the CohortCardViewer's visible columns. The first
# three are pinned (frozen) exactly like the app's pinned name column.
COLUMNS: List[_Column] = [
    _Column("hierarchical_index", "Index", 8, formatters.format_plain, pinned=True),
    _Column("type", "Type", 12, formatters.format_plain, pinned=True),
    _Column("name", "Name", 34, formatters.format_plain, pinned=True),
    _Column("class_name", "Phenotype", 18, formatters.format_plain),
    _Column("expression", "Expression", 22, formatters.format_expression),
    _Column("domain", "Domain", 18, formatters.format_domain),
    _Column("codelist", "Codelists", 30, formatters.format_codelist),
    _Column("relative_time_range", "Relative time ranges", 28, formatters.format_relative_time_range),
    _Column("value_filter", "Value filters", 20, formatters.format_value_filter),
    _Column("categorical_filter", "Categorical filters", 24, formatters.format_categorical_filter),
    _Column("return_date", "Return Date", 12, formatters.format_plain),
    _Column("date_range", "Date Range", 22, formatters.format_date_range),
    _Column("value_aggregation", "Value Aggregation", 20, formatters.format_plain),
]

_TITLE_ROW = 1
_HEADER_ROW = 2
_DATA_START_ROW = 3
_FONT = "Arial"


class CohortReviewTable:
    """Builds a single-cohort review worksheet mirroring the CohortCardViewer.

    Accepts the app's cohort JSON: either the cohort object itself (with a
    ``phenotypes`` list) or a wrapper containing a ``cohort_data`` key.
    """

    def __init__(self, cohort: Dict[str, Any], name: Optional[str] = None):
        cohort_data = cohort.get("cohort_data", cohort) if isinstance(cohort, dict) else {}
        self._cohort = copy.deepcopy(cohort_data) or {}
        self.name = name or self._cohort.get("name") or "Cohort"
        self._phenotypes = self._prepare(self._cohort.get("phenotypes", []) or [])

    # ------------------------------------------------------------------ public
    def to_excel(self, path: str) -> None:
        """Write the cohort to a new .xlsx file at ``path``."""
        workbook = Workbook()
        self.write_sheet(workbook, sheet_name=self.name, replace_active=True)
        workbook.save(path)

    def write_sheet(
        self,
        workbook: Workbook,
        sheet_name: Optional[str] = None,
        replace_active: bool = False,
    ) -> Worksheet:
        """Add (or reuse the active) worksheet in ``workbook`` and populate it."""
        title = self._sheet_title(sheet_name or self.name, workbook)
        if replace_active:
            sheet = workbook.active
            sheet.title = title
        else:
            sheet = workbook.create_sheet(title=title)

        self._write_title(sheet)
        self._write_header(sheet)
        self._write_rows(sheet)
        self._apply_layout(sheet)
        return sheet

    # ------------------------------------------------------------- data prep
    def _prepare(self, phenotypes: List[dict]) -> List[dict]:
        """Order phenotypes hierarchically and (re)compute type/index metadata."""
        phenotypes = [copy.deepcopy(p) for p in phenotypes]
        for i, p in enumerate(phenotypes):
            p.setdefault("id", f"__pheno_{i}")
        by_id = {p["id"]: p for p in phenotypes}

        for p in phenotypes:
            p["effective_type"] = self._resolve_effective_type(p, by_id)
        self._compute_indices(phenotypes)
        return self._order_hierarchically(phenotypes)

    @staticmethod
    def _resolve_effective_type(phenotype: dict, by_id: Dict[str, dict]) -> str:
        if phenotype.get("type") != "component":
            return phenotype.get("type") or "component"
        seen: set = set()
        current = phenotype
        while current.get("type") == "component":
            parents = current.get("parentIds") or []
            if not parents or parents[0] in seen or parents[0] not in by_id:
                return "component"
            seen.add(parents[0])
            current = by_id[parents[0]]
        return current.get("type") or "component"

    def _compute_indices(self, phenotypes: List[dict]) -> None:
        children = self._components_by_parent(phenotypes)

        def assign_children(parent_id: str, parent_index: str) -> None:
            ordered = sorted(children.get(parent_id, []), key=lambda p: p.get("index", 0))
            for i, child in enumerate(ordered, start=1):
                child["hierarchical_index"] = f"{parent_index}.{i}"
                assign_children(child["id"], child["hierarchical_index"])

        for section in SECTION_ORDER:
            tops = [p for p in phenotypes if p.get("type") == section]
            for i, top in enumerate(tops, start=1):
                index = "e" if section == "entry" else str(i)
                top["hierarchical_index"] = index
                assign_children(top["id"], index)

    def _order_hierarchically(self, phenotypes: List[dict]) -> List[dict]:
        children = self._components_by_parent(phenotypes)

        def descendants(parent_id: str) -> List[dict]:
            ordered = sorted(children.get(parent_id, []), key=lambda p: p.get("index", 0))
            result: List[dict] = []
            for child in ordered:
                result.append(child)
                result.extend(descendants(child["id"]))
            return result

        ordered: List[dict] = []
        added: set = set()
        for section in SECTION_ORDER:
            for top in [p for p in phenotypes if p.get("type") == section]:
                ordered.append(top)
                added.add(top["id"])
                for child in descendants(top["id"]):
                    ordered.append(child)
                    added.add(child["id"])

        # Orphan components (parent filtered out or missing) go last.
        for p in phenotypes:
            if p["id"] not in added:
                ordered.append(p)
        return ordered

    @staticmethod
    def _components_by_parent(phenotypes: List[dict]) -> Dict[str, List[dict]]:
        by_parent: Dict[str, List[dict]] = defaultdict(list)
        for p in phenotypes:
            if p.get("type") == "component":
                parents = p.get("parentIds") or []
                if parents:
                    by_parent[parents[0]].append(p)
        return by_parent

    # ------------------------------------------------------------- rendering
    def _write_title(self, sheet: Worksheet) -> None:
        cell = self._set(
            sheet, _TITLE_ROW, 1, self.name, bold=True, size=16, fill="D9D9D9"
        )
        sheet.merge_cells(
            start_row=_TITLE_ROW, start_column=1, end_row=_TITLE_ROW, end_column=len(COLUMNS)
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def _write_header(self, sheet: Worksheet) -> None:
        for col_index, column in enumerate(COLUMNS, start=1):
            self._set(
                sheet, _HEADER_ROW, col_index, column.header,
                bold=True, size=10, fill="F0F0F0", wrap=True,
            )

    def _write_rows(self, sheet: Worksheet) -> None:
        row = _DATA_START_ROW
        previous_section: Optional[str] = None
        for phenotype in self._phenotypes:
            section = phenotype.get("effective_type") or phenotype.get("type")
            if section != previous_section and section in SECTION_TITLES:
                self._write_section_title(sheet, row, section)
                row += 1
                previous_section = section
            self._write_phenotype_row(sheet, row, phenotype)
            row += 1

    def _write_section_title(self, sheet: Worksheet, row: int, section: str) -> None:
        self._set(
            sheet, row, 1, SECTION_TITLES[section].upper(),
            bold=True, size=11, font_color=colors.type_text_color(section),
        )
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS)
        )

    def _write_phenotype_row(self, sheet: Worksheet, row: int, phenotype: dict) -> None:
        effective_type = phenotype.get("effective_type")
        hierarchical_index = phenotype.get("hierarchical_index")
        row_fill = colors.row_fill(effective_type, hierarchical_index)
        na_fill = colors.na_fill(effective_type, hierarchical_index)

        for col_index, column in enumerate(COLUMNS, start=1):
            if not column.pinned and self._is_not_applicable(column.field, phenotype):
                self._set(sheet, row, col_index, None, fill=na_fill)
                continue
            text = column.formatter(phenotype.get(column.field))
            self._set(sheet, row, col_index, text or None, fill=row_fill, wrap=True)

    @staticmethod
    def _is_not_applicable(field: str, phenotype: dict) -> bool:
        # The entry criterion defines the index date, so relative time ranges
        # never apply to it (matches RelativeTimeRangeCellRenderer).
        if field == "relative_time_range" and phenotype.get("type") == "entry":
            return True
        applicable = APPLICABLE_CLASSES.get(field)
        if applicable is None:
            return False
        return phenotype.get("class_name") not in applicable

    # --------------------------------------------------------------- layout
    def _apply_layout(self, sheet: Worksheet) -> None:
        for col_index, column in enumerate(COLUMNS, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = column.width
        pinned_count = sum(1 for column in COLUMNS if column.pinned)
        # Freeze the pinned columns and the title + header rows.
        sheet.freeze_panes = (
            f"{get_column_letter(pinned_count + 1)}{_DATA_START_ROW}"
        )

    # --------------------------------------------------------------- helpers
    def _set(
        self,
        sheet: Worksheet,
        row: int,
        col: int,
        value: Any,
        bold: bool = False,
        size: int = 10,
        fill: Optional[str] = None,
        font_color: Optional[str] = None,
        wrap: bool = False,
    ):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = Font(name=_FONT, bold=bold, size=size, color=font_color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
        if fill:
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        return cell

    @staticmethod
    def _sheet_title(name: str, workbook: Workbook) -> str:
        """Return an Excel-safe, unique (<=31 char) worksheet title."""
        invalid = set(r"[]:*?/\\")
        cleaned = "".join(" " if ch in invalid else ch for ch in str(name)).strip()
        cleaned = (cleaned or "Cohort")[:31]
        existing = set(workbook.sheetnames)
        if cleaned not in existing:
            return cleaned
        stem = cleaned[:28]
        for suffix in range(2, 100):
            candidate = f"{stem} {suffix}"
            if candidate not in existing:
                return candidate
        return cleaned[:31]
