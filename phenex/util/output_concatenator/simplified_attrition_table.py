from pathlib import Path
from typing import List, Optional

from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from phenex.util import create_logger
from .base_sheet_writer import _BaseSheetWriter
from .cohort_group import CohortGroup

logger = create_logger(__name__)


class SimplifiedAttritionTable(_BaseSheetWriter):
    """Writes attrition cards side-by-side, one per cohort/subcohort.

    Each card is a compact table with downward arrows between steps.
    Component rows are filtered out.  A coloured border surrounds each
    card and horizontal dividers separate entry/inclusion/exclusion
    sections.

    Columns per card: Type | Index | Name | % entry | N | Δ | % source
    """

    _ARROW = "\u2193"  # ↓
    _NUM_CARD_COLS = 7

    # Column offsets within a card (0-based from card_start)
    _OFF_TYPE = 0
    _OFF_INDEX = 1
    _OFF_NAME = 2
    _OFF_PCT = 3  # Pct_Remaining
    _OFF_N = 4  # Remaining
    _OFF_DELTA = 5
    _OFF_SRC = 6  # Pct_Source_Database

    _CARD_WIDTHS = [14, 4, 28, 10, 14, 14, 12]

    def write(self, sheet, report_files, cohort_dirs, cohort_groups):
        sheet.sheet_format.defaultRowHeight = 28.0
        sheet.sheet_format.customHeight = True
        dir_to_report = dict(zip(cohort_dirs, report_files))

        expanded = self._build_template(report_files)
        if not expanded:
            return

        title_row = 2
        header_row = 3
        data_start = 4
        sheet.row_dimensions[1].height = self._SPACING_SIZE * 5

        # Row heights
        for i, exp in enumerate(expanded):
            row = data_start + i
            kind = exp["kind"]
            if kind == "spacer":
                sheet.row_dimensions[row].height = 8
            elif kind == "arrow":
                sheet.row_dimensions[row].height = 22
            else:
                sheet.row_dimensions[row].height = 28

        last_data_row = data_start + len(expanded) - 1
        sheet.freeze_panes = sheet.cell(row=data_start, column=1)

        current_col = 1
        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            for cohort_dir in group.all_dirs:
                report_file = dir_to_report.get(cohort_dir)
                if report_file is None:
                    continue

                self._apply_spacing(sheet, spacing_col=current_col)
                current_col += 1

                try:
                    data = self._load_json(report_file)
                    rows = self._filter_components(data.get("rows", []))
                    display = group.display_name(cohort_dir)
                    is_sub = cohort_dir != group.main_dir
                    title = f"{group.name} \u00b7 {display}" if is_sub else display

                    cs = current_col  # card start
                    ce = cs + self._NUM_CARD_COLS - 1

                    self._write_title(sheet, title_row, cs, ce, title, group_color)
                    self._write_headers(sheet, header_row, cs)
                    self._write_data(sheet, data_start, cs, rows, expanded, group_color)
                    self._set_widths(sheet, cs)
                    self._apply_card_border(
                        sheet, title_row, last_data_row, cs, ce, group_color
                    )
                    self._apply_section_dividers(
                        sheet, data_start, cs, ce, expanded, group_color
                    )

                    current_col = ce + 1
                except Exception as e:
                    logger.warning(f"Failed to process {report_file}: {e}")

    # ------------------------------------------------------------------

    @staticmethod
    def _filter_components(rows: list) -> list:
        return [r for r in rows if str(r.get("Type", "")).lower() != "component"]

    def _build_template(self, report_files):
        """Build expanded row template from first available waterfall file."""
        for f in report_files:
            if f is None:
                continue
            try:
                data = self._load_json(f)
                rows = self._filter_components(data.get("rows", []))
                if rows:
                    return self._expand_rows(rows)
            except Exception:
                pass
        return []

    def _expand_rows(self, rows):
        """Build display row template with arrows and spacers interleaved."""
        expanded = []
        prev_type = None

        for i, r in enumerate(rows):
            rtype = str(r.get("Type", "")).lower()
            is_first = i == 0
            is_last = i == len(rows) - 1

            # Database info row (first row)
            if is_first and rtype == "info":
                expanded.append({"kind": "db", "idx": i})
                expanded.append({"kind": "spacer"})
                prev_type = rtype
                continue

            # Spacer before a new type group
            if prev_type and prev_type != rtype and prev_type != "info":
                expanded.append({"kind": "spacer"})

            # Final cohort row – no arrow before it
            if is_last and rtype == "info":
                expanded.append({"kind": "final", "idx": i})
                prev_type = rtype
                continue

            # Arrow row
            show_type = rtype != prev_type
            expanded.append(
                {
                    "kind": "arrow",
                    "type_label": rtype if show_type else None,
                    "idx": i,
                }
            )

            # Data row
            expanded.append({"kind": "data", "idx": i})
            prev_type = rtype

        return expanded

    # ------------------------------------------------------------------

    def _write_title(self, sheet, row, cs, ce, title, color):
        self._write_cell(
            sheet,
            row,
            cs,
            title,
            bold=True,
            italic=True,
            size=16,
            horizontal="left",
            indent=1,
            fill_color=color,
        )
        for c in range(cs + 1, ce + 1):
            self._write_cell(sheet, row, c, None, fill_color=color)
        sheet.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
        sheet.row_dimensions[row].height = 32

    def _write_headers(self, sheet, row, cs):
        for offset, hdr, gray in (
            (self._OFF_PCT, "% entry", False),
            (self._OFF_N, "N", False),
            (self._OFF_DELTA, "\u0394", True),
            (self._OFF_SRC, "% source", True),
        ):
            self._write_cell(
                sheet,
                row,
                cs + offset,
                hdr,
                bold=True,
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT if gray else None,
            )

    def _write_data(self, sheet, start_row, cs, rows, expanded, group_color):
        prev_rem = {}
        prev_val = None
        for i, r in enumerate(rows):
            prev_rem[i] = prev_val
            rtype = str(r.get("Type", "")).lower()
            if i == 0 and rtype == "info":
                prev_val = r.get("N")
            else:
                rem = r.get("Remaining")
                if rem is not None:
                    prev_val = rem

        for ei, exp in enumerate(expanded):
            row = start_row + ei
            kind = exp["kind"]

            if kind == "spacer":
                continue
            if kind == "db":
                self._write_db_row(sheet, row, cs, rows[exp["idx"]])
            elif kind == "arrow":
                self._write_arrow_row(sheet, row, cs, rows, exp, prev_rem)
            elif kind == "final":
                self._write_criterion_row(
                    sheet, row, cs, rows[exp["idx"]], True, group_color
                )
            else:
                self._write_criterion_row(sheet, row, cs, rows[exp["idx"]], False, None)

    def _write_db_row(self, sheet, row, cs, dr):
        db_n = dr.get("N")
        if db_n is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_N,
                self._clean_numeric(db_n),
                size=14,
                horizontal="center",
                number_format=self._number_format_for_value(db_n),
            )
        self._write_cell(
            sheet,
            row,
            cs + self._OFF_SRC,
            100,
            size=14,
            horizontal="center",
            font_color=self._GRAY_TEXT,
        )

    def _write_arrow_row(self, sheet, row, cs, rows, exp, prev_rem):
        tl = exp.get("type_label")
        if tl:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_TYPE,
                tl,
                italic=True,
                size=14,
                horizontal="right",
            )
        self._write_cell(
            sheet,
            row,
            cs + self._OFF_N,
            self._ARROW,
            size=14,
            horizontal="center",
        )
        di = exp["idx"]
        rem = rows[di].get("Remaining")
        prev = prev_rem.get(di)
        if rem is not None and prev is not None:
            delta = rem - prev
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_DELTA,
                self._clean_numeric(delta),
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT,
                number_format=self._number_format_for_value(delta),
            )

    def _write_criterion_row(self, sheet, row, cs, dr, is_final, fill_color):
        if not is_final:
            idx_val = dr.get("Index")
            if idx_val:
                self._write_cell(
                    sheet,
                    row,
                    cs + self._OFF_INDEX,
                    idx_val,
                    size=14,
                    horizontal="center",
                    fill_color=fill_color,
                )
        self._write_cell(
            sheet,
            row,
            cs + self._OFF_NAME,
            dr.get("Name"),
            bold=is_final,
            size=14,
            horizontal="left",
            fill_color=fill_color,
        )
        pct_rem = dr.get("Pct_Remaining")
        if pct_rem is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_PCT,
                self._clean_numeric(pct_rem),
                bold=is_final,
                size=14,
                horizontal="center",
                number_format=self._number_format_for_value(pct_rem),
                fill_color=fill_color,
            )
        remaining = dr.get("Remaining")
        if remaining is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_N,
                self._clean_numeric(remaining),
                bold=is_final,
                size=14,
                horizontal="center",
                number_format=self._number_format_for_value(remaining),
                fill_color=fill_color,
            )
        pct_src = dr.get("Pct_Source_Database")
        if pct_src is not None:
            self._write_cell(
                sheet,
                row,
                cs + self._OFF_SRC,
                self._clean_numeric(pct_src),
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT,
                number_format=self._number_format_for_value(pct_src),
                fill_color=fill_color,
            )
        # Fill remaining empty cells in final row
        if is_final and fill_color:
            for off in range(self._NUM_CARD_COLS):
                cell = sheet.cell(row=row, column=cs + off)
                if cell.value is None:
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid",
                    )

    def _set_widths(self, sheet, cs):
        for i, w in enumerate(self._CARD_WIDTHS):
            sheet.column_dimensions[get_column_letter(cs + i)].width = w

    def _apply_card_border(self, sheet, top_row, bottom_row, cs, ce, color):
        """Draw a thin border around the entire card."""
        side = Side(style="thin", color=color)
        for r in range(top_row, bottom_row + 1):
            for c in range(cs, ce + 1):
                cell = sheet.cell(row=r, column=c)
                existing = cell.border
                cell.border = Border(
                    left=side if c == cs else existing.left,
                    right=side if c == ce else existing.right,
                    top=side if r == top_row else existing.top,
                    bottom=side if r == bottom_row else existing.bottom,
                )

    def _apply_section_dividers(self, sheet, data_start, cs, ce, expanded, color):
        """Draw a horizontal top border on arrow rows that introduce a new type group."""
        side = Side(style="thin", color=color)
        for ei, exp in enumerate(expanded):
            if exp["kind"] == "arrow" and exp.get("type_label"):
                row = data_start + ei
                for c in range(cs, ce + 1):
                    cell = sheet.cell(row=row, column=c)
                    existing = cell.border
                    cell.border = Border(
                        left=existing.left,
                        right=existing.right,
                        top=side,
                        bottom=existing.bottom,
                    )
