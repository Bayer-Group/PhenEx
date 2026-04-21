import json
import re
import zipfile
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from phenex.util import create_logger
from .cohort_group import CohortGroup
from .generic_sheet_writer import GenericSheetWriter
from .info_sheet_writer import InfoSheetWriter
from .table1_numeric_sheet_writer import Table1NumericSheetWriter
from .table1_sheet_writer import Table1SheetWriter
from ._tte_js import _TTE_JS

logger = create_logger(__name__)


class OutputConcatenator:
    """Concatenates per-cohort JSON reports into a single multi-sheet study file.

    Reads ``<cohort_dir>/<reporter_type>.json`` files produced by
    ``Cohort.write_reports_to_json()`` and assembles them into
    ``study_results.xlsx``.
    """

    _SHEET_ORDER_PREFIX = [
        "Info",
        "Waterfall",
        "WaterfallDetailed",
        "Table1Boolean",
        "Table1Numeric",
        "Table1",
        "Table1OutcomesNumeric",
        "Table1Outcomes",
        "Table1OutcomesAll",
        "Table1BooleanDetailed",
        "Table1NumericDetailed",
        "Table1Detailed",
        "Table1OutcomesDetailed",
        "Table1OutcomesAllDetailed",
    ]
    _CANONICAL_NAMES = {
        "waterfall": "Waterfall",
        "waterfall_detailed": "WaterfallDetailed",
        "table1": "Table1",
        "table1_detailed": "Table1Detailed",
        "table1_outcomes": "Table1Outcomes",
        "table1_outcomes_detailed": "Table1OutcomesDetailed",
    }
    _SHEET_DISPLAY_NAMES = {
        "Info": "OVERVIEW",
        "Waterfall": "INCLUSION EXCLUSION",
        "WaterfallDetailed": "INCLUSION EXCLUSION (detailed)",
        "Table1": "CHARACTERISTICS all",
        "Table1Detailed": "CHARACTERISTICS all (detailed)",
        "Table1Boolean": "CHARACTERISTICS boolean",
        "Table1BooleanDetailed": "CHARACTERISTICS bool (detailed)",
        "Table1Numeric": "CHARACTERISTICS numeric",
        "Table1NumericDetailed": "CHARACTERISTICS num (detailed)",
        "Table1OutcomesAll": "OUTCOMES all",
        "Table1OutcomesAllDetailed": "OUTCOMES all (detailed)",
        "Table1Outcomes": "OUTCOMES boolean",
        "Table1OutcomesDetailed": "OUTCOMES boolean (detailed)",
        "Table1OutcomesNumeric": "OUTCOMES numeric",
    }

    def __init__(
        self,
        study_execution_path: str,
        study_name: str = "study",
        cohort_names: Optional[List[str]] = None,
        description: Optional[str] = None,
    ) -> None:
        self.study_path = Path(study_execution_path)
        self.cohort_names = cohort_names
        self.description = description
        timestamp = self.study_path.name
        self.output_file = self.study_path / f"results_{study_name}_{timestamp}.xlsx"
        self._info_writer = InfoSheetWriter()
        self._generic_writer = GenericSheetWriter()
        self._table1_writer = Table1SheetWriter()
        self._numeric_writer = Table1NumericSheetWriter()
        self._attrition_writer = GenericSheetWriter()

    # ------------------------------------------------------------------

    def concatenate_all_reports(self) -> None:
        cohort_dirs = self._get_cohort_directories()
        if not cohort_dirs:
            logger.warning(f"No cohort directories found in {self.study_path}")
            return

        reports_by_type = self._group_reports_by_type(cohort_dirs)
        if not reports_by_type:
            logger.warning("No report files found in cohort directories")
            return

        # Add boolean views of Table1 data (same source files, N+Pct only)
        _BOOLEAN_SOURCES = {
            "Table1Boolean": "Table1",
            "Table1BooleanDetailed": "Table1Detailed",
        }
        for bool_type, source_type in _BOOLEAN_SOURCES.items():
            if source_type in reports_by_type:
                reports_by_type[bool_type] = reports_by_type[source_type]

        # Add all-columns outcome views (same source files as boolean outcomes)
        _ALL_SOURCES = {
            "Table1OutcomesAll": "Table1Outcomes",
            "Table1OutcomesAllDetailed": "Table1OutcomesDetailed",
        }
        for all_type, source_type in _ALL_SOURCES.items():
            if source_type in reports_by_type:
                reports_by_type[all_type] = reports_by_type[source_type]

        # Add numeric views of Table1 data (same source files, numeric rows only)
        _NUMERIC_SOURCES = {
            "Table1Numeric": "Table1",
            "Table1NumericDetailed": "Table1Detailed",
            "Table1OutcomesNumeric": "Table1Outcomes",
        }
        for num_type, source_type in _NUMERIC_SOURCES.items():
            if source_type in reports_by_type:
                reports_by_type[num_type] = reports_by_type[source_type]

        cohort_groups = self._group_cohorts(cohort_dirs)

        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)

        # Info sheet is always first
        info_sheet = output_wb.create_sheet(
            title=self._SHEET_DISPLAY_NAMES.get("Info", "Info")
        )
        info_sheet.sheet_view.showGridLines = False
        self._info_writer.write(
            info_sheet, cohort_dirs, self.study_path, self.description, cohort_groups
        )

        for report_type in self._sheet_order(reports_by_type):
            if report_type in self._SANKEY_TYPES or report_type in self._TTE_TYPES:
                continue  # handled separately as HTML
            display_name = self._SHEET_DISPLAY_NAMES.get(report_type, report_type)
            display_name = display_name[:31]  # Excel sheet name limit
            logger.info(f"Concatenating {report_type} reports...")
            sheet = output_wb.create_sheet(title=display_name)
            sheet.sheet_view.showGridLines = False
            self._write_sheet(
                sheet,
                report_type,
                reports_by_type[report_type],
                cohort_dirs,
                cohort_groups,
            )

        output_wb.save(self.output_file)
        self._suppress_number_as_text_warnings(self.output_file)
        logger.info(f"Successfully created: {self.output_file}")

        for report_type in self._SANKEY_TYPES:
            if report_type in reports_by_type:
                self._generate_sankey_html(
                    report_type, reports_by_type[report_type], cohort_dirs
                )

        for report_type in self._TTE_TYPES:
            if report_type in reports_by_type:
                self._generate_tte_html(
                    report_type, reports_by_type[report_type], cohort_dirs
                )

    # ------------------------------------------------------------------

    def _sheet_order(self, reports_by_type: Dict[str, List[Path]]) -> List[str]:
        prefix = [n for n in self._SHEET_ORDER_PREFIX if n in reports_by_type]
        rest = sorted(
            k
            for k in reports_by_type
            if k not in self._SHEET_ORDER_PREFIX
            and k not in self._SANKEY_TYPES
            and k not in self._TTE_TYPES
        )
        return prefix + rest

    def _generate_sankey_html(
        self,
        report_type: str,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
    ) -> None:
        """Generate a combined sankey HTML for all cohorts that have data."""
        from phenex.reporting.treatment_pattern_analysis_sankey import (
            _build_sankey_html,
        )

        all_entries = []
        for cohort_dir, json_file in zip(cohort_dirs, report_files):
            if json_file is None:
                continue
            try:
                with json_file.open() as f:
                    data = json.load(f)
                for entry in data.get("sankey_data", []):
                    labeled = dict(entry)
                    labeled["tpa_name"] = (
                        f"{cohort_dir.name} — {entry.get('tpa_name', '')}"
                    )
                    all_entries.append(labeled)
            except Exception as e:
                logger.warning(f"Could not read sankey data from {json_file}: {e}")

        if not all_entries:
            logger.warning(
                f"No sankey data found for {report_type}; skipping HTML generation."
            )
            return

        version = self._info_writer._read_phenex_version(self.study_path)

        html_path = self.output_file.with_name(
            self.output_file.stem + f"_{report_type}.html"
        )
        html_path.write_text(
            _build_sankey_html(all_entries, version=version), encoding="utf-8"
        )
        logger.info(f"Generated sankey HTML: {html_path}")

    def _generate_tte_html(
        self,
        report_type: str,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
    ) -> None:
        """Generate a combined time-to-event KM curves HTML for all cohorts."""
        all_cohort_data = []
        for cohort_dir, json_file in zip(cohort_dirs, report_files):
            if json_file is None:
                continue
            try:
                with json_file.open() as f:
                    data = json.load(f)
                rows = data.get("rows", [])
                if rows:
                    all_cohort_data.append(
                        {"cohort_name": cohort_dir.name, "rows": rows}
                    )
            except Exception as e:
                logger.warning(f"Could not read TTE data from {json_file}: {e}")

        if not all_cohort_data:
            logger.warning(
                f"No time-to-event data found for {report_type}; skipping HTML generation."
            )
            return

        version = self._info_writer._read_phenex_version(self.study_path)
        html = self._build_tte_html(all_cohort_data, version=version)

        html_path = self.output_file.with_name(
            self.output_file.stem + f"_{report_type}.html"
        )
        html_path.write_text(html, encoding="utf-8")
        logger.info(f"Generated time-to-event HTML: {html_path}")

    @staticmethod
    def _build_tte_html(all_cohort_data: List[dict], version: str = "unknown") -> str:
        """Build interactive HTML with multi-select cohort dropdown and KM curves."""
        import base64
        from html import escape

        icon_path = (
            Path(__file__).resolve().parent.parent.parent
            / "docs"
            / "assets"
            / "bird_icon.png"
        )
        if icon_path.exists():
            icon_b64 = base64.b64encode(icon_path.read_bytes()).decode("ascii")
            icon_data_uri = f"data:image/png;base64,{icon_b64}"
        else:
            icon_data_uri = ""

        version_escaped = escape(version)

        if icon_data_uri:
            footer_html = (
                f'<div class="phenex-footer">'
                f'<img src="{icon_data_uri}" alt="PhenEx">'
                f"<span>Generated with PhenEx v{version_escaped}</span></div>"
            )
        else:
            footer_html = (
                f'<div class="phenex-footer">'
                f"<span>Generated with PhenEx v{version_escaped}</span></div>"
            )

        data_json = json.dumps(all_cohort_data, default=str)

        # The JS is kept as a plain string to avoid f-string brace conflicts
        return (
            '<!DOCTYPE html>\n<html lang="en">\n<head><meta charset="UTF-8">\n'
            "<title>Time to Event \u2014 Kaplan-Meier Curves</title>\n"
            "<style>\n"
            "body{font-family:Arial,sans-serif;background:#fff;margin:0;"
            "padding:20px 20px 60px 20px}\n"
            ".controls{margin-bottom:24px;display:flex;flex-wrap:wrap;gap:8px;"
            "align-items:center}\n"
            ".controls label{font-size:13px;font-weight:bold;color:#555;"
            "margin-right:4px}\n"
            ".cohort-btn{padding:5px 14px;border-radius:16px;border:2px solid #ccc;"
            "background:#fff;font-size:12px;cursor:pointer;transition:all .15s}\n"
            ".cohort-btn.active{color:#fff}\n"
            ".outcome-section{margin-bottom:40px}\n"
            ".outcome-title{font-size:16px;font-weight:bold;color:#333;"
            "margin:0 0 8px 0}\n"
            ".risk-table{border-collapse:collapse;font-size:11px;margin-top:2px}\n"
            ".risk-table td{padding:1px 0;text-align:center;min-width:50px}\n"
            ".risk-table .label-cell{text-align:right;padding-right:8px;"
            "font-weight:bold;white-space:nowrap}\n"
            ".phenex-footer{position:fixed;bottom:0;left:0;padding:10px 16px;"
            "display:flex;align-items:center;gap:8px;"
            "background:rgba(255,255,255,0.9);z-index:9999}\n"
            ".phenex-footer img{height:24px;width:auto}\n"
            ".phenex-footer span{font-size:11px;color:#999}\n"
            "</style>\n"
            "</head>\n<body>\n"
            '<h1 style="margin-bottom:8px">Kaplan-Meier Survival Curves</h1>\n'
            '<div class="controls" id="controls">'
            "<label>Cohorts:</label></div>\n"
            '<div id="charts"></div>\n' + footer_html + "\n<script>\n"
            "var DATA = "
            + data_json
            + ";\n"
            + _TTE_JS
            + "\n</script>\n</body>\n</html>"
        )

    _TABLE1_TYPES = {
        "Table1",
        "Table1Detailed",
        "Table1OutcomesAll",
        "Table1OutcomesAllDetailed",
        "Table1Outcomes",
        "Table1OutcomesDetailed",
        "Table1Boolean",
        "Table1BooleanDetailed",
    }
    _TABLE1_BOOLEAN_TYPES = {
        "Table1Boolean",
        "Table1BooleanDetailed",
        "Table1Outcomes",
        "Table1OutcomesDetailed",
    }
    _TABLE1_NUMERIC_TYPES = {
        "Table1Numeric",
        "Table1NumericDetailed",
        "Table1OutcomesNumeric",
    }
    _SANKEY_TYPES = {
        "TreatmentPatternSankey",
    }
    _TTE_TYPES = {
        "TimeToEvent",
    }

    def _write_sheet(
        self,
        sheet,
        report_type: str,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List[CohortGroup],
    ) -> None:
        if report_type == "Waterfall":
            self._attrition_writer.write(
                sheet, report_files, cohort_dirs, cohort_groups
            )
        elif report_type in self._TABLE1_NUMERIC_TYPES:
            self._numeric_writer.write(sheet, report_files, cohort_dirs, cohort_groups)
        elif report_type in self._TABLE1_TYPES:
            boolean_only = report_type in self._TABLE1_BOOLEAN_TYPES
            self._table1_writer.write(
                sheet,
                report_files,
                cohort_dirs,
                cohort_groups,
                boolean_only=boolean_only,
            )
        else:
            self._generic_writer.write(sheet, report_files, cohort_dirs, cohort_groups)

    def _get_cohort_directories(self) -> List[Path]:
        dirs = {
            d.name: d
            for d in self.study_path.iterdir()
            if d.is_dir() and not d.name.startswith(".")
        }
        if self.cohort_names:
            return [dirs[name] for name in self.cohort_names if name in dirs]
        return sorted(dirs.values(), key=lambda x: x.name)

    def _group_cohorts(self, cohort_dirs: List[Path]) -> List[CohortGroup]:
        """Group cohort directories into main cohorts with their subcohorts.

        Main cohorts are those whose name does not start with
        ``<other_name>__``.  Subcohorts are matched to their parent by
        the ``parent__child`` naming convention.
        """
        names = [d.name for d in cohort_dirs]
        dir_by_name = {d.name: d for d in cohort_dirs}

        main_names = [
            name
            for name in names
            if not any(
                name.startswith(other + "__") for other in names if other != name
            )
        ]

        groups: List[CohortGroup] = []
        for main_name in main_names:
            subcohort_dirs = [
                dir_by_name[n] for n in names if n.startswith(main_name + "__")
            ]
            groups.append(
                CohortGroup(
                    name=main_name,
                    main_dir=dir_by_name[main_name],
                    subcohort_dirs=subcohort_dirs,
                )
            )
        return groups

    def _group_reports_by_type(
        self, cohort_dirs: List[Path]
    ) -> Dict[str, List[Optional[Path]]]:
        """Group report files by type, aligned to cohort_dirs.

        Returns a dict mapping report type name to a list of Optional[Path] with
        one entry per cohort dir (None when that cohort has no file of that type).
        """
        # {report_type: {cohort_dir: path}}
        type_to_cohort_path: Dict[str, Dict[Path, Path]] = {}
        for cohort_dir in cohort_dirs:
            for json_file in sorted(cohort_dir.glob("*.json")):
                if json_file.name.startswith("frozen_"):
                    continue
                report_type = self._CANONICAL_NAMES.get(
                    json_file.stem.lower(), json_file.stem
                )
                type_to_cohort_path.setdefault(report_type, {})[cohort_dir] = json_file
        # Build aligned lists; None where a cohort has no file for a given type
        return {
            report_type: [cohort_paths.get(d) for d in cohort_dirs]
            for report_type, cohort_paths in type_to_cohort_path.items()
        }

    def _suppress_number_as_text_warnings(self, xlsx_path: Path) -> None:
        """Inject <ignoredErrors> into each sheet XML to suppress green triangles."""
        tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
        try:
            with zipfile.ZipFile(xlsx_path, "r") as zin:
                with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename.startswith(
                            "xl/worksheets/sheet"
                        ) and item.filename.endswith(".xml"):
                            data = self._inject_ignored_errors(data)
                        zout.writestr(item, data)
            tmp_path.replace(xlsx_path)
        except Exception as e:
            logger.warning(f"Could not suppress number-as-text warnings: {e}")
            if tmp_path.exists():
                tmp_path.unlink()

    @staticmethod
    def _inject_ignored_errors(xml_bytes: bytes) -> bytes:
        text = xml_bytes.decode("utf-8")
        if "<ignoredErrors>" in text:
            return xml_bytes
        m = re.search(r'<dimension ref="([^"]+)"', text)
        ref = m.group(1) if m else "A1:XFD1048576"
        snippet = (
            f"<ignoredErrors>"
            f'<ignoredError sqref="{ref}" numberStoredAsText="1"/>'
            f"</ignoredErrors>"
        )
        return text.replace("</worksheet>", snippet + "</worksheet>").encode("utf-8")
