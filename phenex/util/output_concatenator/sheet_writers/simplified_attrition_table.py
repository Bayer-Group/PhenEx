from pathlib import Path
from typing import Dict, List, Optional

from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from phenex.util import create_logger
from .base_sheet_writer import _BaseSheetWriter
from ..cohort_group import CohortGroup

logger = create_logger(__name__)


class SimplifiedAttritionTable(_BaseSheetWriter):
    """Writes attrition tables grouped by parent cohort.

    Layout::

        Rows 1-2 : frozen merged headers (source, remaining, Δ, of_entry)
        Row 3+   : cohort groups, each with title rows then data rows

    Per cohort group::

        Spacer row   : empty row between groups
        Title row    : parent cohort name at col G, right-justified
        Subtitle row : "main" / subcohort names at each card's start
        Separator    : colored row per card (15px, matching border extents)
        Data rows    : one row per waterfall entry, height = max across all cohorts
        Final row    : pinned to bottom of bordered box

    Main cohort columns (frozen A-G)::

        A: spacer | B: Type | C: Index | D: Name | E: source% | F: remaining% | G: remainingN

    Continuing (unfrozen)::

        H: Δ | I: of_entry% | J: of_entryN

    Sub-cohort columns::

        spacer | Index | source% | remaining% | remainingN | Δ | of_entry% | of_entryN
    """

    # Font sizes
    _FONT_SIZE_TITLE1 = 18
    _FONT_SIZE_TITLE2 = 14
    _FONT_SIZE_CONTENT = 10

    # Frozen area: columns A-G (1-7)
    _COL_SPACER = 1
    _COL_TYPE = 2
    _COL_INDEX = 3
    _COL_NAME = 4
    _COL_DATA_START = 5  # first data column (source %)
    _FREEZE_COL = 8  # freeze after col G

    # Data column offsets (0-based from a card's data start)
    _DATA_SRC_PCT = 0
    _DATA_REM_PCT = 1
    _DATA_REM_N = 2
    _DATA_DELTA = 3
    _DATA_ENTRY_PCT = 4
    _DATA_ENTRY_N = 5
    _NUM_DATA_COLS = 6

    # Header rows (frozen)
    _HEADER_ROW_1 = 1  # merged category names
    _HEADER_ROW_2 = 2  # sub-labels
    _FREEZE_ROW = 3  # first unfrozen row

    # Main cohort column widths: spacer, Type, Index, Name, src%, rem%, remN, Δ, entry%, entryN
    _MAIN_WIDTHS = [3, 14, 4, 28, 10, 10, 14, 14, 10, 14]
    # Sub-cohort column widths: spacer, Index, Name, src%, rem%, remN, Δ, entry%, entryN
    _SUB_WIDTHS = [3, 4, 28, 10, 10, 14, 14, 10, 14]

    _SEPARATOR_HEIGHT = 15
    _TITLE_ROW_HEIGHT = 28
    _SPACER_ROW_HEIGHT = 40

    _TYPE_COLORS = {
        "entry": "27607C",
        "inclusion": "1A4225",
        "exclusion": "C22D4E",
    }
    _TEXT_COLOR_SUBCOHORT = "B0B0B0"

    # Header definitions: (label_row1, merge_count, sub_labels)
    # sub_labels: (text, horizontal_alignment)
    _HEADER_GROUPS = [
        ("source", 1, [("%", "right")]),
        ("remaining", 2, [("%", "right"), ("N", "left")]),
        ("", 1, [("\u0394", "right")]),
        ("of_entry", 2, [("%", "right"), ("N", "left")]),
    ]

    def write(self, sheet, report_files, cohort_dirs, cohort_groups):
        sheet.sheet_format.defaultRowHeight = 28.0
        sheet.sheet_format.customHeight = True
        dir_to_report = dict(zip(cohort_dirs, report_files))

        # Freeze panes at H3 (cols A-G and rows 1-2 frozen)
        sheet.freeze_panes = sheet.cell(row=self._FREEZE_ROW, column=self._FREEZE_COL)

        # Write frozen column headers for main cohort data cols
        self._write_header_block(sheet, self._COL_DATA_START)

        # Set main cohort column widths (A through J)
        for i, w in enumerate(self._MAIN_WIDTHS):
            sheet.column_dimensions[get_column_letter(self._COL_SPACER + i)].width = w

        # Determine max number of sub-cohorts across all groups (shared column slots)
        current_row = self._FREEZE_ROW
        max_subs = max(
            (
                sum(1 for cd in g.subcohort_dirs if dir_to_report.get(cd) is not None)
                for g in cohort_groups
            ),
            default=0,
        )

        # Set up shared sub-cohort column slots
        sub_slot_starts: List[int] = []
        next_sub_col = self._COL_SPACER + len(self._MAIN_WIDTHS)
        for slot in range(max_subs):
            sub_slot_starts.append(next_sub_col)
            for i, w in enumerate(self._SUB_WIDTHS):
                sheet.column_dimensions[
                    get_column_letter(next_sub_col + i)
                ].width = w
            data_start_col = next_sub_col + 3  # skip spacer, index, name
            self._write_header_block(sheet, data_start_col)
            next_sub_col += len(self._SUB_WIDTHS)

        # Write each cohort group
        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            main_report = dir_to_report.get(group.main_dir)
            if main_report is None:
                continue

            try:
                main_data = self._load_json(main_report)
            except Exception:
                continue
            main_rows = self._filter_components(main_data.get("rows", []))
            if not main_rows:
                continue

            # Load sub-cohort data (assigned to shared column slots)
            sub_entries: List[Dict] = []
            for cd in group.subcohort_dirs:
                rf = dir_to_report.get(cd)
                if rf is None:
                    continue
                try:
                    data = self._load_json(rf)
                    slot_idx = len(sub_entries)
                    sub_entries.append(
                        {
                            "dir": cd,
                            "name": cd.name.replace(f"{group.name}__", ""),
                            "rows": self._filter_components(data.get("rows", [])),
                            "col_start": sub_slot_starts[slot_idx],
                        }
                    )
                except Exception:
                    continue

            # --- Spacer row above each group ---
            sheet.row_dimensions[current_row].height = self._SPACER_ROW_HEIGHT
            current_row += 1

            # --- Title row: parent cohort name at column G ---
            title_row = current_row
            self._write_cell(
                sheet,
                title_row,
                7,  # column G
                group.name,
                bold=True,
                size=self._FONT_SIZE_TITLE1,
                horizontal="right",
                indent=1,
            )
            sheet.row_dimensions[title_row].height = self._TITLE_ROW_HEIGHT
            current_row += 1

            # --- Subtitle row: cohort names at each card's start ---
            subtitle_row = current_row
            self._write_cell(
                sheet,
                subtitle_row,
                self._COL_TYPE,
                "main",
                bold=True,
                size=self._FONT_SIZE_TITLE2,
                horizontal="left",
            )
            for sub in sub_entries:
                self._write_cell(
                    sheet,
                    subtitle_row,
                    sub["col_start"] + 1,
                    sub["name"],
                    bold=True,
                    size=self._FONT_SIZE_TITLE2,
                    horizontal="left",
                )
            sheet.row_dimensions[subtitle_row].height = self._TITLE_ROW_HEIGHT
            current_row += 1

            # --- Colored separator row (per card, matching border extents) ---
            sep_row = current_row
            # Main card separator
            main_card_start = self._COL_TYPE
            main_card_end = self._COL_DATA_START + self._NUM_DATA_COLS - 1
            for c in range(main_card_start, main_card_end + 1):
                self._write_cell(sheet, sep_row, c, None, fill_color=group_color)
            # Sub-cohort card separators
            for sub in sub_entries:
                sub_card_start = sub["col_start"] + 1
                sub_card_end = sub["col_start"] + len(self._SUB_WIDTHS) - 1
                for c in range(sub_card_start, sub_card_end + 1):
                    self._write_cell(sheet, sep_row, c, None, fill_color=group_color)
            sheet.row_dimensions[sep_row].height = self._SEPARATOR_HEIGHT
            current_row += 1

            # --- Determine box height (max rows across all cohorts) ---
            all_row_counts = [len(main_rows)]
            for sub in sub_entries:
                all_row_counts.append(len(sub["rows"]))
            box_height = max(all_row_counts)

            # --- Data rows ---
            data_start_row = current_row
            last_data_row = data_start_row + box_height - 1

            # Write main cohort rows (final row pinned to bottom)
            self._write_cohort_in_box(
                sheet, data_start_row, last_data_row,
                main_rows, group_color, is_main=True, col_start=None,
            )

            # Write sub-cohort rows (final row pinned to bottom)
            # Rows at positions 0..main_body_count-1 are shared with the parent
            # and should be rendered in light gray.
            main_body_count = len(main_rows) - 1  # exclude final info row
            for sub in sub_entries:
                self._write_cohort_in_box(
                    sheet, data_start_row, last_data_row,
                    sub["rows"], group_color, is_main=False, col_start=sub["col_start"],
                    shared_row_count=main_body_count,
                )

            # --- Borders around each card ---
            self._apply_card_border(
                sheet, data_start_row, last_data_row,
                main_card_start, main_card_end, group_color,
            )
            for sub in sub_entries:
                sub_card_start = sub["col_start"] + 1
                sub_card_end = sub["col_start"] + len(self._SUB_WIDTHS) - 1
                self._apply_card_border(
                    sheet, data_start_row, last_data_row,
                    sub_card_start, sub_card_end, group_color,
                )

            current_row = last_data_row + 1

    # ------------------------------------------------------------------

    def _write_cohort_in_box(
        self, sheet, data_start_row, last_data_row,
        rows, group_color, is_main, col_start,
        shared_row_count=0,
    ):
        """Write cohort rows with final row pinned to bottom of the box."""
        num_rows = len(rows)
        body_rows = rows[:-1] if num_rows > 1 else []
        final_row_data = rows[-1] if num_rows > 0 else None

        # Sparsify type labels across body rows
        prev_type = None

        # Write body rows starting at data_start_row
        for ri, dr in enumerate(body_rows):
            row = data_start_row + ri
            rtype = str(dr.get("Type", "")).lower()
            show_type = rtype != prev_type and rtype not in ("info", "component")
            type_label = rtype if show_type else None
            if rtype not in ("info", "component"):
                prev_type = rtype
            if is_main:
                self._write_main_data_row(sheet, row, dr, type_label, False, group_color)
            else:
                is_shared = ri < shared_row_count
                self._write_sub_data_row(sheet, row, dr, False, col_start, is_shared=is_shared)

        # Write final row at the bottom of the box
        if final_row_data is not None:
            rtype = str(final_row_data.get("Type", "")).lower()
            is_final = rtype == "info"
            if is_main:
                self._write_main_data_row(
                    sheet, last_data_row, final_row_data, None, is_final, group_color
                )
            else:
                self._write_sub_data_row(
                    sheet, last_data_row, final_row_data, is_final, col_start
                )

    # ------------------------------------------------------------------

    def _write_header_block(self, sheet, data_col_start: int):
        """Write the merged row-1 and sub-label row-2 headers for one set of data columns."""
        col = data_col_start
        for label, span, subs in self._HEADER_GROUPS:
            if label:
                self._write_cell(
                    sheet,
                    self._HEADER_ROW_1,
                    col,
                    label,
                    bold=True,
                    size=self._FONT_SIZE_CONTENT,
                    horizontal="center",
                )
                if span > 1:
                    sheet.merge_cells(
                        start_row=self._HEADER_ROW_1,
                        start_column=col,
                        end_row=self._HEADER_ROW_1,
                        end_column=col + span - 1,
                    )
            for i, (sub, align) in enumerate(subs):
                is_gray = sub == "\u0394"
                self._write_cell(
                    sheet,
                    self._HEADER_ROW_2,
                    col + i,
                    sub,
                    bold=not is_gray,
                    size=self._FONT_SIZE_CONTENT,
                    horizontal=align,
                    font_color=self._GRAY_TEXT if is_gray else None,
                )
            col += span

    # ------------------------------------------------------------------

    @staticmethod
    def _filter_components(rows: list) -> list:
        return [r for r in rows if str(r.get("Type", "")).lower() != "component"]

    def _write_main_data_row(self, sheet, row, dr, type_label, is_final, group_color):
        """Write one waterfall row for the main cohort (cols B-J)."""
        data_col = self._COL_DATA_START
        rtype = str(dr.get("Type", "")).lower()
        fc = self._TYPE_COLORS.get(rtype)

        # Type label (sparsified)
        if type_label:
            self._write_cell(
                sheet, row, self._COL_TYPE, type_label,
                italic=True,
                size=self._FONT_SIZE_CONTENT,
                horizontal="right",
                font_color=fc,
            )
        # Index
        idx_val = dr.get("Index")
        if idx_val:
            self._write_cell(
                sheet, row, self._COL_INDEX, idx_val,
                size=self._FONT_SIZE_CONTENT,
                horizontal="center",
                font_color=fc,
            )
        # Name
        self._write_cell(
            sheet, row, self._COL_NAME, dr.get("Name"),
            bold=is_final,
            size=self._FONT_SIZE_CONTENT,
            horizontal="right",
            font_color=fc,
        )
        # Data columns
        self._write_data_cells(sheet, row, data_col, dr, is_final, None, fc)

    def _write_sub_data_row(self, sheet, row, dr, is_final, col_start, is_shared=False):
        """Write one waterfall row for a sub-cohort (Index + Name + data cols)."""
        idx_col = col_start + 1
        name_col = col_start + 2
        data_col = col_start + 3
        rtype = str(dr.get("Type", "")).lower()
        fc = self._TEXT_COLOR_SUBCOHORT if is_shared else self._TYPE_COLORS.get(rtype)

        idx_val = dr.get("Index")
        if idx_val and not is_final:
            self._write_cell(
                sheet, row, idx_col, idx_val,
                size=self._FONT_SIZE_CONTENT,
                horizontal="center",
                font_color=fc,
            )
        self._write_cell(
            sheet, row, name_col, dr.get("Name"),
            bold=is_final,
            size=self._FONT_SIZE_CONTENT,
            horizontal="right",
            font_color=fc,
        )
        self._write_data_cells(sheet, row, data_col, dr, is_final, None, fc)

    def _write_data_cells(self, sheet, row, data_col, dr, is_final, fill_color, font_color=None):
        """Write the 6 data columns for a waterfall row."""
        pct_src = dr.get("Pct_Source_Database")
        if pct_src is not None:
            self._write_cell(
                sheet, row, data_col + self._DATA_SRC_PCT,
                self._clean_numeric(pct_src),
                size=self._FONT_SIZE_CONTENT,
                horizontal="right",
                font_color=font_color or self._GRAY_TEXT,
                number_format=self._number_format_for_value(pct_src),
                fill_color=fill_color,
            )
        pct_rem = dr.get("Pct_Remaining")
        if pct_rem is not None:
            self._write_cell(
                sheet, row, data_col + self._DATA_REM_PCT,
                self._clean_numeric(pct_rem),
                bold=is_final,
                size=self._FONT_SIZE_CONTENT,
                horizontal="right",
                font_color=font_color,
                number_format=self._number_format_for_value(pct_rem),
                fill_color=fill_color,
            )
        remaining = dr.get("Remaining")
        if remaining is not None:
            self._write_cell(
                sheet, row, data_col + self._DATA_REM_N,
                self._clean_numeric(remaining),
                bold=is_final,
                size=self._FONT_SIZE_CONTENT,
                horizontal="left",
                font_color=font_color,
                number_format=self._number_format_for_value(remaining),
                fill_color=fill_color,
            )
        delta = dr.get("Delta")
        if delta is not None:
            self._write_cell(
                sheet, row, data_col + self._DATA_DELTA,
                self._clean_numeric(delta),
                size=self._FONT_SIZE_CONTENT,
                horizontal="right",
                font_color=font_color or self._GRAY_TEXT,
                number_format=self._number_format_for_value(delta),
                fill_color=fill_color,
            )
        pct_n = dr.get("Pct_N")
        if pct_n is not None:
            self._write_cell(
                sheet, row, data_col + self._DATA_ENTRY_PCT,
                self._clean_numeric(pct_n),
                size=self._FONT_SIZE_CONTENT,
                horizontal="right",
                font_color=font_color,
                number_format=self._number_format_for_value(pct_n),
                fill_color=fill_color,
            )
        n_val = dr.get("N")
        if n_val is not None:
            self._write_cell(
                sheet, row, data_col + self._DATA_ENTRY_N,
                self._clean_numeric(n_val),
                size=self._FONT_SIZE_CONTENT,
                horizontal="left",
                font_color=font_color,
                number_format=self._number_format_for_value(n_val),
                fill_color=fill_color,
            )

    # ------------------------------------------------------------------

    def _apply_card_border(self, sheet, top_row, bottom_row, cs, ce, color):
        """Draw a thin border around a card area."""
        side = Side(style="thin", color=color)
        for r in range(top_row, bottom_row + 1):
            for c in range(cs, ce + 1):
                cell = sheet.cell(row=r, column=c)
                existing = cell.border
                cell.border = Border(
                    left=side if c == cs else existing.left,
                    right=side if c == ce else existing.right,
                    top=side if r == top_row else existing.top,
                    bottom=side if r == bottom_row else existing.bottom,
                )
