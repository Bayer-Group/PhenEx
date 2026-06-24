import colorsys
import json
import math
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class _BaseSheetWriter:
    """Low-level cell and sheet helpers shared by all writers."""

    # Row layout constants (1-indexed Excel rows)
    _SPACING_ROW = 1
    _TITLE_ROW = 2
    _SUBTITLE_ROW = 3
    _HEADER_ROW = 4
    _DATA_START_ROW = 5

    _SPACING_SIZE = 3
    _ROW_BACKGROUND_1 = "ededed"
    _SECTION_HEADER_COLOR = "FFFFFF"
    _GRAY_TEXT = "D9D9D9"
    _FONT = "Arial"
    _BOOLEAN_COLUMNS = {"N", "Pct"}
    _COLUMN_DISPLAY_NAMES = {"Pct": "%"}

    _COHORT_COLORS: List[str] = [
        "D6E4F0",  # light blue
        "D9EAD3",  # light green
        "FDE5CC",  # light orange
        "F4CCCC",  # light red
        "D9D2E9",  # light purple
        "D0E0E3",  # light teal
        "FFF2CC",  # light yellow
        "EAD1DC",  # light pink
        "CFE2F3",  # light sky
        "E6D8C3",  # light tan
    ]

    # Row type -> background fill colour (ARGB hex, no leading #)
    _WATERFALL_COLORS: Dict[str, str] = {
        "info": "DCDCDC",
        "entry": "9DC3E6",
        "inclusion": "C5E0B4",
        "exclusion": "F4B8A0",
        "component": "E2EFD9",
    }

    def _write_cell(
        self,
        sheet,
        row: int,
        col: int,
        value,
        bold: bool = False,
        italic: bool = False,
        size: int = 10,
        horizontal: str = "left",
        vertical: str = "center",
        indent: int = 0,
        fill_color: Optional[str] = None,
        number_format: Optional[str] = None,
        wrap_text: bool = False,
        font_color: Optional[str] = None,
    ):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = Font(
            name=self._FONT, bold=bold, italic=italic, size=size, color=font_color
        )
        cell.alignment = Alignment(
            horizontal=horizontal, vertical="center", indent=indent, wrap_text=wrap_text
        )
        if fill_color:
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
        if number_format:
            cell.number_format = number_format
        return cell

    def _add_group_title(
        self, sheet, title: str, start_col: int, end_col: int, color: str = "D3D3D3"
    ):
        """Write main cohort name in the title row, merged across group columns."""
        cell = self._write_cell(
            sheet,
            self._TITLE_ROW,
            start_col,
            title,
            italic=True,
            size=18,
            horizontal="left",
            indent=2,
            fill_color=color,
        )
        cell.border = Border()
        if end_col > start_col:
            sheet.merge_cells(
                start_row=self._TITLE_ROW,
                start_column=start_col,
                end_row=self._TITLE_ROW,
                end_column=end_col,
            )

    def _add_subcohort_header(
        self,
        sheet,
        name: str,
        start_col: int,
        num_cols: int,
        color: str = "E8E8E8",
        font_color: Optional[str] = None,
    ):
        """Write individual cohort/subcohort name in the subtitle row."""
        self._write_cell(
            sheet,
            self._SUBTITLE_ROW,
            start_col,
            name,
            italic=True,
            size=14,
            horizontal="left",
            indent=2,
            fill_color=color,
            font_color=font_color,
        )
        if num_cols > 1:
            sheet.merge_cells(
                start_row=self._SUBTITLE_ROW,
                start_column=start_col,
                end_row=self._SUBTITLE_ROW,
                end_column=start_col + num_cols - 1,
            )

    def _apply_group_border(
        self, sheet, start_col: int, end_col: int, color: str = "000000"
    ):
        """Draw a thin border around the cohort group from title row to last data row."""
        max_row = sheet.max_row
        top_row = self._TITLE_ROW
        side = Side(style="thin", color=color)
        for row in range(top_row, max_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                existing = cell.border
                cell.border = Border(
                    left=side if col == start_col else existing.left,
                    right=side if col == end_col else existing.right,
                    top=side if row == top_row else existing.top,
                    bottom=side if row == max_row else existing.bottom,
                )

    def _apply_spacing(self, sheet, spacing_col: int):
        """Set spacing column width and spacing row height."""
        sheet.column_dimensions[get_column_letter(spacing_col)].width = (
            self._SPACING_SIZE
        )
        sheet.row_dimensions[self._SPACING_ROW].height = self._SPACING_SIZE * 5

    def _apply_left_border_to_column(self, sheet, col: int, color: str = "000000"):
        """Draw a thin left border on every cell in a column."""
        side = Side(style="thin", color=color)
        for row in range(self._TITLE_ROW, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            existing = cell.border
            cell.border = Border(
                left=side,
                right=existing.right,
                top=existing.top,
                bottom=existing.bottom,
            )

    @staticmethod
    def _number_format_for_value(value) -> Optional[str]:
        """Return an Excel number format code suitable for value."""
        if value is None:
            return None
        if isinstance(value, int):
            return "#,##0"
        if isinstance(value, float):
            return "0.0"
        return None

    @staticmethod
    def _clean_numeric(value):
        """Convert whole-number floats to int (e.g. 98.0 -> 98)."""
        if isinstance(value, float) and not math.isnan(value) and value == int(value):
            return int(value)
        return value

    @staticmethod
    def _level_to_gray_hex(level) -> Optional[str]:
        """Return a 6-char ARGB hex fill colour for a component nesting level, or None."""
        try:
            lvl = int(level)
        except (TypeError, ValueError):
            return None
        if lvl <= 0:
            return None
        value = max(235 - 20 * (lvl - 1), 100)
        return f"{value:02X}{value:02X}{value:02X}"

    @staticmethod
    def _cohort_text_colors(hex_color: str) -> Tuple[str, str]:
        """Return (text_dark, text_light) derived from a cohort background hex colour.

        Both share the same hue as *hex_color*.  ``text_dark`` is very dark
        (for prominent text like %), ``text_light`` is medium-dark (for
        secondary text like N).
        """
        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        # text_dark – deep, saturated
        rd, gd, bd = colorsys.hls_to_rgb(h, 0.25, min(s * 1.5, 1.0))
        dark = f"{int(rd * 255):02X}{int(gd * 255):02X}{int(bd * 255):02X}"
        # text_light – medium
        rl, gl, bl = colorsys.hls_to_rgb(h, 0.45, min(s * 1.3, 1.0))
        light = f"{int(rl * 255):02X}{int(gl * 255):02X}{int(bl * 255):02X}"
        return dark, light

    def _set_default_row_height(self, sheet, height: float = 24.0) -> None:
        sheet.sheet_format.defaultRowHeight = height
        sheet.sheet_format.customHeight = True
        sheet.row_dimensions[self._TITLE_ROW].height = 34
        sheet.row_dimensions[self._SUBTITLE_ROW].height = 28

    @staticmethod
    def _load_json(path: Path) -> Dict:
        with path.open() as f:
            return json.load(f)
