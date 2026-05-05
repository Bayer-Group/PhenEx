from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter

from phenex.util import create_logger
from .base_sheet_writer import _BaseSheetWriter
from ..cohort_group import CohortGroup

logger = create_logger(__name__)


class Table1SheetWriter(_BaseSheetWriter):
    """Writes Table1 JSON files into a single aligned worksheet.

    Layout::

        Row 1  : spacing row
        Row 2  : main cohort name banner per group (merged)
        Row 3  : individual cohort/subcohort name per block
        Row 4  : column labels per cohort block
        Row 5+ : characteristic rows; section-header rows interleaved

    Column A is a narrow spacer that mirrors row colours.
    Column B holds the characteristic names.

    Section headers appear as full-width blue-grey rows when the first cohort
    JSON contains a ``sections`` dict mapping section names to characteristic
    display names.
    """

    def write(
        self,
        sheet,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List[CohortGroup],
        boolean_only: bool = False,
    ):
        self._set_default_row_height(sheet)
        master_names, sections, name_to_level = self._build_master_names_and_sections(
            report_files
        )
        if not master_names:
            return

        expanded = self._build_expanded_rows(master_names, sections, name_to_level)
        alt_mask = self._compute_alternate_fills(expanded, name_to_level)
        self._write_name_column(sheet, expanded, name_to_level, alt_mask)

        dir_to_report = dict(zip(cohort_dirs, report_files))
        current_col = 3
        spacing_cols: List[int] = []

        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            text_dark, text_light = self._cohort_text_colors(group_color)
            spacing_cols.append(current_col)
            self._apply_spacing(sheet, spacing_col=current_col)
            current_col += 1
            group_start_col = current_col

            for cohort_dir in group.all_dirs:
                report_file = dir_to_report.get(cohort_dir)
                if report_file is None:
                    continue
                is_subcohort = cohort_dir != group.main_dir
                try:
                    data = self._load_json(report_file)
                    rows = data.get("rows", [])
                    columns = [
                        c
                        for c in (rows[0].keys() if rows else [])
                        if c not in ("Name", "_level")
                    ]
                    if boolean_only:
                        columns = [c for c in columns if c in self._BOOLEAN_COLUMNS]
                    if not columns:
                        continue
                    num_value_cols = len(columns)

                    display_name = group.display_name(cohort_dir)
                    self._add_subcohort_header(
                        sheet,
                        display_name,
                        current_col,
                        num_value_cols,
                        color=group_color,
                        font_color=text_dark,
                    )
                    pct_offset = self._write_column_labels(
                        sheet,
                        columns,
                        current_col,
                        text_dark,
                        text_light,
                    )
                    self._write_value_rows(
                        sheet,
                        rows,
                        columns,
                        expanded,
                        current_col,
                        pct_offset,
                        name_to_level,
                        alt_mask,
                        group_color,
                        text_dark,
                        text_light,
                    )
                    self._set_value_column_widths(sheet, rows, columns, current_col)

                    if is_subcohort:
                        self._apply_left_border_to_column(
                            sheet, current_col, color=group_color
                        )

                    current_col += num_value_cols

                except Exception as e:
                    logger.warning(f"Failed to process {report_file}: {e}")

            group_end_col = current_col - 1
            if group_end_col >= group_start_col:
                self._add_group_title(
                    sheet, group.name, group_start_col, group_end_col, color=group_color
                )
                self._apply_group_border(
                    sheet, group_start_col, group_end_col, color=group_color
                )

        # Apply section-header spans, skipping spacer columns
        self._apply_section_header_spans(
            sheet, expanded, sheet.max_column, spacing_cols
        )

    # ------------------------------------------------------------------

    def _build_master_names_and_sections(
        self, report_files: List[Optional[Path]]
    ) -> Tuple[List[str], Optional[Dict], Dict[str, int]]:
        """Return (master_names, sections, name_to_level) aggregated across all JSON files."""
        master_names: List[str] = []
        existing_names: set = set()
        first_file_done = False
        sections = None
        name_to_level: Dict[str, int] = {}
        for report_file in report_files:
            if report_file is None:
                continue
            try:
                data = self._load_json(report_file)
                if sections is None:
                    sections = data.get("sections")
                for row in data.get("rows", []):
                    name = row.get("Name")
                    if name is not None:
                        name_str = str(name)
                        if not first_file_done:
                            # First file: preserve all names including duplicates
                            master_names.append(name_str)
                            existing_names.add(name_str)
                        elif name_str not in existing_names:
                            # Subsequent files: only add genuinely new names
                            master_names.append(name_str)
                            existing_names.add(name_str)
                        level = row.get("_level", 0)
                        if name_str not in name_to_level and level:
                            name_to_level[name_str] = int(level)
                first_file_done = True
            except Exception as e:
                logger.warning(f"Could not read {report_file}: {e}")
        return master_names, sections, name_to_level

    def _build_expanded_rows(
        self,
        master_names: List[str],
        sections: Optional[Dict],
        name_to_level: Dict[str, int] = None,
    ) -> List[Tuple[str, str]]:
        """Return list of (row_type, value) with section headers interleaved.

        ``row_type`` is ``"section"`` for section-header rows and ``"row"``
        for data rows.
        """
        if not sections:
            return [("row", n) for n in master_names]

        insert_at: Dict[int, str] = {}
        search_start = 0
        for section_name, char_names in sections.items():
            for idx in range(search_start, len(master_names)):
                if self._name_belongs_to_chars(master_names[idx], char_names):
                    insert_at[idx] = section_name
                    # Advance past all rows belonging to this section.
                    # A row belongs if it exactly matches or is a binned
                    # variant (startswith char+"=") of a char_name.
                    # Stop when we hit a row that doesn't belong, OR
                    # when we see an exact char_name duplicate (new TPA
                    # period reusing the same names).
                    seen_exact: set = set()
                    end = idx
                    while end < len(master_names):
                        rname = master_names[end]
                        belongs = False
                        is_exact = False
                        for char in char_names:
                            if rname == char:
                                belongs = True
                                is_exact = True
                                break
                            if rname.startswith(char + "="):
                                belongs = True
                                break
                        if not belongs:
                            break
                        if is_exact and rname in seen_exact:
                            break  # duplicate exact name → next TPA period
                        if is_exact:
                            seen_exact.add(rname)
                        end += 1
                    search_start = end
                    break

        result: List[Tuple[str, str]] = []
        section_start_indices = sorted(insert_at.keys())

        # Build a mapping: master_names index -> which section it belongs to
        current_section_idx = -1
        owner: List[int] = []
        for idx in range(len(master_names)):
            if idx in insert_at:
                current_section_idx = idx
            owner.append(current_section_idx)

        for idx, name in enumerate(master_names):
            if idx in insert_at:
                result.append(("section", insert_at[idx]))
            result.append(("row", name))
            # Add spacer after the last row of this section
            is_last = idx == len(master_names) - 1
            next_in_different_section = not is_last and owner[idx + 1] != owner[idx]
            if owner[idx] >= 0 and (is_last or next_in_different_section):
                result.append(("spacer", ""))
        return self._insert_binned_spacers(result, name_to_level or {})

    @staticmethod
    def _is_binned(name: str, name_to_level: Dict[str, int]) -> bool:
        """A row is binned if it contains '=' and is not a component (level 0)."""
        return "=" in name and name_to_level.get(name, 0) == 0

    @staticmethod
    def _insert_binned_spacers(
        expanded: List[Tuple[str, str]],
        name_to_level: Dict[str, int],
    ) -> List[Tuple[str, str]]:
        """Insert a spacer row before each group of binned rows.

        A spacer is not inserted if the binned group is the first data row
        or immediately follows a section header.
        """
        result: List[Tuple[str, str]] = []
        for i, (row_type, value) in enumerate(expanded):
            if row_type == "row" and "=" in value and name_to_level.get(value, 0) == 0:
                prev_type = result[-1][0] if result else None
                prev_is_binned = (
                    (
                        prev_type == "row"
                        and "=" in result[-1][1]
                        and name_to_level.get(result[-1][1], 0) == 0
                    )
                    if result
                    else False
                )
                if not prev_is_binned and prev_type not in (None, "section"):
                    result.append(("spacer", ""))
            result.append((row_type, value))
        return result

    def _compute_alternate_fills(
        self,
        expanded: List[Tuple[str, str]],
        name_to_level: Dict[str, int],
    ) -> List[bool]:
        """Return per-row boolean mask: True where an alternating fill applies.

        Resets at each section.  The first eligible row after a section
        header gets no fill; subsequent eligible rows alternate.
        Component rows (level > 0) and spacer rows are always False.
        """
        mask: List[bool] = []
        counter = 0
        for row_type, value in expanded:
            if row_type == "section":
                counter = 0
                mask.append(False)
            elif row_type == "spacer":
                mask.append(False)
            else:
                if name_to_level.get(value, 0) > 0:
                    mask.append(False)
                else:
                    mask.append(counter % 2 == 1)
                    counter += 1
        return mask

    @staticmethod
    def _name_belongs_to_chars(name: str, char_names: List[str]) -> bool:
        for char in char_names:
            if name == char or name.startswith(char + "="):
                return True
        return False

    _NAME_SPACER_COL = 1
    _NAME_COL = 2

    def _write_name_column(
        self,
        sheet,
        expanded: List[Tuple[str, str]],
        name_to_level: Dict[str, int] = None,
        alt_mask: List[bool] = None,
    ):
        """Populate col B with characteristic names; col A is a narrow colour spacer."""
        name_to_level = name_to_level or {}
        sheet.column_dimensions[get_column_letter(self._NAME_SPACER_COL)].width = (
            self._SPACING_SIZE
        )

        for i, (row_type, value) in enumerate(expanded):
            out_row = self._DATA_START_ROW + i
            if row_type == "section":
                self._write_cell(
                    sheet,
                    out_row,
                    self._NAME_COL,
                    value,
                    bold=True,
                    size=10,
                    horizontal="left",
                    indent=2,
                    fill_color=self._SECTION_HEADER_COLOR,
                )
                self._write_cell(
                    sheet,
                    out_row,
                    self._NAME_SPACER_COL,
                    None,
                    fill_color=self._SECTION_HEADER_COLOR,
                )
                sheet.row_dimensions[out_row].height = 36
            elif row_type == "spacer":
                sheet.cell(row=out_row, column=self._NAME_COL, value=None)
                sheet.cell(row=out_row, column=self._NAME_SPACER_COL, value=None)
            else:
                is_cohort = value == "Cohort"
                level = name_to_level.get(value, 0)
                gray = self._level_to_gray_hex(level)
                alt = self._ROW_BACKGROUND_1 if (alt_mask and alt_mask[i]) else None
                fill = gray if gray else alt
                display = None if is_cohort else value
                if display and "=" in display and name_to_level.get(value, 0) == 0:
                    self._write_binned_name_cell(
                        sheet,
                        out_row,
                        self._NAME_COL,
                        display,
                        fill_color=fill,
                    )
                else:
                    self._write_cell(
                        sheet,
                        out_row,
                        self._NAME_COL,
                        display,
                        bold=is_cohort,
                        size=10,
                        horizontal="right",
                        indent=1,
                        fill_color=fill,
                    )
                self._write_cell(
                    sheet,
                    out_row,
                    self._NAME_SPACER_COL,
                    None,
                    fill_color=fill,
                )

        max_name_len = max((len(t[1]) for t in expanded if t[0] == "row"), default=10)
        sheet.column_dimensions[get_column_letter(self._NAME_COL)].width = min(
            max(max_name_len * 2.4, 28), 60
        )
        sheet.freeze_panes = sheet.cell(
            row=self._DATA_START_ROW + 1, column=self._NAME_COL + 1
        )

    def _write_binned_name_cell(
        self,
        sheet,
        row: int,
        col: int,
        value: str,
        fill_color: Optional[str] = None,
    ):
        """Write a binned phenotype name with the part after '=' in bold."""
        prefix, suffix = value.split("=", 1)
        rich = CellRichText(
            TextBlock(InlineFont(rFont=self._FONT, sz=10), prefix + "="),
            TextBlock(InlineFont(rFont=self._FONT, sz=10, b=True), suffix),
        )
        cell = sheet.cell(row=row, column=col, value=rich)
        cell.alignment = Alignment(horizontal="right", indent=1)
        if fill_color:
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    def _write_column_labels(
        self,
        sheet,
        columns: List[str],
        start_col: int,
        text_dark: Optional[str] = None,
        text_light: Optional[str] = None,
    ) -> Optional[int]:
        """Write column labels in the header row; return 0-based offset of the Pct column."""
        pct_offset = None
        for offset, col_name in enumerate(columns):
            is_pct = col_name.strip().lower() == "pct"
            is_n = col_name.strip().lower() == "n"
            if is_pct:
                pct_offset = offset
            display_name = self._COLUMN_DISPLAY_NAMES.get(col_name, col_name)
            horizontal = "right" if is_n else ("left" if is_pct else "center")
            fc = (
                text_light
                if is_n and text_light
                else (text_dark if is_pct and text_dark else None)
            )
            cell = self._write_cell(
                sheet,
                self._HEADER_ROW,
                start_col + offset,
                display_name,
                bold=is_pct,
                size=10,
                horizontal=horizontal,
                font_color=fc,
            )
            cell.border = Border()
        return pct_offset

    def _write_value_rows(
        self,
        sheet,
        rows,
        columns: List[str],
        expanded: List[Tuple[str, str]],
        start_col: int,
        pct_offset: Optional[int],
        name_to_level: Dict[str, int] = None,
        alt_mask: List[bool] = None,
        group_color: Optional[str] = None,
        text_dark: Optional[str] = None,
        text_light: Optional[str] = None,
    ):
        name_to_level = name_to_level or {}
        # Build per-name list of rows to handle duplicate names (e.g. TPA periods)
        _name_row_lists: Dict[str, List] = defaultdict(list)
        for r in rows:
            if "Name" in r:
                _name_row_lists[str(r["Name"])].append(r)
        _name_cursors: Dict[str, int] = defaultdict(int)
        for i, (row_type, value) in enumerate(expanded):
            out_row = self._DATA_START_ROW + i
            if row_type == "section":
                for offset in range(len(columns)):
                    self._write_cell(
                        sheet,
                        out_row,
                        start_col + offset,
                        None,
                        fill_color=self._SECTION_HEADER_COLOR,
                    )
                continue
            if row_type == "spacer":
                continue

            cursor = _name_cursors[value]
            entries = _name_row_lists.get(value, [])
            row_data = entries[cursor] if cursor < len(entries) else None
            _name_cursors[value] = cursor + 1
            is_cohort = value == "Cohort"
            level = name_to_level.get(value, 0)
            gray = self._level_to_gray_hex(level)
            should_alt = alt_mask[i] if alt_mask else False
            row_fill = gray if gray else (group_color if should_alt else None)
            for offset, col_name in enumerate(columns):
                raw_value = (
                    self._clean_numeric(row_data.get(col_name)) if row_data else None
                )
                is_pct_col = offset == pct_offset
                is_n_col = col_name.strip().lower() == "n"
                horizontal = (
                    "right" if is_n_col else ("left" if is_pct_col else "center")
                )
                fmt = self._number_format_for_value(raw_value)
                fc = text_light if is_n_col else (text_dark if is_pct_col else None)
                self._write_cell(
                    sheet,
                    out_row,
                    start_col + offset,
                    raw_value,
                    bold=is_pct_col,
                    size=10,
                    horizontal=horizontal,
                    fill_color=row_fill,
                    number_format=fmt,
                    font_color=fc,
                )

    def _apply_section_header_spans(
        self,
        sheet,
        expanded: List[Tuple[str, str]],
        max_col: int,
        spacing_cols: List[int] = None,
    ):
        """Apply fill colour across the full row width for section headers."""
        if max_col < 2:
            return
        skip = set(spacing_cols or [])
        fill = PatternFill(
            start_color=self._SECTION_HEADER_COLOR,
            end_color=self._SECTION_HEADER_COLOR,
            fill_type="solid",
        )
        for i, (row_type, _) in enumerate(expanded):
            if row_type == "section":
                out_row = self._DATA_START_ROW + i
                for col in range(1, max_col + 1):
                    if col in skip:
                        continue
                    cell = sheet.cell(row=out_row, column=col)
                    if (
                        not cell.fill
                        or cell.fill.fgColor is None
                        or cell.fill.fgColor.rgb == "00000000"
                    ):
                        cell.fill = fill

    def _set_value_column_widths(self, sheet, rows, columns: List[str], start_col: int):
        font_scale = 14 / 11
        for offset, col_name in enumerate(columns):
            if col_name.strip().lower() == "n":
                width = 14  # fits "10,000,000"
            else:
                display_name = self._COLUMN_DISPLAY_NAMES.get(col_name, col_name)
                values = [
                    str(r.get(col_name, ""))
                    for r in rows
                    if r.get(col_name) is not None
                ]
                max_val_len = max((len(v) for v in values), default=0)
                content_len = max(len(display_name), max_val_len)
                width = min(max(content_len * font_scale + 1, 6), 18)
            sheet.column_dimensions[get_column_letter(start_col + offset)].width = width
