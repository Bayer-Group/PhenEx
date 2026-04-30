import math
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .base_sheet_writer import _BaseSheetWriter
from ..cohort_group import CohortGroup


class Table1NumericSheetWriter(_BaseSheetWriter):
    """Writes numeric and categorical characteristics in horizontally-stacked blocks.

    Layout::

        Col A  : narrow spacer (no fill)
        Col B  : cohort display names (cohort colour, right-justified, frozen)
        Col C+ : phenotype blocks separated by spacer columns

    Header rows (all frozen)::

        Row 1  : spacing
        Row 2  : phenotype names (large text, merged across block columns)
        Row 3  : stat column labels (numeric) / category names (categorical)
        Row 4  : blank (numeric) / N % sub-headers (categorical)
        Row 5+ : one row per cohort/subcohort
    """

    _NUMERIC_COL_ORDER = [
        "Mean",
        "STD",
        "Min",
        "P10",
        "P25",
        "Median",
        "P75",
        "P90",
        "Max",
        "N",
        "Pct",
    ]
    _GRAYED_STATS = {"STD", "N", "Pct"}

    _SPACER_COL = 1
    _NAME_COL = 2
    _DATA_START_COL = 3

    def write(
        self,
        sheet,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List[CohortGroup],
    ):
        self._set_default_row_height(sheet)
        dir_to_report = dict(zip(cohort_dirs, report_files))

        blocks = self._collect_blocks(report_files)
        if not blocks:
            return

        # Build ordered cohort entries with pre-loaded data
        cohort_entries: List[Dict] = []
        for gi, group in enumerate(cohort_groups):
            gc = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]
            td, tl = self._cohort_text_colors(gc)
            for cd in group.all_dirs:
                rf = dir_to_report.get(cd)
                if rf is None:
                    continue
                try:
                    data = self._load_json(rf)
                except Exception:
                    continue
                cohort_entries.append(
                    {
                        "display_name": group.display_name(cd),
                        "group_color": gc,
                        "text_dark": td,
                        "text_light": tl,
                        "is_subcohort": cd != group.main_dir,
                        "data": data,
                    }
                )

        if not cohort_entries:
            return

        # Fixed columns
        sheet.column_dimensions[get_column_letter(self._SPACER_COL)].width = (
            self._SPACING_SIZE
        )
        sheet.column_dimensions[get_column_letter(self._NAME_COL)].width = 24
        sheet.row_dimensions[self._SPACING_ROW].height = self._SPACING_SIZE * 5
        sheet.row_dimensions[self._TITLE_ROW].height = 32

        # Write cohort name column
        for ri, entry in enumerate(cohort_entries):
            row = self._DATA_START_ROW + ri
            self._write_cell(
                sheet,
                row,
                self._NAME_COL,
                entry["display_name"],
                size=14,
                horizontal="right",
                indent=6 if entry["is_subcohort"] else 4,
                fill_color=entry["group_color"],
            )

        # Write phenotype blocks
        current_col = self._DATA_START_COL
        last_data_row = self._DATA_START_ROW + len(cohort_entries) - 1
        for block in blocks:
            self._apply_spacing(sheet, spacing_col=current_col)
            current_col += 1
            block_start = current_col
            if block["kind"] == "numeric":
                current_col = self._write_numeric_block(
                    sheet,
                    current_col,
                    block,
                    cohort_entries,
                )
            else:
                current_col = self._write_categorical_block(
                    sheet,
                    current_col,
                    block,
                    cohort_entries,
                )
            self._apply_block_border(
                sheet,
                self._TITLE_ROW,
                last_data_row,
                block_start,
                current_col - 1,
            )

        # Freeze header rows + cohort name columns
        sheet.freeze_panes = sheet.cell(
            row=self._DATA_START_ROW,
            column=self._NAME_COL + 1,
        )

    # ------------------------------------------------------------------

    def _collect_blocks(self, report_files: List[Optional[Path]]) -> List[Dict]:
        """Return phenotype block descriptors (numeric and categorical) in data order."""
        order: List[Tuple[str, str]] = []
        seen: set = set()
        stat_cols_available: set = set()
        categorical_phenos: Dict[str, Dict[str, None]] = {}

        for f in report_files:
            if f is None:
                continue
            try:
                data = self._load_json(f)
                for row in data.get("rows", []):
                    name = row.get("Name", "")
                    level = row.get("_level", 0)
                    if "=" in name and int(level) == 0:
                        pheno, cat = name.split("=", 1)
                        if pheno not in categorical_phenos:
                            categorical_phenos[pheno] = {}
                        if pheno not in seen:
                            order.append(("categorical", pheno))
                            seen.add(pheno)
                        if cat not in categorical_phenos[pheno]:
                            categorical_phenos[pheno][cat] = None
                    elif name and "=" not in name and self._has_numeric_mean(row):
                        if name not in seen:
                            order.append(("numeric", name))
                            seen.add(name)
                        for col in self._NUMERIC_COL_ORDER:
                            if row.get(col) is not None:
                                stat_cols_available.add(col)
            except Exception:
                pass

        available = [c for c in self._NUMERIC_COL_ORDER if c in stat_cols_available]

        blocks: List[Dict] = []
        for kind, name in order:
            if kind == "numeric":
                blocks.append({"kind": "numeric", "name": name, "stat_cols": available})
            else:
                blocks.append(
                    {
                        "kind": "categorical",
                        "name": name,
                        "categories": list(categorical_phenos[name].keys()),
                    }
                )
        return blocks

    def _write_numeric_block(
        self,
        sheet,
        start_col: int,
        block: Dict,
        cohort_entries: List[Dict],
    ) -> int:
        """Write headers and data for one numeric phenotype. Returns next free column."""
        name = block["name"]
        stat_cols = block["stat_cols"]
        num_cols = len(stat_cols)
        end_col = start_col + num_cols - 1

        # Row 2: phenotype name (merged)
        self._write_cell(
            sheet,
            self._TITLE_ROW,
            start_col,
            name,
            bold=True,
            size=16,
            horizontal="center",
        )
        if num_cols > 1:
            for c in range(start_col + 1, end_col + 1):
                sheet.cell(row=self._TITLE_ROW, column=c)
            sheet.merge_cells(
                start_row=self._TITLE_ROW,
                start_column=start_col,
                end_row=self._TITLE_ROW,
                end_column=end_col,
            )

        # Row 3: stat column labels
        for i, col_name in enumerate(stat_cols):
            display = self._COLUMN_DISPLAY_NAMES.get(col_name, col_name)
            grayed = col_name in self._GRAYED_STATS
            self._write_cell(
                sheet,
                self._SUBTITLE_ROW,
                start_col + i,
                display,
                bold=not grayed,
                size=14,
                horizontal="center",
                font_color=self._GRAY_TEXT if grayed else None,
            )

        # Data rows + track max content length for widths
        font_scale = 14 / 11
        col_max_lens = {c: len(self._COLUMN_DISPLAY_NAMES.get(c, c)) for c in stat_cols}

        for ri, entry in enumerate(cohort_entries):
            row = self._DATA_START_ROW + ri
            row_data = self._find_char_row(entry["data"], name)
            alt_fill = entry["group_color"] if ri % 2 == 1 else None
            td, tl = entry["text_dark"], entry["text_light"]
            for i, col_name in enumerate(stat_cols):
                value = (
                    self._clean_numeric(row_data.get(col_name)) if row_data else None
                )
                fmt = self._number_format_for_value(value)
                if col_name == "STD":
                    fc = self._GRAY_TEXT
                elif col_name == "N":
                    fc = tl
                elif col_name == "Pct":
                    fc = td
                else:
                    fc = td
                self._write_cell(
                    sheet,
                    row,
                    start_col + i,
                    value,
                    size=14,
                    horizontal="center",
                    fill_color=alt_fill,
                    number_format=fmt,
                    font_color=fc,
                )
                if value is not None:
                    col_max_lens[col_name] = max(
                        col_max_lens[col_name],
                        len(str(value)),
                    )

        # Column widths
        for i, col_name in enumerate(stat_cols):
            if col_name == "N":
                w = 14
            else:
                content_len = col_max_lens.get(col_name, 4)
                w = min(max(content_len * font_scale + 1, 6), 18)
            sheet.column_dimensions[get_column_letter(start_col + i)].width = w

        return end_col + 1

    def _write_categorical_block(
        self,
        sheet,
        start_col: int,
        block: Dict,
        cohort_entries: List[Dict],
    ) -> int:
        """Write headers and data for one categorical phenotype. Returns next free column."""
        name = block["name"]
        categories = block["categories"]
        num_cols = len(categories) * 2
        end_col = start_col + num_cols - 1

        # Row 2: phenotype name (merged)
        self._write_cell(
            sheet,
            self._TITLE_ROW,
            start_col,
            name,
            bold=True,
            size=16,
            horizontal="center",
        )
        if num_cols > 1:
            for c in range(start_col + 1, end_col + 1):
                sheet.cell(row=self._TITLE_ROW, column=c)
            sheet.merge_cells(
                start_row=self._TITLE_ROW,
                start_column=start_col,
                end_row=self._TITLE_ROW,
                end_column=end_col,
            )

        # Row 3: category names (merged over N/% pairs)
        for ci, cat_name in enumerate(categories):
            n_col = start_col + ci * 2
            pct_col = n_col + 1
            cell = self._write_cell(
                sheet,
                self._SUBTITLE_ROW,
                n_col,
                cat_name,
                bold=True,
                size=14,
                horizontal="center",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="bottom",
                wrap_text=True,
            )
            sheet.merge_cells(
                start_row=self._SUBTITLE_ROW,
                start_column=n_col,
                end_row=self._SUBTITLE_ROW,
                end_column=pct_col,
            )

        # Adjust row 3 height for wrapped category names
        font_scale = 14 / 11
        chars_per_col = int(14 / font_scale)
        max_lines = 1
        for cat in categories:
            lines = max(1, math.ceil(len(cat) / max(chars_per_col, 1)))
            max_lines = max(max_lines, lines)
        needed = max(24, max_lines * 20)
        current = sheet.row_dimensions[self._SUBTITLE_ROW].height or 24
        sheet.row_dimensions[self._SUBTITLE_ROW].height = max(current, needed)

        # Row 4: N/% sub-headers
        for ci in range(len(categories)):
            n_col = start_col + ci * 2
            pct_col = n_col + 1
            self._write_cell(
                sheet,
                self._HEADER_ROW,
                n_col,
                "N",
                size=14,
                horizontal="right",
                font_color=self._GRAY_TEXT,
            )
            self._write_cell(
                sheet,
                self._HEADER_ROW,
                pct_col,
                "%",
                bold=True,
                size=14,
                horizontal="left",
            )

        # Data rows
        for ri, entry in enumerate(cohort_entries):
            row = self._DATA_START_ROW + ri
            cat_data = self._get_category_data(entry["data"], name, categories)
            alt_fill = entry["group_color"] if ri % 2 == 1 else None
            td, tl = entry["text_dark"], entry["text_light"]
            for ci, cat in enumerate(categories):
                n_col = start_col + ci * 2
                pct_col = n_col + 1
                vals = cat_data.get(cat, {})
                n_val = self._clean_numeric(vals.get("N"))
                pct_val = self._clean_numeric(vals.get("Pct"))
                self._write_cell(
                    sheet,
                    row,
                    n_col,
                    n_val,
                    size=14,
                    horizontal="right",
                    fill_color=alt_fill,
                    number_format=self._number_format_for_value(n_val),
                    font_color=tl,
                )
                self._write_cell(
                    sheet,
                    row,
                    pct_col,
                    pct_val,
                    bold=True,
                    size=14,
                    horizontal="left",
                    fill_color=alt_fill,
                    number_format=self._number_format_for_value(pct_val),
                    font_color=td,
                )

        # Column widths
        for ci in range(len(categories)):
            n_col = start_col + ci * 2
            pct_col = n_col + 1
            sheet.column_dimensions[get_column_letter(n_col)].width = 14
            sheet.column_dimensions[get_column_letter(pct_col)].width = 6

        return end_col + 1

    # ------------------------------------------------------------------

    def _apply_block_border(
        self,
        sheet,
        top_row: int,
        bottom_row: int,
        start_col: int,
        end_col: int,
    ):
        """Draw a thin black border around a phenotype block."""
        side = Side(style="thin", color="000000")
        for row in range(top_row, bottom_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                existing = cell.border
                cell.border = Border(
                    left=side if col == start_col else existing.left,
                    right=side if col == end_col else existing.right,
                    top=side if row == top_row else existing.top,
                    bottom=side if row == bottom_row else existing.bottom,
                )

    @staticmethod
    def _has_numeric_mean(row: Dict) -> bool:
        """Return True if the row has a real (non-NaN) Mean value."""
        val = row.get("Mean")
        if val is None:
            return False
        if isinstance(val, float) and math.isnan(val):
            return False
        if isinstance(val, str) and val.lower() in ("nan", ""):
            return False
        return True

    @staticmethod
    def _find_char_row(data: Dict, char_name: str) -> Optional[Dict]:
        for row in data.get("rows", []):
            if row.get("Name") == char_name:
                return row
        return None

    @staticmethod
    def _get_category_data(
        data: Dict,
        phenotype_name: str,
        categories: List[str],
    ) -> Dict[str, Dict[str, object]]:
        result: Dict[str, Dict] = {}
        for row in data.get("rows", []):
            name = row.get("Name", "")
            if "=" in name:
                pheno, cat = name.split("=", 1)
                if pheno == phenotype_name and cat in categories:
                    result[cat] = {"N": row.get("N"), "Pct": row.get("Pct")}
        return result
