import json
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl.utils import get_column_letter

from phenex.util import create_logger
from .base_sheet_writer import _BaseSheetWriter
from .cohort_group import CohortGroup

logger = create_logger(__name__)


class GenericSheetWriter(_BaseSheetWriter):
    """Writes multiple JSON report files side-by-side into a single worksheet.

    For Waterfall reports the ``Type`` column drives row background colours.
    Columns are reordered and renamed according to ``_COLUMN_ORDER`` and
    ``_COLUMN_HEADERS``.  Columns not listed in the order list are appended
    at the end in their original order.
    """

    # Desired column order and display headers for Waterfall-type reports.
    _COLUMN_ORDER = [
        "Type",
        "Index",
        "Name",
        "Pct_Source_Database",
        "Pct_Remaining",
        "Remaining",
        "Delta",
        "Pct_N",
        "N",
    ]
    _COLUMN_HEADERS: Dict[str, str] = {
        "Type": "",
        "Index": "",
        "Name": "",
        "Pct_Source_Database": "% source",
        "Pct_Remaining": "%",
        "Remaining": "remaining",
        "Delta": "\u0394",
        "Pct_N": "%",
        "N": "N",
    }

    # Columns that should be sized for large numbers (2-digit millions).
    _WIDE_COLUMNS = {"N", "Remaining", "Delta"}

    def write(
        self,
        sheet,
        report_files: List[Optional[Path]],
        cohort_dirs: List[Path],
        cohort_groups: List[CohortGroup],
    ):
        self._set_default_row_height(sheet)
        raw_columns = self._get_raw_column_order(report_files)
        if not raw_columns:
            return

        columns = self._reorder_columns(raw_columns)
        dir_to_report = dict(zip(cohort_dirs, report_files))
        num_cols = len(columns)
        current_col = 1
        sheet.freeze_panes = sheet.cell(row=self._DATA_START_ROW, column=1)

        for gi, group in enumerate(cohort_groups):
            group_color = self._COHORT_COLORS[gi % len(self._COHORT_COLORS)]

            for cohort_dir in group.all_dirs:
                report_file = dir_to_report.get(cohort_dir)
                if report_file is None:
                    continue
                is_subcohort = cohort_dir != group.main_dir

                # Spacer before every cohort card
                self._apply_spacing(sheet, spacing_col=current_col)
                current_col += 1

                try:
                    data = self._load_json(report_file)
                    rows = data.get("rows", [])
                    display_name = group.display_name(cohort_dir)
                    title = (
                        f"{group.name} \u00b7 {display_name}"
                        if is_subcohort
                        else display_name
                    )

                    card_start = current_col
                    card_end = card_start + num_cols - 1

                    # Title row (row 2) – cohort color, merged
                    self._write_cell(
                        sheet,
                        self._TITLE_ROW,
                        card_start,
                        title,
                        bold=True,
                        italic=True,
                        size=16,
                        horizontal="left",
                        indent=1,
                        fill_color=group_color,
                    )
                    for c in range(card_start + 1, card_end + 1):
                        self._write_cell(
                            sheet, self._TITLE_ROW, c, None, fill_color=group_color
                        )
                    if card_end > card_start:
                        sheet.merge_cells(
                            start_row=self._TITLE_ROW,
                            start_column=card_start,
                            end_row=self._TITLE_ROW,
                            end_column=card_end,
                        )

                    self._write_column_headers(sheet, columns, card_start)
                    self._write_data_rows(sheet, rows, columns, card_start)
                    self._set_column_widths(sheet, rows, columns, card_start)
                    self._apply_group_border(
                        sheet, card_start, card_end, color=group_color
                    )

                    current_col = card_end + 1
                except Exception as e:
                    logger.warning(f"Failed to process {report_file}: {e}")

    # ------------------------------------------------------------------

    def _reorder_columns(self, raw_columns: List[str]) -> List[str]:
        """Return columns in the preferred order, appending any extras."""
        ordered = [c for c in self._COLUMN_ORDER if c in raw_columns]
        extras = [c for c in raw_columns if c not in self._COLUMN_ORDER]
        return ordered + extras

    @staticmethod
    def _get_raw_column_order(report_files: List[Optional[Path]]) -> List[str]:
        for f in report_files:
            if f is None:
                continue
            try:
                with f.open() as fh:
                    data = json.load(fh)
                rows = data.get("rows", [])
                if rows:
                    return list(rows[0].keys())
            except Exception:
                pass
        return []

    def _write_column_headers(self, sheet, columns: List[str], start_col: int):
        for offset, col_name in enumerate(columns):
            display = self._COLUMN_HEADERS.get(col_name, col_name)
            col = start_col + offset
            self._write_cell(
                sheet,
                self._HEADER_ROW,
                col,
                display,
                bold=True,
                size=14,
                horizontal="center",
            )

    def _write_data_rows(self, sheet, rows: list, columns: List[str], start_col: int):
        display_rows = self._sparsify_type(rows) if "Type" in columns else rows
        for row_idx, (orig_row, disp_row) in enumerate(
            zip(rows, display_rows),
            start=self._DATA_START_ROW,
        ):
            row_type = str(orig_row.get("Type", "")).lower()
            fill_color = (
                None if row_type == "info" else self._WATERFALL_COLORS.get(row_type)
            )
            for offset, col_name in enumerate(columns):
                value = self._clean_numeric(disp_row.get(col_name))
                fmt = self._number_format_for_value(value)
                self._write_cell(
                    sheet,
                    row_idx,
                    start_col + offset,
                    value,
                    size=14,
                    horizontal="center",
                    fill_color=fill_color,
                    number_format=fmt,
                )

    @staticmethod
    def _sparsify_type(rows: list) -> list:
        """Return a copy of rows with Type showing only on the first row of each group."""
        previous_type = None
        result = []
        for row in rows:
            row_copy = dict(row)
            type_val = str(row_copy.get("Type", ""))
            if (
                type_val
                and type_val != previous_type
                and type_val.lower() not in ("component", "info")
            ):
                previous_type = type_val
            else:
                row_copy["Type"] = ""
            result.append(row_copy)
        return result

    def _set_column_widths(self, sheet, rows: list, columns: List[str], start_col: int):
        font_scale = 14 / 11
        for offset, col_name in enumerate(columns):
            if col_name in self._WIDE_COLUMNS:
                width = 14  # fits "10,000,000"
            else:
                display = self._COLUMN_HEADERS.get(col_name, col_name)
                values = [
                    str(r.get(col_name, ""))
                    for r in rows
                    if r.get(col_name) is not None
                ]
                max_val_len = max((len(v) for v in values), default=0)
                content_len = max(len(display), max_val_len)
                width = min(max(content_len * font_scale + 2, 6), 40)
            sheet.column_dimensions[get_column_letter(start_col + offset)].width = width
